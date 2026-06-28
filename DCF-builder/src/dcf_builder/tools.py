"""Local artifact tools for the DCF Builder agent."""

from __future__ import annotations

import io
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from financial_agent_runtime import (
    artifact_exists,
    backend_is_daytona,
    contains_task_timestamp_dir,
    ensure_artifact_dir,
    materialize_file_artifact,
    read_bytes_artifact,
    read_text_artifact,
    write_text_artifact,
)
from langchain_core.tools import tool

from dcf_builder.config import file_storage_root


_TASK_OUTPUT_DIRS: dict[str, Path] = {}


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _workspace_root() -> Path:
    return file_storage_root()


def _resolve_output_dir(output_dir: str) -> Path:
    path = Path(output_dir)
    return path if path.is_absolute() else _workspace_root() / path


def _ensure_out_dir(output_dir: str) -> Path:
    out = _resolve_output_dir(output_dir)
    _ensure_dir(out)
    return out


def _ensure_dir(path: Path) -> None:
    ensure_artifact_dir(path)


def _write_text(path: Path, text: str) -> None:
    write_text_artifact(path, text, encoding="utf-8")


def _save_workbook(wb: Any, path: Path) -> None:
    materialize_file_artifact(path, lambda target: wb.save(target))


def _load_workbook(path: Path, **kwargs: Any):
    import openpyxl

    if backend_is_daytona():
        return openpyxl.load_workbook(io.BytesIO(read_bytes_artifact(path)), **kwargs)
    return openpyxl.load_workbook(path, **kwargs)


def _timestamped_output_dir(output_dir: str, exact_output_dir: bool = False) -> Path:
    base = _resolve_output_dir(output_dir)
    if exact_output_dir:
        _ensure_dir(base)
        return base

    if contains_task_timestamp_dir(base):
        _ensure_dir(base)
        return base

    key = str(base.resolve())
    existing = _TASK_OUTPUT_DIRS.get(key)
    if existing:
        _ensure_dir(existing)
        return existing

    timestamp = os.getenv("DCF_BUILDER_OUTPUT_TIMESTAMP")
    if not timestamp:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

    candidate = base / timestamp
    if artifact_exists(candidate) and os.getenv("DCF_BUILDER_OUTPUT_TIMESTAMP") is None:
        suffix = 2
        while artifact_exists(base / f"{timestamp}-{suffix}"):
            suffix += 1
        candidate = base / f"{timestamp}-{suffix}"

    _ensure_dir(candidate)
    _TASK_OUTPUT_DIRS[key] = candidate
    return candidate


def _slugify(text: str, fallback: str = "company") -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", str(text).strip()).strip("-").lower()
    return slug or fallback


def _parse_json(value: str, expected: str) -> Any:
    try:
        return json.loads(value)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{expected} must be valid JSON: {exc}") from exc


def _as_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").replace("%", "").strip()
    if cleaned in {"", "[UNSOURCED]", "UNSOURCED"}:
        return default
    try:
        return float(cleaned)
    except ValueError:
        return default


def _as_decimal(value: Any, default: float = 0.0) -> float:
    val = _as_float(value, default)
    if abs(val) > 1.5:
        return val / 100.0
    return val


def _as_series(value: Any, periods: int, default: float, decimal: bool = True) -> list[float]:
    converter = _as_decimal if decimal else _as_float
    if isinstance(value, list):
        values = [converter(item, default) for item in value]
    else:
        values = [converter(value, default)]
    if not values:
        values = [default]
    while len(values) < periods:
        values.append(values[-1])
    return values[:periods]


def _scenario(scenarios: dict[str, Any], name: str) -> dict[str, Any]:
    for key, value in scenarios.items():
        if key.lower() == name.lower() and isinstance(value, dict):
            return value
    return {}


def _field(data: dict[str, Any], *names: str, default: Any = None) -> Any:
    for name in names:
        if name in data and data[name] not in (None, ""):
            return data[name]
    return default


def _is_missing_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        text = value.strip()
        return text == "" or text.upper() in {"[UNSOURCED]", "UNSOURCED"}
    return False


def _parse_number(value: Any) -> float | None:
    if _is_missing_value(value) or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").replace("%", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_decimal(value: Any) -> float | None:
    parsed = _parse_number(value)
    if parsed is None:
        return None
    if abs(parsed) > 1.5:
        return parsed / 100.0
    return parsed


def _require_source(errors: list[str], value: Any, path: str) -> None:
    if _is_missing_value(value):
        errors.append(f"{path}: source is required")


def _require_text(errors: list[str], value: Any, path: str) -> None:
    if _is_missing_value(value):
        errors.append(f"{path}: value is required")


def _require_number(
    errors: list[str],
    value: Any,
    path: str,
    *,
    positive: bool = False,
    nonnegative: bool = False,
    decimal: bool = False,
) -> float | None:
    parsed = _parse_decimal(value) if decimal else _parse_number(value)
    if parsed is None:
        errors.append(f"{path}: valid numeric value is required")
        return None
    if positive and parsed <= 0:
        errors.append(f"{path}: value must be greater than 0")
    if nonnegative and parsed < 0:
        errors.append(f"{path}: value must be greater than or equal to 0")
    return parsed


def _require_ratio(
    errors: list[str],
    value: Any,
    path: str,
    *,
    allow_negative: bool = False,
) -> float | None:
    parsed = _require_number(errors, value, path, decimal=True)
    if parsed is None:
        return None
    lower_bound = -1.0 if allow_negative else 0.0
    if parsed < lower_bound or parsed > 1.0:
        errors.append(f"{path}: decimal ratio must be between {lower_bound:g} and 1")
    return parsed


def _require_series(
    errors: list[str],
    value: Any,
    path: str,
    periods: int,
    *,
    allow_negative: bool = False,
) -> list[float]:
    if not isinstance(value, list):
        parsed = _require_ratio(errors, value, path, allow_negative=allow_negative)
        return [parsed] * periods if parsed is not None else []

    if len(value) != periods:
        errors.append(f"{path}: expected {periods} values, got {len(value)}")

    values: list[float] = []
    for idx, item in enumerate(value):
        item_path = f"{path}[{idx}]"
        parsed = _require_ratio(errors, item, item_path, allow_negative=allow_negative)
        if parsed is not None:
            values.append(parsed)
    return values


_MARKET_DATA_ALIASES: dict[str, tuple[str, ...]] = {
    "current_stock_price": (
        "current_price",
        "stock_price",
        "latest_price",
        "last_price",
        "price",
        "close",
        "closing_price",
        "最新价",
        "现价",
        "当前价",
        "收盘价",
        "最新收盘价",
        "股票价格",
    ),
    "debt": (
        "total_debt",
        "interest_bearing_debt",
        "interest_bearing_liabilities",
        "borrowings",
        "有息负债",
        "带息债务",
        "总债务",
        "总负债",
    ),
    "cash": (
        "cash_and_equivalents",
        "cash_equivalents",
        "cash_and_cash_equivalents",
        "monetary_funds",
        "cash_balance",
        "货币资金",
        "现金及现金等价物",
        "现金等价物",
        "现金",
    ),
    "shares_outstanding": (
        "shares",
        "total_shares",
        "share_outstanding",
        "outstanding_shares",
        "share_capital",
        "total_share_capital",
        "total_capital_stock",
        "diluted_shares",
        "总股本",
        "总股份",
        "股本",
        "股份总数",
    ),
    "beta": (
        "Beta",
        "beta_coefficient",
        "beta_value",
        "贝塔",
        "贝塔系数",
    ),
    "risk_free_rate": (
        "riskfree_rate",
        "rfr",
        "treasury_yield",
        "government_bond_yield",
        "10y_treasury_yield",
        "10y_government_bond_yield",
        "无风险利率",
        "国债收益率",
        "十年期国债收益率",
        "10年期国债收益率",
    ),
    "equity_risk_premium": (
        "market_risk_premium",
        "erp",
        "risk_premium",
        "股权风险溢价",
        "权益风险溢价",
        "市场风险溢价",
    ),
    "pretax_cost_of_debt": (
        "pre_tax_cost_of_debt",
        "cost_of_debt",
        "debt_cost",
        "borrowing_rate",
        "loan_rate",
        "bond_yield",
        "税前债务成本",
        "债务成本",
        "借款利率",
        "融资成本",
    ),
    "tax_rate": (
        "effective_tax_rate",
        "income_tax_rate",
        "statutory_tax_rate",
        "corporate_tax_rate",
        "actual_tax_rate",
        "实际所得税率",
        "所得税率",
        "企业所得税率",
        "实际税率",
        "税率",
    ),
    "source": (
        "sources",
        "data_source",
        "source_text",
        "数据来源",
        "来源",
    ),
}


def _normalize_market_key(key: Any) -> str:
    return re.sub(r"[\s_\-./()（）:：,，%]+", "", str(key).casefold())


def _source_text(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(str(item) for item in value if item not in (None, ""))
    return str(value) if value not in (None, "") else ""


def _normalize_market_data(market_data: Any) -> dict[str, Any]:
    if not isinstance(market_data, dict):
        return {}

    normalized = dict(market_data)
    values_by_key: dict[str, Any] = {}
    for key, value in market_data.items():
        normalized_key = _normalize_market_key(key)
        if normalized_key and normalized_key not in values_by_key and value not in (None, ""):
            values_by_key[normalized_key] = value

    for canonical, aliases in _MARKET_DATA_ALIASES.items():
        if canonical in normalized and normalized[canonical] not in (None, ""):
            continue
        for alias in (canonical, *aliases):
            value = values_by_key.get(_normalize_market_key(alias))
            if value not in (None, ""):
                normalized[canonical] = _source_text(value) if canonical == "source" else value
                break

    return normalized


def _peer_field(data: dict[str, Any], *names: str, default: Any = None) -> Any:
    return _field(data, *names, default=default)


def _validate_comps_inputs(companies: list[Any]) -> None:
    errors: list[str] = []
    for idx, co in enumerate(companies):
        path = f"data_json.companies[{idx}]"
        if not isinstance(co, dict):
            errors.append(f"{path}: object is required")
            continue

        _require_text(
            errors,
            _peer_field(co, "company", "company_name", "name", "short_name", "证券简称", "公司名称"),
            f"{path}.company",
        )
        _require_text(
            errors,
            _peer_field(co, "ticker", "symbol", "stock_code", "证券代码", "股票代码"),
            f"{path}.ticker",
        )
        _require_source(
            errors,
            _peer_field(co, "source", "sources", "data_source", "数据来源", "来源"),
            f"{path}.source",
        )
        _require_number(
            errors,
            _peer_field(co, "revenue_ltm", "revenue", "营业收入"),
            f"{path}.revenue",
            positive=True,
        )
        _require_ratio(
            errors,
            _peer_field(co, "revenue_growth_pct", "revenue_growth", "收入增速"),
            f"{path}.revenue_growth",
            allow_negative=True,
        )
        _require_number(
            errors,
            _peer_field(co, "ebitda_ltm", "ebitda", "EBITDA"),
            f"{path}.ebitda",
        )
        _require_number(
            errors,
            _peer_field(co, "net_income", "净利润"),
            f"{path}.net_income",
        )
        _require_number(
            errors,
            _peer_field(co, "market_cap", "market_value", "市值", "总市值"),
            f"{path}.market_cap",
            positive=True,
        )
        _require_number(
            errors,
            _peer_field(co, "enterprise_value", "ev", "EV", "企业价值"),
            f"{path}.enterprise_value",
            positive=True,
        )

    if errors:
        raise ValueError("comps data validation failed:\n- " + "\n- ".join(errors))


def _validate_projection_periods(payload: dict[str, Any]) -> int:
    raw_periods = payload.get("projection_periods", payload.get("projection_years", 5))
    parsed = _parse_number(raw_periods)
    if parsed is None or int(parsed) != parsed:
        raise ValueError(
            "dcf_json validation failed:\n"
            "- dcf_json.projection_periods: integer value is required"
        )
    return max(3, min(int(parsed), 7))


def _validate_dcf_inputs(
    payload: dict[str, Any],
    historicals: list[Any],
    market_data: dict[str, Any],
    scenarios: dict[str, Any],
    periods: int,
) -> None:
    errors: list[str] = []

    _require_text(errors, payload.get("company") or payload.get("company_name"), "dcf_json.company")
    _require_text(errors, payload.get("ticker"), "dcf_json.ticker")

    market_path = "dcf_json.market_data"
    _require_source(errors, market_data.get("source"), f"{market_path}.source")
    _require_number(errors, market_data.get("current_stock_price"), f"{market_path}.current_stock_price", positive=True)
    _require_number(errors, market_data.get("debt"), f"{market_path}.debt", nonnegative=True)
    _require_number(errors, market_data.get("cash"), f"{market_path}.cash", nonnegative=True)
    _require_number(errors, market_data.get("shares_outstanding"), f"{market_path}.shares_outstanding", positive=True)
    _require_number(errors, market_data.get("beta"), f"{market_path}.beta", positive=True)
    _require_ratio(errors, market_data.get("risk_free_rate"), f"{market_path}.risk_free_rate", allow_negative=True)
    _require_ratio(errors, market_data.get("equity_risk_premium"), f"{market_path}.equity_risk_premium")
    _require_ratio(errors, market_data.get("pretax_cost_of_debt"), f"{market_path}.pretax_cost_of_debt")
    _require_ratio(errors, market_data.get("tax_rate"), f"{market_path}.tax_rate")

    for idx, hist in enumerate(historicals):
        hist_path = f"dcf_json.historicals[{idx}]"
        if not isinstance(hist, dict):
            errors.append(f"{hist_path}: object is required")
            continue

        _require_source(errors, hist.get("source"), f"{hist_path}.source")
        if _year_int(hist.get("year"), 0) == 0:
            errors.append(f"{hist_path}.year: four-digit year is required")
        _require_number(errors, hist.get("revenue"), f"{hist_path}.revenue", positive=True)

        ebit = hist.get("ebit")
        ebit_margin = hist.get("ebit_margin")
        if _is_missing_value(ebit) and _is_missing_value(ebit_margin):
            errors.append(f"{hist_path}.ebit: ebit or ebit_margin is required")
        elif not _is_missing_value(ebit):
            _require_number(errors, ebit, f"{hist_path}.ebit")
        else:
            _require_ratio(errors, ebit_margin, f"{hist_path}.ebit_margin", allow_negative=True)

        _require_number(errors, _field(hist, "da", "d_and_a", "depreciation_amortization"), f"{hist_path}.da", nonnegative=True)
        _require_number(errors, hist.get("capex"), f"{hist_path}.capex")
        _require_number(errors, _field(hist, "nwc_change", "change_in_nwc"), f"{hist_path}.nwc_change")
        _require_number(errors, hist.get("debt"), f"{hist_path}.debt", nonnegative=True)
        _require_number(errors, hist.get("cash"), f"{hist_path}.cash", nonnegative=True)
        _require_number(errors, _field(hist, "shares_outstanding", "shares"), f"{hist_path}.shares_outstanding", positive=True)

    for scenario_name in ("Bear", "Base", "Bull"):
        scenario = _scenario(scenarios, scenario_name)
        scenario_path = f"dcf_json.scenarios.{scenario_name}"
        if not scenario:
            errors.append(f"{scenario_path}: scenario object is required")
            continue

        _require_source(errors, scenario.get("source"), f"{scenario_path}.source")
        revenue_growth = _field(scenario, "revenue_growth", "revenue_growth_pct")
        ebit_margin = _field(scenario, "ebit_margin", "operating_margin")
        tax_rate = scenario.get("tax_rate")
        da_pct = _field(scenario, "da_pct_revenue", "d_and_a_pct_revenue")
        capex_pct = scenario.get("capex_pct_revenue")
        nwc_pct = scenario.get("nwc_pct_delta_revenue")
        wacc = scenario.get("wacc")
        terminal_growth = _field(scenario, "terminal_growth", "terminal_growth_rate")

        _require_series(errors, revenue_growth, f"{scenario_path}.revenue_growth", periods, allow_negative=True)
        _require_series(errors, ebit_margin, f"{scenario_path}.ebit_margin", periods, allow_negative=True)
        _require_series(errors, tax_rate, f"{scenario_path}.tax_rate", periods)
        _require_series(errors, da_pct, f"{scenario_path}.da_pct_revenue", periods)
        _require_series(errors, capex_pct, f"{scenario_path}.capex_pct_revenue", periods)
        _require_series(errors, nwc_pct, f"{scenario_path}.nwc_pct_delta_revenue", periods, allow_negative=True)
        wacc_values = _require_series(errors, wacc, f"{scenario_path}.wacc", periods)
        terminal_growth_values = _require_series(errors, terminal_growth, f"{scenario_path}.terminal_growth", periods)
        for idx, (wacc_value, terminal_value) in enumerate(zip(wacc_values, terminal_growth_values)):
            if terminal_value >= wacc_value:
                errors.append(
                    f"{scenario_path}.terminal_growth[{idx}]: must be lower than "
                    f"{scenario_path}.wacc[{idx}]"
                )

    if errors:
        raise ValueError("dcf_json validation failed:\n- " + "\n- ".join(errors))


def _col_letter(col_index: int) -> str:
    letters = ""
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _comment(source: str):
    from openpyxl.comments import Comment

    text = f"Source: {source or '[UNSOURCED]'}"
    comment = Comment(text, "dcf-builder")
    comment.width = 240
    comment.height = 70
    return comment


def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    return {
        "navy_fill": PatternFill("solid", fgColor="1F4E79"),
        "blue_fill": PatternFill("solid", fgColor="D9E1F2"),
        "gray_fill": PatternFill("solid", fgColor="F2F2F2"),
        "input_fill": PatternFill("solid", fgColor="EAF3F8"),
        "base_fill": PatternFill("solid", fgColor="BDD7EE"),
        "white_font": Font(name="Times New Roman", bold=True, color="FFFFFF", size=11),
        "bold_font": Font(name="Times New Roman", bold=True, color="000000", size=11),
        "normal_font": Font(name="Times New Roman", color="000000", size=11),
        "input_font": Font(name="Times New Roman", color="0000FF", size=11),
        "formula_font": Font(name="Times New Roman", color="000000", size=11),
        "link_font": Font(name="Times New Roman", color="008000", size=11),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "thin_border": Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        ),
    }


def _section_header(ws, row: int, title: str, end_col: int, styles: dict[str, Any]) -> None:
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).fill = styles["navy_fill"]
    ws.cell(row=row, column=1).font = styles["white_font"]
    ws.cell(row=row, column=1).alignment = styles["left"]
    for col in range(2, end_col + 1):
        ws.cell(row=row, column=col).fill = styles["navy_fill"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)


def _input_cell(cell, value: Any, source: str, fmt: str = "") -> None:
    styles = _styles()
    cell.value = value
    cell.font = styles["input_font"]
    cell.fill = styles["input_fill"]
    cell.alignment = styles["center"]
    if fmt:
        cell.number_format = fmt
    cell.comment = _comment(source)


def _formula_cell(cell, formula: str, fmt: str = "", link: bool = False) -> None:
    styles = _styles()
    cell.value = formula
    cell.font = styles["link_font"] if link else styles["formula_font"]
    cell.alignment = styles["center"]
    if fmt:
        cell.number_format = fmt


@tool
def build_comps_excel(
    data_json: str,
    sector: str,
    output_dir: str = "./out",
    exact_output_dir: bool = False,
) -> str:
    """Build a comparable company analysis workbook.

    Args:
        data_json: JSON list or {"companies": [...]} with raw peer data.
        sector: Sector or peer-set label used in the file name.
        output_dir: Base output directory. A task timestamp subdirectory is reused
            unless exact_output_dir is true.
        exact_output_dir: When true, write directly into output_dir and use
            the canonical filename comps.xlsx.

    Returns:
        Path to the written workbook.
    """
    try:
        import openpyxl
    except ImportError as exc:
        return f"ERROR: openpyxl is not installed - {exc}"

    payload = _parse_json(data_json, "data_json")
    companies = payload.get("companies", payload) if isinstance(payload, dict) else payload
    if not isinstance(companies, list) or not companies:
        raise ValueError("data_json must contain a non-empty company list")
    _validate_comps_inputs(companies)

    out_dir = _timestamped_output_dir(output_dir, exact_output_dir)
    filename = (
        "comps.xlsx"
        if exact_output_dir
        else f"comps-{_slugify(sector, 'sector')}-{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    filepath = out_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comps"
    styles = _styles()

    ws["A1"] = f"{sector.upper()} - COMPARABLE COMPANY ANALYSIS"
    ws["A1"].fill = styles["navy_fill"]
    ws["A1"].font = styles["white_font"]
    ws["A2"] = " | ".join(
        f"{co.get('company', co.get('ticker', ''))} ({co.get('ticker', '')})"
        for co in companies
    )
    ws["A3"] = f"As of {datetime.now().strftime('%Y-%m-%d')} | Values in stated units except percentages and multiples"
    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")
    ws.merge_cells("A3:L3")

    headers = [
        "Company",
        "Ticker",
        "Revenue",
        "Revenue Growth",
        "EBITDA",
        "EBITDA Margin",
        "Net Income",
        "Market Cap",
        "Enterprise Value",
        "EV/Revenue",
        "EV/EBITDA",
        "P/E",
    ]
    _section_header(ws, 5, "OPERATING STATISTICS & VALUATION MULTIPLES", len(headers), styles)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.fill = styles["blue_fill"]
        cell.font = styles["bold_font"]
        cell.alignment = styles["center"]

    for idx, co in enumerate(companies, start=7):
        source = _source_text(_peer_field(co, "source", "sources", "data_source", "数据来源", "来源"))
        ws.cell(idx, 1, _peer_field(co, "company", "company_name", "name", "short_name", "证券简称", "公司名称"))
        ws.cell(idx, 2, _peer_field(co, "ticker", "symbol", "stock_code", "证券代码", "股票代码"))
        _input_cell(ws.cell(idx, 3), _as_float(_peer_field(co, "revenue_ltm", "revenue", "营业收入")), source, "#,##0")
        _input_cell(ws.cell(idx, 4), _as_decimal(_peer_field(co, "revenue_growth_pct", "revenue_growth", "收入增速")), source, "0.0%")
        _input_cell(ws.cell(idx, 5), _as_float(_peer_field(co, "ebitda_ltm", "ebitda", "EBITDA")), source, "#,##0")
        _formula_cell(ws.cell(idx, 6), f"=IF(C{idx}>0,E{idx}/C{idx},\"\")", "0.0%")
        _input_cell(ws.cell(idx, 7), _as_float(_peer_field(co, "net_income", "净利润")), source, "#,##0")
        _input_cell(ws.cell(idx, 8), _as_float(_peer_field(co, "market_cap", "market_value", "市值", "总市值")), source, "#,##0")
        _input_cell(ws.cell(idx, 9), _as_float(_peer_field(co, "enterprise_value", "ev", "EV", "企业价值")), source, "#,##0")
        _formula_cell(ws.cell(idx, 10), f"=IF(C{idx}>0,I{idx}/C{idx},\"\")", '0.0"x"')
        _formula_cell(ws.cell(idx, 11), f"=IF(E{idx}>0,I{idx}/E{idx},\"\")", '0.0"x"')
        _formula_cell(ws.cell(idx, 12), f"=IF(G{idx}>0,H{idx}/G{idx},\"\")", '0.0"x"')

    data_start = 7
    data_end = 6 + len(companies)
    stats_start = data_end + 2
    stats = [
        ("Maximum", "MAX"),
        ("75th Percentile", "QUARTILE.INC({range},3)"),
        ("Median", "MEDIAN"),
        ("25th Percentile", "QUARTILE.INC({range},1)"),
        ("Minimum", "MIN"),
    ]
    for row_offset, (label, template) in enumerate(stats):
        row = stats_start + row_offset
        ws.cell(row, 1, label).font = styles["bold_font"]
        ws.cell(row, 1).fill = styles["gray_fill"]
        for col in [4, 6, 10, 11, 12]:
            range_ref = f"{_col_letter(col)}{data_start}:{_col_letter(col)}{data_end}"
            formula = f"={template}({range_ref})" if "{range}" not in template else "=" + template.replace("{range}", range_ref)
            _formula_cell(ws.cell(row, col), formula)
            ws.cell(row, col).fill = styles["gray_fill"]
            ws.cell(row, col).number_format = ws.cell(data_start, col).number_format

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Notes & Methodology"
    notes["A1"].font = styles["bold_font"]
    notes["A3"] = "Sources"
    notes["A3"].font = styles["bold_font"]
    sources = sorted({co.get("source", "") for co in companies if co.get("source")})
    for row, source in enumerate(sources or ["[UNSOURCED]"], start=4):
        notes.cell(row, 1, source)
    notes["A10"] = "Derived metrics are formulas. Raw inputs carry source comments."

    for sheet in [ws, notes]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = styles["thin_border"]
        for col in range(1, 13):
            sheet.column_dimensions[_col_letter(col)].width = 16
    ws.column_dimensions["A"].width = 28

    _save_workbook(wb, filepath)
    return str(filepath)


@tool
def build_dcf_model(
    dcf_json: str,
    output_dir: str = "./out",
    exact_output_dir: bool = False,
) -> str:
    """Build a deterministic DCF workbook from structured JSON.

    Args:
        dcf_json: JSON object containing company, historicals, market_data,
            scenarios, and optional comps_summary.
        output_dir: Base output directory. A task timestamp subdirectory is reused
            unless exact_output_dir is true.
        exact_output_dir: When true, write directly into output_dir and use
            the canonical filename dcf_model.xlsx.

    Returns:
        Path to the written workbook.
    """
    try:
        import openpyxl
    except ImportError as exc:
        return f"ERROR: openpyxl is not installed - {exc}"

    payload = _parse_json(dcf_json, "dcf_json")
    if not isinstance(payload, dict):
        raise ValueError("dcf_json must be a JSON object")

    historicals = payload.get("historicals") or payload.get("historical_financials") or []
    if not isinstance(historicals, list) or not historicals:
        raise ValueError("dcf_json must include non-empty historicals")

    scenarios = payload.get("scenarios") or {}
    if not isinstance(scenarios, dict):
        raise ValueError("dcf_json.scenarios must be an object")

    periods = _validate_projection_periods(payload)
    company = payload.get("company") or payload.get("company_name") or "Company"
    ticker = payload.get("ticker") or "TICKER"
    currency = payload.get("currency") or "CNY"
    unit = payload.get("unit") or "millions"
    fiscal_year_end = payload.get("fiscal_year_end") or "Dec"
    market_data = _normalize_market_data(payload.get("market_data") or {})
    _validate_dcf_inputs(payload, historicals, market_data, scenarios, periods)
    latest_hist = historicals[-1]
    latest_year = _year_int(latest_hist.get("year"), datetime.now().year)
    projection_years = payload.get("projection_years_list") or [latest_year + i for i in range(1, periods + 1)]

    case_data = {
        "Bear": _normalize_case(_scenario(scenarios, "Bear"), periods, "Bear"),
        "Base": _normalize_case(_scenario(scenarios, "Base"), periods, "Base"),
        "Bull": _normalize_case(_scenario(scenarios, "Bull"), periods, "Bull"),
    }

    out_dir = _timestamped_output_dir(output_dir, exact_output_dir)
    filename = (
        "dcf_model.xlsx"
        if exact_output_dir
        else f"{_slugify(ticker, 'ticker')}_DCF_Model_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    filepath = out_dir / filename

    wb = openpyxl.Workbook()
    styles = _styles()
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_dcf = wb.create_sheet("DCF")
    ws_wacc = wb.create_sheet("WACC")
    ws_checks = wb.create_sheet("Checks")

    _write_inputs_sheet(ws_inputs, historicals, market_data, company, ticker, currency, unit, fiscal_year_end, styles)
    row_map = _write_dcf_sheet(ws_dcf, historicals, market_data, case_data, company, ticker, currency, unit, fiscal_year_end, projection_years, styles)
    _write_wacc_sheet(ws_wacc, row_map, styles)
    _write_checks_sheet(ws_checks, row_map, styles)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = styles["thin_border"]
                if cell.value is not None and cell.alignment is None:
                    cell.alignment = styles["center"]
        for col in range(1, 14):
            sheet.column_dimensions[_col_letter(col)].width = 16
        sheet.column_dimensions["A"].width = 28
        sheet.freeze_panes = "B2"

    _save_workbook(wb, filepath)
    return str(filepath)


def _year_int(value: Any, default: int) -> int:
    match = re.search(r"(19|20)\d{2}", str(value or ""))
    return int(match.group(0)) if match else default


def _normalize_case(case: dict[str, Any], periods: int, name: str) -> dict[str, Any]:
    defaults = {
        "Bear": {
            "revenue_growth": 0.02,
            "ebit_margin": 0.08,
            "tax_rate": 0.25,
            "da_pct_revenue": 0.03,
            "capex_pct_revenue": 0.04,
            "nwc_pct_delta_revenue": 0.08,
            "wacc": 0.105,
            "terminal_growth": 0.015,
        },
        "Base": {
            "revenue_growth": 0.05,
            "ebit_margin": 0.12,
            "tax_rate": 0.25,
            "da_pct_revenue": 0.03,
            "capex_pct_revenue": 0.035,
            "nwc_pct_delta_revenue": 0.06,
            "wacc": 0.095,
            "terminal_growth": 0.025,
        },
        "Bull": {
            "revenue_growth": 0.08,
            "ebit_margin": 0.16,
            "tax_rate": 0.25,
            "da_pct_revenue": 0.03,
            "capex_pct_revenue": 0.03,
            "nwc_pct_delta_revenue": 0.05,
            "wacc": 0.085,
            "terminal_growth": 0.03,
        },
    }[name]
    source = case.get("source") or "[ASSUMPTION]"
    return {
        "source": source,
        "revenue_growth": _as_series(_field(case, "revenue_growth", "revenue_growth_pct"), periods, defaults["revenue_growth"]),
        "ebit_margin": _as_series(_field(case, "ebit_margin", "operating_margin"), periods, defaults["ebit_margin"]),
        "tax_rate": _as_series(case.get("tax_rate"), periods, defaults["tax_rate"]),
        "da_pct_revenue": _as_series(_field(case, "da_pct_revenue", "d_and_a_pct_revenue"), periods, defaults["da_pct_revenue"]),
        "capex_pct_revenue": _as_series(case.get("capex_pct_revenue"), periods, defaults["capex_pct_revenue"]),
        "nwc_pct_delta_revenue": _as_series(case.get("nwc_pct_delta_revenue"), periods, defaults["nwc_pct_delta_revenue"]),
        "wacc": _as_series(case.get("wacc"), periods, defaults["wacc"]),
        "terminal_growth": _as_series(_field(case, "terminal_growth", "terminal_growth_rate"), periods, defaults["terminal_growth"]),
    }


def _write_inputs_sheet(ws, historicals, market_data, company, ticker, currency, unit, fiscal_year_end, styles) -> None:
    ws["A1"] = "DCF INPUTS"
    ws["A1"].fill = styles["navy_fill"]
    ws["A1"].font = styles["white_font"]
    ws.merge_cells("A1:J1")

    meta = [
        ("Company", company),
        ("Ticker", ticker),
        ("Currency", currency),
        ("Unit", unit),
        ("Fiscal Year End", fiscal_year_end),
    ]
    for row, (label, value) in enumerate(meta, start=3):
        ws.cell(row, 1, label).font = styles["bold_font"]
        _input_cell(ws.cell(row, 2), value, "User or iFind identity metadata")

    _section_header(ws, 10, "MARKET DATA", 4, styles)
    market_rows = [
        ("Current Stock Price", "current_stock_price", "0.00"),
        ("Debt", "debt", "#,##0"),
        ("Cash", "cash", "#,##0"),
        ("Shares Outstanding", "shares_outstanding", "#,##0.0"),
        ("Beta", "beta", "0.00"),
        ("Risk-Free Rate", "risk_free_rate", "0.0%"),
        ("Equity Risk Premium", "equity_risk_premium", "0.0%"),
        ("Pre-Tax Cost of Debt", "pretax_cost_of_debt", "0.0%"),
        ("Tax Rate", "tax_rate", "0.0%"),
    ]
    source = market_data.get("source", "")
    rate_keys = {"risk_free_rate", "equity_risk_premium", "pretax_cost_of_debt", "tax_rate"}
    for row, (label, key, fmt) in enumerate(market_rows, start=11):
        ws.cell(row, 1, label).font = styles["bold_font"]
        val = _as_decimal(market_data.get(key)) if key in rate_keys else _as_float(market_data.get(key))
        if key == "beta":
            val = _as_float(market_data.get(key), 1.0)
        _input_cell(ws.cell(row, 2), val, source, fmt)

    _section_header(ws, 23, "HISTORICAL FINANCIALS", 10, styles)
    headers = ["Year", "Revenue", "EBIT", "D&A", "CapEx", "NWC Change", "Debt", "Cash", "Shares", "Source"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(24, col, header)
        cell.fill = styles["blue_fill"]
        cell.font = styles["bold_font"]
        cell.alignment = styles["center"]

    for idx, hist in enumerate(historicals, start=25):
        source = hist.get("source", "")
        _input_cell(ws.cell(idx, 1), hist.get("year", ""), source)
        _input_cell(ws.cell(idx, 2), _as_float(hist.get("revenue")), source, "#,##0")
        ebit = hist.get("ebit")
        if ebit in (None, "") and hist.get("ebit_margin") not in (None, ""):
            ebit = _as_float(hist.get("revenue")) * _as_decimal(hist.get("ebit_margin"))
        _input_cell(ws.cell(idx, 3), _as_float(ebit), source, "#,##0")
        _input_cell(ws.cell(idx, 4), _as_float(_field(hist, "da", "d_and_a", "depreciation_amortization")), source, "#,##0")
        _input_cell(ws.cell(idx, 5), abs(_as_float(hist.get("capex"))), source, "#,##0")
        _input_cell(ws.cell(idx, 6), _as_float(_field(hist, "nwc_change", "change_in_nwc")), source, "#,##0")
        _input_cell(ws.cell(idx, 7), _as_float(hist.get("debt", market_data.get("debt"))), source, "#,##0")
        _input_cell(ws.cell(idx, 8), _as_float(hist.get("cash", market_data.get("cash"))), source, "#,##0")
        _input_cell(ws.cell(idx, 9), _as_float(hist.get("shares_outstanding", hist.get("shares", market_data.get("shares_outstanding")))), source, "#,##0.0")
        ws.cell(idx, 10, source or "[UNSOURCED]")


def _write_dcf_sheet(ws, historicals, market_data, case_data, company, ticker, currency, unit, fiscal_year_end, projection_years, styles) -> dict[str, int]:
    periods = len(projection_years)
    latest_row = 25 + len(historicals) - 1
    latest_year = _year_int(historicals[-1].get("year"), datetime.now().year)

    ws["A1"] = f"{company} DCF Model"
    ws["A1"].fill = styles["navy_fill"]
    ws["A1"].font = styles["white_font"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=periods + 2)
    ws["A2"] = f"Ticker: {ticker} | Date: {datetime.now().strftime('%Y-%m-%d')} | FYE: {fiscal_year_end} | Currency: {currency} {unit}"

    ws["A4"] = "Case Selector (1=Bear, 2=Base, 3=Bull)"
    _input_cell(ws["B4"], 2, "User case selector")
    ws["A5"] = "Selected Case"
    _formula_cell(ws["B5"], '=IF($B$4=1,"Bear",IF($B$4=2,"Base","Bull"))')

    _section_header(ws, 7, "MARKET DATA", 4, styles)
    market_labels = [
        ("Current Stock Price", "=Inputs!$B$11", "0.00"),
        ("Shares Outstanding", "=Inputs!$B$14", "#,##0.0"),
        ("Market Cap", "=B8*B9", "#,##0"),
        ("Debt", "=Inputs!$B$12", "#,##0"),
        ("Cash", "=Inputs!$B$13", "#,##0"),
        ("Net Debt", "=B11-B12", "#,##0"),
        ("Beta", "=Inputs!$B$15", "0.00"),
    ]
    for row, (label, formula, fmt) in enumerate(market_labels, start=8):
        ws.cell(row, 1, label).font = styles["bold_font"]
        _formula_cell(ws.cell(row, 2), formula, fmt, link=formula.startswith("=") and "Inputs!" in formula)

    scen_rows: dict[str, int] = {}
    _section_header(ws, 17, "SCENARIO ASSUMPTIONS", periods + 1, styles)
    scenario_start = {"Bear": 19, "Base": 30, "Bull": 41}
    assumption_labels = [
        ("Revenue Growth", "revenue_growth", "0.0%"),
        ("EBIT Margin", "ebit_margin", "0.0%"),
        ("Tax Rate", "tax_rate", "0.0%"),
        ("D&A % Revenue", "da_pct_revenue", "0.0%"),
        ("CapEx % Revenue", "capex_pct_revenue", "0.0%"),
        ("NWC % Delta Revenue", "nwc_pct_delta_revenue", "0.0%"),
        ("WACC", "wacc", "0.0%"),
        ("Terminal Growth", "terminal_growth", "0.0%"),
    ]
    for scenario_name, start_row in scenario_start.items():
        scen_rows[scenario_name] = start_row
        ws.cell(start_row, 1, scenario_name).fill = styles["blue_fill"]
        ws.cell(start_row, 1).font = styles["bold_font"]
        for idx, year in enumerate(projection_years, start=2):
            cell = ws.cell(start_row, idx, f"{year}E")
            cell.fill = styles["blue_fill"]
            cell.font = styles["bold_font"]
            cell.alignment = styles["center"]
        for row_offset, (label, key, fmt) in enumerate(assumption_labels, start=1):
            row = start_row + row_offset
            ws.cell(row, 1, label).font = styles["bold_font"]
            values = case_data[scenario_name][key]
            for col_offset in range(periods):
                cell = ws.cell(row, 2 + col_offset)
                _input_cell(cell, values[col_offset], case_data[scenario_name]["source"], fmt)

    selected_start = 52
    _section_header(ws, selected_start, "SELECTED CASE ASSUMPTIONS", periods + 1, styles)
    for idx, year in enumerate(projection_years, start=2):
        ws.cell(selected_start + 1, idx, f"{year}E").fill = styles["blue_fill"]
        ws.cell(selected_start + 1, idx).font = styles["bold_font"]
    selected_rows: dict[str, int] = {}
    for row_offset, (label, key, fmt) in enumerate(assumption_labels, start=2):
        row = selected_start + row_offset
        selected_rows[key] = row
        ws.cell(row, 1, label).font = styles["bold_font"]
        for col in range(2, periods + 2):
            bear = ws.cell(scenario_start["Bear"] + row_offset - 1, col).coordinate
            base = ws.cell(scenario_start["Base"] + row_offset - 1, col).coordinate
            bull = ws.cell(scenario_start["Bull"] + row_offset - 1, col).coordinate
            _formula_cell(ws.cell(row, col), f"=IF($B$4=1,{bear},IF($B$4=2,{base},{bull}))", fmt)

    fin_start = 64
    _section_header(ws, fin_start, "HISTORICAL & PROJECTED FINANCIALS", periods + 2, styles)
    ws.cell(fin_start + 1, 2, f"{latest_year}A").fill = styles["blue_fill"]
    ws.cell(fin_start + 1, 2).font = styles["bold_font"]
    for idx, year in enumerate(projection_years, start=3):
        ws.cell(fin_start + 1, idx, f"{year}E").fill = styles["blue_fill"]
        ws.cell(fin_start + 1, idx).font = styles["bold_font"]

    rows = {
        "revenue": fin_start + 2,
        "revenue_growth": fin_start + 3,
        "ebit_margin": fin_start + 4,
        "ebit": fin_start + 5,
        "tax": fin_start + 6,
        "nopat": fin_start + 7,
        "da": fin_start + 8,
        "capex": fin_start + 9,
        "nwc_change": fin_start + 10,
        "fcf": fin_start + 11,
    }
    labels = {
        "revenue": "Revenue",
        "revenue_growth": "Revenue Growth",
        "ebit_margin": "EBIT Margin",
        "ebit": "EBIT",
        "tax": "Tax",
        "nopat": "NOPAT",
        "da": "D&A",
        "capex": "CapEx",
        "nwc_change": "NWC Change",
        "fcf": "Unlevered FCF",
    }
    for key, row in rows.items():
        ws.cell(row, 1, labels[key]).font = styles["bold_font"]

    _formula_cell(ws.cell(rows["revenue"], 2), f"=Inputs!$B${latest_row}", "#,##0", link=True)
    _formula_cell(ws.cell(rows["ebit"], 2), f"=Inputs!$C${latest_row}", "#,##0", link=True)
    _formula_cell(ws.cell(rows["ebit_margin"], 2), f"=B{rows['ebit']}/B{rows['revenue']}", "0.0%")
    _formula_cell(ws.cell(rows["tax"], 2), f"=MAX(0,B{rows['ebit']})*Inputs!$B$19", "#,##0", link=True)
    _formula_cell(ws.cell(rows["nopat"], 2), f"=B{rows['ebit']}-B{rows['tax']}", "#,##0")
    _formula_cell(ws.cell(rows["da"], 2), f"=Inputs!$D${latest_row}", "#,##0", link=True)
    _formula_cell(ws.cell(rows["capex"], 2), f"=Inputs!$E${latest_row}", "#,##0", link=True)
    _formula_cell(ws.cell(rows["nwc_change"], 2), f"=Inputs!$F${latest_row}", "#,##0", link=True)
    _formula_cell(ws.cell(rows["fcf"], 2), f"=B{rows['nopat']}+B{rows['da']}-B{rows['capex']}-B{rows['nwc_change']}", "#,##0")

    for col in range(3, periods + 3):
        prev_col = _col_letter(col - 1)
        cur_col = _col_letter(col)
        assum_col = _col_letter(col - 1)
        _formula_cell(ws.cell(rows["revenue"], col), f"={prev_col}{rows['revenue']}*(1+{assum_col}${selected_rows['revenue_growth']})", "#,##0")
        _formula_cell(ws.cell(rows["revenue_growth"], col), f"={cur_col}{rows['revenue']}/{prev_col}{rows['revenue']}-1", "0.0%")
        _formula_cell(ws.cell(rows["ebit_margin"], col), f"={assum_col}${selected_rows['ebit_margin']}", "0.0%")
        _formula_cell(ws.cell(rows["ebit"], col), f"={cur_col}{rows['revenue']}*{cur_col}{rows['ebit_margin']}", "#,##0")
        _formula_cell(ws.cell(rows["tax"], col), f"=MAX(0,{cur_col}{rows['ebit']})*{assum_col}${selected_rows['tax_rate']}", "#,##0")
        _formula_cell(ws.cell(rows["nopat"], col), f"={cur_col}{rows['ebit']}-{cur_col}{rows['tax']}", "#,##0")
        _formula_cell(ws.cell(rows["da"], col), f"={cur_col}{rows['revenue']}*{assum_col}${selected_rows['da_pct_revenue']}", "#,##0")
        _formula_cell(ws.cell(rows["capex"], col), f"={cur_col}{rows['revenue']}*{assum_col}${selected_rows['capex_pct_revenue']}", "#,##0")
        _formula_cell(ws.cell(rows["nwc_change"], col), f"=({cur_col}{rows['revenue']}-{prev_col}{rows['revenue']})*{assum_col}${selected_rows['nwc_pct_delta_revenue']}", "#,##0")
        _formula_cell(ws.cell(rows["fcf"], col), f"={cur_col}{rows['nopat']}+{cur_col}{rows['da']}-{cur_col}{rows['capex']}-{cur_col}{rows['nwc_change']}", "#,##0")

    val_start = 80
    _section_header(ws, val_start, "DCF VALUATION", periods + 2, styles)
    for col, year in enumerate(projection_years, start=3):
        ws.cell(val_start + 1, col, f"{year}E").fill = styles["blue_fill"]
        ws.cell(val_start + 1, col).font = styles["bold_font"]
    valuation_rows = {
        "fcf": val_start + 2,
        "period": val_start + 3,
        "wacc": val_start + 4,
        "discount_factor": val_start + 5,
        "pv_fcf": val_start + 6,
        "sum_pv_fcf": val_start + 8,
        "terminal_fcf": val_start + 9,
        "terminal_value": val_start + 10,
        "pv_terminal_value": val_start + 11,
        "enterprise_value": val_start + 12,
        "net_debt": val_start + 13,
        "equity_value": val_start + 14,
        "shares": val_start + 15,
        "implied_price": val_start + 16,
        "current_price": val_start + 17,
        "upside": val_start + 18,
        "tv_pct_ev": val_start + 19,
    }
    val_labels = {
        "fcf": "Unlevered FCF",
        "period": "Period",
        "wacc": "WACC",
        "discount_factor": "Discount Factor",
        "pv_fcf": "PV of FCF",
        "sum_pv_fcf": "Sum of PV FCFs",
        "terminal_fcf": "Terminal FCF",
        "terminal_value": "Terminal Value",
        "pv_terminal_value": "PV Terminal Value",
        "enterprise_value": "Enterprise Value",
        "net_debt": "Net Debt",
        "equity_value": "Equity Value",
        "shares": "Shares Outstanding",
        "implied_price": "Implied Price",
        "current_price": "Current Price",
        "upside": "Upside/(Downside)",
        "tv_pct_ev": "Terminal Value % EV",
    }
    for key, row in valuation_rows.items():
        ws.cell(row, 1, val_labels[key]).font = styles["bold_font"]

    for offset, col in enumerate(range(3, periods + 3), start=1):
        cur_col = _col_letter(col)
        assum_col = _col_letter(col - 1)
        _formula_cell(ws.cell(valuation_rows["fcf"], col), f"={cur_col}{rows['fcf']}", "#,##0")
        _formula_cell(ws.cell(valuation_rows["period"], col), f"={offset}", "0.0")
        _formula_cell(ws.cell(valuation_rows["wacc"], col), f"={assum_col}${selected_rows['wacc']}", "0.0%")
        _formula_cell(ws.cell(valuation_rows["discount_factor"], col), f"=1/(1+{cur_col}{valuation_rows['wacc']})^{cur_col}{valuation_rows['period']}", "0.000")
        _formula_cell(ws.cell(valuation_rows["pv_fcf"], col), f"={cur_col}{valuation_rows['fcf']}*{cur_col}{valuation_rows['discount_factor']}", "#,##0")

    first_proj_col = "C"
    last_proj_col = _col_letter(periods + 2)
    last_assum_col = _col_letter(periods + 1)
    _formula_cell(ws.cell(valuation_rows["sum_pv_fcf"], 2), f"=SUM({first_proj_col}{valuation_rows['pv_fcf']}:{last_proj_col}{valuation_rows['pv_fcf']})", "#,##0")
    _formula_cell(ws.cell(valuation_rows["terminal_fcf"], 2), f"={last_proj_col}{valuation_rows['fcf']}*(1+{last_assum_col}${selected_rows['terminal_growth']})", "#,##0")
    _formula_cell(ws.cell(valuation_rows["terminal_value"], 2), f"=IF({last_assum_col}${selected_rows['wacc']}>{last_assum_col}${selected_rows['terminal_growth']},B{valuation_rows['terminal_fcf']}/({last_assum_col}${selected_rows['wacc']}-{last_assum_col}${selected_rows['terminal_growth']}),NA())", "#,##0")
    _formula_cell(ws.cell(valuation_rows["pv_terminal_value"], 2), f"=B{valuation_rows['terminal_value']}/(1+{last_assum_col}${selected_rows['wacc']})^{periods}", "#,##0")
    _formula_cell(ws.cell(valuation_rows["enterprise_value"], 2), f"=B{valuation_rows['sum_pv_fcf']}+B{valuation_rows['pv_terminal_value']}", "#,##0")
    _formula_cell(ws.cell(valuation_rows["net_debt"], 2), "=B13", "#,##0")
    _formula_cell(ws.cell(valuation_rows["equity_value"], 2), f"=B{valuation_rows['enterprise_value']}-B{valuation_rows['net_debt']}", "#,##0")
    _formula_cell(ws.cell(valuation_rows["shares"], 2), "=B9", "#,##0.0")
    _formula_cell(ws.cell(valuation_rows["implied_price"], 2), f"=B{valuation_rows['equity_value']}/B{valuation_rows['shares']}", "0.00")
    _formula_cell(ws.cell(valuation_rows["current_price"], 2), "=B8", "0.00")
    _formula_cell(ws.cell(valuation_rows["upside"], 2), f"=B{valuation_rows['implied_price']}/B{valuation_rows['current_price']}-1", "0.0%")
    _formula_cell(ws.cell(valuation_rows["tv_pct_ev"], 2), f"=B{valuation_rows['pv_terminal_value']}/B{valuation_rows['enterprise_value']}", "0.0%")

    sens_start = val_start + 24
    _section_header(ws, sens_start, "SENSITIVITY: WACC VS TERMINAL GROWTH", 6, styles)
    ws.cell(sens_start + 1, 1, "WACC / g").fill = styles["blue_fill"]
    axis_offsets = [-0.01, -0.005, 0, 0.005, 0.01]
    wacc_offsets = [-0.02, -0.01, 0, 0.01, 0.02]
    for idx, offset in enumerate(axis_offsets, start=2):
        cell = ws.cell(sens_start + 1, idx)
        _formula_cell(cell, f"=MAX(0,{last_assum_col}${selected_rows['terminal_growth']}+{offset})", "0.0%")
        cell.fill = styles["blue_fill"]
        cell.font = styles["bold_font"]
    for idx, offset in enumerate(wacc_offsets, start=2):
        row = sens_start + idx
        _formula_cell(ws.cell(row, 1), f"=MAX(0.0001,{last_assum_col}${selected_rows['wacc']}+{offset})", "0.0%")
        ws.cell(row, 1).fill = styles["blue_fill"]
        ws.cell(row, 1).font = styles["bold_font"]
        for col in range(2, 7):
            fcf_terms = "+".join(
                f"{_col_letter(proj_col)}{valuation_rows['fcf']}/(1+$A{row})^{_col_letter(proj_col)}{valuation_rows['period']}"
                for proj_col in range(3, periods + 3)
            )
            formula = (
                f"=IF($A{row}>B${sens_start + 1},"
                f"(({fcf_terms})+({last_proj_col}{valuation_rows['fcf']}*(1+B${sens_start + 1})/($A{row}-B${sens_start + 1}))/(1+$A{row})^{periods}"
                f"-B{valuation_rows['net_debt']})/B{valuation_rows['shares']},NA())"
            )
            formula = formula.replace("B$", f"{_col_letter(col)}$")
            _formula_cell(ws.cell(row, col), formula, "0.00")
            if idx == 4 and col == 4:
                ws.cell(row, col).fill = styles["base_fill"]
                ws.cell(row, col).font = styles["bold_font"]

    return {
        "selected_wacc_row": selected_rows["wacc"],
        "selected_terminal_growth_row": selected_rows["terminal_growth"],
        "valuation_rows": valuation_rows,
        "last_assumption_col": periods + 1,
        "periods": periods,
    }


def _write_wacc_sheet(ws, row_map, styles) -> None:
    ws["A1"] = "WACC BUILD"
    ws["A1"].fill = styles["navy_fill"]
    ws["A1"].font = styles["white_font"]
    ws.merge_cells("A1:C1")
    rows = [
        ("Risk-Free Rate", "=Inputs!$B$16", "0.0%"),
        ("Beta", "=Inputs!$B$15", "0.00"),
        ("Equity Risk Premium", "=Inputs!$B$17", "0.0%"),
        ("Cost of Equity", "=B3+B4*B5", "0.0%"),
        ("Pre-Tax Cost of Debt", "=Inputs!$B$18", "0.0%"),
        ("Tax Rate", "=Inputs!$B$19", "0.0%"),
        ("After-Tax Cost of Debt", "=B7*(1-B8)", "0.0%"),
        ("Market Cap", "=DCF!$B$10", "#,##0"),
        ("Net Debt", "=DCF!$B$13", "#,##0"),
        ("Enterprise Value", "=B10+B11", "#,##0"),
        ("Equity Weight", "=B10/B12", "0.0%"),
        ("Net Debt Weight", "=B11/B12", "0.0%"),
        ("Calculated WACC", "=B13*B6+B14*B9", "0.0%"),
    ]
    for idx, (label, formula, fmt) in enumerate(rows, start=3):
        ws.cell(idx, 1, label).font = styles["bold_font"]
        _formula_cell(ws.cell(idx, 2), formula, fmt, link="Inputs!" in formula or "DCF!" in formula)
    ws["A18"] = "DCF selected-case WACC"
    _formula_cell(ws["B18"], f"=DCF!{_col_letter(row_map['last_assumption_col'])}${row_map['selected_wacc_row']}", "0.0%", link=True)


def _write_checks_sheet(ws, row_map, styles) -> None:
    ws["A1"] = "MODEL CHECKS"
    ws["A1"].fill = styles["navy_fill"]
    ws["A1"].font = styles["white_font"]
    ws.merge_cells("A1:C1")
    valuation_rows = row_map["valuation_rows"]
    last_col = _col_letter(row_map["last_assumption_col"])
    checks = [
        ("Terminal growth < WACC", f"=DCF!{last_col}${row_map['selected_terminal_growth_row']}<DCF!{last_col}${row_map['selected_wacc_row']}"),
        ("Implied price positive", f"=DCF!$B${valuation_rows['implied_price']}>0"),
        ("Terminal value % EV below 80%", f"=DCF!$B${valuation_rows['tv_pct_ev']}<80%"),
        ("Net debt linked", f"=DCF!$B${valuation_rows['net_debt']}=DCF!$B$13"),
    ]
    for idx, (label, formula) in enumerate(checks, start=3):
        ws.cell(idx, 1, label).font = styles["bold_font"]
        _formula_cell(ws.cell(idx, 2), formula)


@tool
def validate_dcf_model(excel_path: str) -> str:
    """Validate a generated DCF workbook and write validation.json beside it."""
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        return json.dumps({"status": "ERROR", "message": f"openpyxl is not installed - {exc}"})

    path = Path(excel_path)
    if not artifact_exists(path):
        raise FileNotFoundError(f"File not found: {excel_path}")

    wb = _load_workbook(path, data_only=False)
    errors: list[str] = []
    warnings: list[str] = []
    info: list[str] = []
    required = ["Inputs", "DCF", "WACC", "Checks"]
    for sheet in required:
        if sheet not in wb.sheetnames:
            errors.append(f"Missing required sheet: {sheet}")
        else:
            info.append(f"Found sheet: {sheet}")

    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    for token in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!"]:
                        if token in cell.value:
                            errors.append(f"{token} appears in formula at {ws.title}!{cell.coordinate}")
    info.append(f"Total formulas: {formula_count}")

    if "DCF" in wb.sheetnames and "Inputs" in wb.sheetnames:
        model_metrics = _compute_base_case_from_workbook(wb)
        if model_metrics:
            wacc = model_metrics["wacc"]
            terminal_growth = model_metrics["terminal_growth"]
            tv_pct_ev = model_metrics["terminal_value_pct_ev"]
            if terminal_growth >= wacc:
                errors.append("Terminal growth is greater than or equal to WACC")
            if not 0.05 <= wacc <= 0.20:
                warnings.append(f"WACC outside typical range: {wacc:.1%}")
            if tv_pct_ev > 0.80:
                warnings.append(f"Terminal value is high as % of EV: {tv_pct_ev:.1%}")
            elif tv_pct_ev < 0.40:
                warnings.append(f"Terminal value is low as % of EV: {tv_pct_ev:.1%}")
            info.append(f"Base implied price: {model_metrics['implied_price']:.2f}")
            info.append(f"Terminal value % EV: {tv_pct_ev:.1%}")

    results = {
        "file": str(path),
        "validation_date": datetime.now().isoformat(),
        "status": "PASS" if not errors else "FAIL",
        "error_count": len(errors),
        "warning_count": len(warnings),
        "errors": errors,
        "warnings": warnings,
        "info": info,
    }
    validation_path = path.parent / "validation.json"
    _write_text(validation_path, json.dumps(results, indent=2))
    results["validation_path"] = str(validation_path)
    return json.dumps(results, indent=2)


def _compute_base_case_from_workbook(wb) -> dict[str, float] | None:
    ws_inputs = wb["Inputs"]
    ws_dcf = wb["DCF"]
    latest_row = ws_inputs.max_row
    for row in range(ws_inputs.max_row, 24, -1):
        if ws_inputs.cell(row, 2).value not in (None, ""):
            latest_row = row
            break

    revenue = _as_float(ws_inputs.cell(latest_row, 2).value)
    debt = _as_float(ws_inputs["B12"].value)
    cash = _as_float(ws_inputs["B13"].value)
    shares = _as_float(ws_inputs["B14"].value, 1.0)
    if shares == 0:
        return None

    base_start = 30
    periods = 5
    for col in range(2, 10):
        if ws_dcf.cell(base_start, col).value in (None, ""):
            periods = col - 2
            break

    growth = [_as_decimal(ws_dcf.cell(base_start + 1, col).value) for col in range(2, 2 + periods)]
    margin = [_as_decimal(ws_dcf.cell(base_start + 2, col).value) for col in range(2, 2 + periods)]
    tax = [_as_decimal(ws_dcf.cell(base_start + 3, col).value) for col in range(2, 2 + periods)]
    da_pct = [_as_decimal(ws_dcf.cell(base_start + 4, col).value) for col in range(2, 2 + periods)]
    capex_pct = [_as_decimal(ws_dcf.cell(base_start + 5, col).value) for col in range(2, 2 + periods)]
    nwc_pct = [_as_decimal(ws_dcf.cell(base_start + 6, col).value) for col in range(2, 2 + periods)]
    wacc = _as_decimal(ws_dcf.cell(base_start + 7, 1 + periods).value)
    terminal_growth = _as_decimal(ws_dcf.cell(base_start + 8, 1 + periods).value)

    fcf_values: list[float] = []
    prev_revenue = revenue
    for idx in range(periods):
        revenue = prev_revenue * (1 + growth[idx])
        ebit = revenue * margin[idx]
        nopat = ebit - max(0, ebit) * tax[idx]
        da = revenue * da_pct[idx]
        capex = revenue * capex_pct[idx]
        nwc = (revenue - prev_revenue) * nwc_pct[idx]
        fcf = nopat + da - capex - nwc
        fcf_values.append(fcf)
        prev_revenue = revenue

    pv_fcf = sum(fcf / ((1 + wacc) ** (idx + 1)) for idx, fcf in enumerate(fcf_values))
    if wacc <= terminal_growth:
        return {
            "wacc": wacc,
            "terminal_growth": terminal_growth,
            "terminal_value_pct_ev": 1.0,
            "implied_price": 0.0,
        }
    terminal_fcf = fcf_values[-1] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (wacc - terminal_growth)
    pv_terminal_value = terminal_value / ((1 + wacc) ** periods)
    enterprise_value = pv_fcf + pv_terminal_value
    equity_value = enterprise_value - (debt - cash)
    implied_price = equity_value / shares
    tv_pct_ev = pv_terminal_value / enterprise_value if enterprise_value else 0.0
    return {
        "wacc": wacc,
        "terminal_growth": terminal_growth,
        "terminal_value_pct_ev": tv_pct_ev,
        "implied_price": implied_price,
    }


@tool
def write_valuation_summary(
    summary_json: str,
    output_dir: str = "./out",
    exact_output_dir: bool = False,
) -> str:
    """Write a concise valuation summary markdown file."""
    payload = _parse_json(summary_json, "summary_json")
    if not isinstance(payload, dict):
        raise ValueError("summary_json must be a JSON object")

    out_dir = _timestamped_output_dir(output_dir, exact_output_dir)
    filepath = out_dir / "valuation-summary.md"
    validation = payload.get("validation")
    if not validation:
        validation_path = out_dir / "validation.json"
        validation_text = read_text_artifact(validation_path, missing_ok=True)
        if validation_text is not None:
            try:
                validation = json.loads(validation_text)
            except json.JSONDecodeError:
                validation = None

    company = payload.get("company") or payload.get("company_name") or "Company"
    ticker = payload.get("ticker") or ""
    valuation_date = payload.get("valuation_date") or datetime.now().strftime("%Y-%m-%d")
    lines = [
        f"# {company} ({ticker}) DCF Valuation Summary".strip(),
        "",
        f"- Valuation date: {valuation_date}",
        f"- Current price: {payload.get('current_price', '[UNSOURCED]')}",
        f"- Base implied price: {payload.get('base_implied_price', '[UNSOURCED]')}",
        f"- Upside/(downside): {payload.get('upside_downside', '[UNSOURCED]')}",
        f"- WACC: {payload.get('wacc', '[UNSOURCED]')}",
        f"- Terminal growth: {payload.get('terminal_growth', '[UNSOURCED]')}",
        f"- Terminal value % EV: {payload.get('terminal_value_pct_ev', '[UNSOURCED]')}",
        "",
        "## Key Drivers",
    ]
    for item in payload.get("key_drivers", []) or ["[UNSOURCED]"]:
        lines.append(f"- {item}")
    lines.extend(["", "## Sources"])
    for item in payload.get("sources", []) or ["[UNSOURCED]"]:
        lines.append(f"- {item}")
    lines.extend(["", "## Validation"])
    if isinstance(validation, dict):
        lines.append(f"- Status: {validation.get('status', '[UNSOURCED]')}")
        if validation.get("error_count") is not None:
            lines.append(f"- Errors: {validation.get('error_count')}")
        if validation.get("warning_count") is not None:
            lines.append(f"- Warnings: {validation.get('warning_count')}")
        for warning in validation.get("warnings", []) or []:
            lines.append(f"- Warning: {warning}")
        for error in validation.get("errors", []) or []:
            lines.append(f"- Error: {error}")
    else:
        lines.append(f"- {validation or '[UNSOURCED]'}")
    lines.append("")

    _write_text(filepath, "\n".join(lines))
    return str(filepath)


@tool
def write_assumption_analysis(
    assumption_markdown: str,
    output_dir: str = "./out",
    filename: str = "assumption-analysis.md",
    exact_output_dir: bool = False,
) -> str:
    """Write the DCF assumption analysis Markdown pack to the task output dir."""
    if not isinstance(assumption_markdown, str) or not assumption_markdown.strip():
        raise ValueError("assumption_markdown must be non-empty Markdown text")

    safe_name = Path(filename).name or "assumption-analysis.md"
    if not safe_name.endswith(".md"):
        safe_name += ".md"

    out_dir = _timestamped_output_dir(output_dir, exact_output_dir)
    filepath = out_dir / safe_name
    _write_text(filepath, assumption_markdown.rstrip() + "\n")
    return str(filepath)
