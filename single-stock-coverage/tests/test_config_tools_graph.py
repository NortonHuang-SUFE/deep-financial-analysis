import asyncio
import importlib
import json
import sys
from pathlib import Path
from types import SimpleNamespace

import single_stock_coverage_agent.config as config_module
from single_stock_coverage_agent.agent_registry import (
    ToolGroupResolver,
    agent_uses_tool_group,
    describe_agent,
    load_agent_registry,
    mcp_tool_group_names,
    mcp_tool_group_server_names,
)
from single_stock_coverage_agent.config import (
    PROJECT_ROOT,
    WORKSPACE_ROOT,
    enabled_mcp_server_configs,
    file_storage_root,
    load_config,
)
from single_stock_coverage_agent import tools

DCF_SRC_ROOT = WORKSPACE_ROOT / "DCF-builder" / "src"
if str(DCF_SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(DCF_SRC_ROOT))


def _statement_payload(statement_type: str) -> dict:
    canonical = {
        "income_statement": [
            "revenue_total",
            "gross_profit",
            "ebit",
            "ebitda",
            "interest_expense",
            "pretax_income",
            "tax_expense",
            "net_income",
            "da_total",
        ],
        "balance_sheet": [
            "cash_and_equivalents",
            "total_current_assets",
            "total_assets",
            "total_current_liabilities",
            "total_debt",
            "retained_earnings",
            "total_equity",
            "total_liabilities_and_equity",
        ],
        "cash_flow": [
            "net_income_cf",
            "da_addback",
            "nwc_change",
            "cfo_total",
            "capex",
            "cfi_total",
            "debt_proceeds_repayments",
            "dividends",
            "cff_total",
            "beginning_cash",
            "ending_cash",
        ],
    }[statement_type]
    dependencies = {
        "income_statement": [
            "revenue_build.total_revenue",
            "debt_interest.interest_expense",
            "share_count.diluted_shares",
        ],
        "balance_sheet": [
            "cash_flow.ending_cash",
            "income_statement.net_income",
            "share_count.dividends",
        ],
        "cash_flow": [
            "income_statement.net_income",
            "ppe_da.da_total",
            "balance_sheet.cash_and_equivalents",
        ],
    }[statement_type]
    values = {
        "revenue_total": 1000,
        "gross_profit": 500,
        "ebit": 290,
        "ebitda": 320,
        "interest_expense": 10,
        "pretax_income": 280,
        "tax_expense": 70,
        "net_income": 210,
        "da_total": 30,
        "cash_and_equivalents": 120,
        "total_current_assets": 300,
        "total_assets": 900,
        "total_current_liabilities": 150,
        "total_debt": 200,
        "retained_earnings": 300,
        "total_equity": 450,
        "total_liabilities_and_equity": 900,
        "net_income_cf": 210,
        "da_addback": 30,
        "nwc_change": -10,
        "cfo_total": 230,
        "capex": 45,
        "cfi_total": -45,
        "debt_proceeds_repayments": 0,
        "dividends": 0,
        "cff_total": 0,
        "beginning_cash": 100,
        "ending_cash": 120,
    }
    payload = {
        "company": "Example Co",
        "ticker": "EXM",
        "market": "US",
        "currency": "USD",
        "unit": "millions",
        "fiscal_year_end": "Dec",
        "statement_type": statement_type,
        "canonical_row_keys": canonical,
        "line_items": [{"name": key} for key in canonical],
        "historical_inputs": [
            {
                "period": "FY2023A",
                "canonical_key": key,
                "value": values.get(key, 0),
                "source": "FY2023 annual report",
                "currency": "USD",
                "unit": "millions",
            }
            for key in canonical
        ],
        "forecast_logic": {"method": "formula-driven"},
        "assumption_requirements": ["revenue growth"],
        "cross_statement_dependencies": dependencies,
        "source_coverage": {"status": "sourced"},
        "unsourced_items": [],
        "validation_status": "draft_checked",
    }
    if statement_type == "income_statement":
        payload["revenue_build_spec"] = {
            "statement_type": "revenue_build",
            "segments": [],
        }
    return payload


def _minimal_model_input() -> dict:
    return {
        "company": "Example Co",
        "ticker": "EXM",
        "market": "US",
        "currency": "USD",
        "unit": "millions",
        "fiscal_year_end": "Dec",
        "projection_periods": 2,
        "historicals": [
            {
                "period": "FY2023A",
                "year": 2023,
                "revenue": 1000,
                "gross_profit": 500,
                "operating_expenses": 180,
                "da": 30,
                "ebit": 290,
                "ebitda": 320,
                "interest_expense": 10,
                "pretax_income": 280,
                "tax_expense": 70,
                "net_income": 210,
                "capex": 45,
                "cash": 120,
                "debt": 200,
                "retained_earnings": 300,
                "shares": 100,
                "source": "FY2023 annual report",
            }
        ],
        "assumptions": {
            "revenue_growth": 0.05,
            "gross_margin": 0.5,
            "opex_pct_revenue": 0.18,
            "tax_rate": 0.25,
            "da_pct_revenue": 0.03,
            "capex_pct_revenue": 0.04,
        },
    }


def _nested_a_share_model_input() -> dict:
    financial_facts = {
        "company": {
            "legal_name": "成都新易盛通信技术股份有限公司",
            "short_name": "新易盛",
            "ticker": "300502.SZ",
            "market": "A-share",
            "currency": "CNY",
            "reporting_unit": "亿元",
            "fiscal_year_end": "12-31",
        },
        "historicals": [
            {
                "year": "2022",
                "is": {
                    "revenue": 33.1057,
                    "cogs": 20.9686,
                    "gross_profit": 12.1372,
                    "finance_expenses": -1.2452,
                    "interest_income": 0.2151,
                    "operating_profit": 10.2778,
                    "pretax_income": 10.2786,
                    "income_tax": 1.2428,
                    "net_income": 9.0358,
                    "ebit": 10.2924,
                    "ebitda": 11.3133,
                },
                "bs": {
                    "cash": 17.8497,
                    "accounts_receivable": 6.6196,
                    "inventory": 14.692,
                    "current_assets": 42.0537,
                    "ppe_net": 6.6102,
                    "total_assets": 58.7606,
                    "accounts_payable": 4.3614,
                    "short_term_debt": 0.0,
                    "long_term_debt": 0.0,
                    "current_liabilities": 9.3173,
                    "total_liabilities": 10.4689,
                    "total_equity": 48.2917,
                },
                "cf": {"capex": 3.6472},
                "derived": {"da_estimated": 1.0209},
            },
            {
                "year": "2023",
                "is": {
                    "revenue": 30.9761,
                    "cogs": 21.3766,
                    "gross_profit": 9.5995,
                    "finance_expenses": -1.0824,
                    "interest_income": 0.7456,
                    "operating_profit": 7.848,
                    "pretax_income": 7.887,
                    "income_tax": 1.003,
                    "net_income": 6.884,
                    "ebit": 7.9,
                    "ebitda": 8.92,
                },
                "bs": {
                    "cash": 25.15,
                    "accounts_receivable": 7.15,
                    "inventory": 9.63,
                    "current_assets": 43.35,
                    "ppe_net": 13.1,
                    "total_assets": 64.4,
                    "accounts_payable": 5.98,
                    "short_term_debt": 0.0,
                    "long_term_debt": 0.0,
                    "current_liabilities": 8.67,
                    "total_liabilities": 9.74,
                    "total_equity": 54.66,
                },
                "cf": {"capex": 5.54},
                "derived": {"da_estimated": 1.02},
            },
            {
                "year": "2024",
                "is": {
                    "revenue": 86.4683,
                    "cogs": 47.802,
                    "gross_profit": 38.6663,
                    "finance_expenses": -1.7932,
                    "interest_income": 0.7284,
                    "operating_profit": 32.315,
                    "pretax_income": 32.342,
                    "income_tax": 3.964,
                    "net_income": 28.378,
                    "ebit": 32.36,
                    "ebitda": 34.0,
                },
                "bs": {
                    "cash": 16.0,
                    "accounts_receivable": 25.39,
                    "inventory": 41.32,
                    "current_assets": 89.25,
                    "ppe_net": 20.22,
                    "total_assets": 122.67,
                    "accounts_payable": 21.94,
                    "short_term_debt": 0.0,
                    "long_term_debt": 0.0,
                    "current_liabilities": 38.25,
                    "total_liabilities": 39.38,
                    "total_equity": 83.28,
                },
                "cf": {"capex": 14.76},
                "derived": {"da_estimated": 1.64},
            },
            {
                "year": "2025",
                "is": {
                    "revenue": 248.4185,
                    "cogs": 129.6605,
                    "gross_profit": 118.758,
                    "finance_expenses": -3.1818,
                    "interest_income": 0.7795,
                    "operating_profit": 108.5927,
                    "pretax_income": 108.6448,
                    "income_tax": 13.1244,
                    "net_income": 95.5328,
                    "ebit": 108.6567,
                    "ebitda": 112.1171,
                },
                "bs": {
                    "cash": 81.5561,
                    "accounts_receivable": 44.3751,
                    "inventory": 72.3433,
                    "current_assets": 205.1599,
                    "ppe_net": 33.5386,
                    "total_assets": 258.8107,
                    "accounts_payable": 37.2637,
                    "interest_bearing_debt": 0.1493887,
                    "short_term_debt_raw": 15.5258,
                    "debt_to_market_equity": 0.0035,
                    "current_liabilities": 68.9625,
                    "total_liabilities": 78.1682,
                    "total_equity": 180.6425,
                },
                "cf": {"capex": 13.197},
                "derived": {"da_estimated": 3.4604},
            },
        ],
        "shares": {
            "2022": {"total_shares_期末": 5.0709},
            "2023": {"total_shares_期末": 7.0992},
            "2024": {"total_shares_期末": 7.0881},
            "2025": {"total_shares_期末": 9.9401},
        },
    }
    revenue_build_spec = {
        "segments": [
            {
                "label": "800G Revenue",
                "historical_revenue": [
                    {"period": "2022", "value": 28},
                    {"period": "2023", "value": 22},
                    {"period": "2024", "value": 72},
                    {"period": "2025", "value": 185},
                ],
                "forecast_revenue": [
                    {"period": "2026E", "value": 190},
                    {"period": "2027E", "value": 180},
                    {"period": "2028E", "value": 160},
                ],
            },
            {
                "label": "1.6T Revenue",
                "historical_revenue": [
                    {"period": "2022", "value": 0},
                    {"period": "2023", "value": 0},
                    {"period": "2024", "value": 5},
                    {"period": "2025", "value": 50},
                ],
                "forecast_revenue": [
                    {"period": "2026E", "value": 160},
                    {"period": "2027E", "value": 290},
                    {"period": "2028E", "value": 420},
                ],
            },
            {
                "label": "Other Revenue",
                "historical_revenue": [
                    {"period": "2022", "value": 5.1057},
                    {"period": "2023", "value": 8.9761},
                    {"period": "2024", "value": 9.4683},
                    {"period": "2025", "value": 13.4185},
                ],
                "forecast_revenue": [
                    {"period": "2026E", "value": 30},
                    {"period": "2027E", "value": 40},
                    {"period": "2028E", "value": 60},
                ],
            },
        ],
        "total_revenue_reconciliation": {
            "forecast": [
                {"period": "2026E", "total": 380},
                {"period": "2027E", "total": 510},
                {"period": "2028E", "total": 640},
            ]
        },
    }
    statement_spec_pack = {
        "status": "PASS",
        "statement_specs": {
            "income_statement": {
                "revenue_build_spec": revenue_build_spec,
                "forecast_logic": {
                    "cogs": {
                        "assumptions": [
                            "2026E: gross margin 48.0%",
                            "2027E: gross margin 47.0%",
                            "2028E: gross margin 46.0%",
                        ]
                    },
                    "finance_expenses": {
                        "assumptions": [
                            "2026E: -4.5亿",
                            "2027E: -5.5亿",
                            "2028E: -6.5亿",
                        ]
                    },
                },
            },
            "balance_sheet": {
                "assumption_requirements": [
                    {
                        "name": "AR Days (DSO)",
                        "forecast_values": {"2026E": 60, "2027E": 58, "2028E": 55},
                    },
                    {
                        "name": "Inventory Days (DIO)",
                        "forecast_values": {"2026E": 100, "2027E": 95, "2028E": 90},
                    },
                    {
                        "name": "AP Days (DPO)",
                        "forecast_values": {"2026E": 52, "2027E": 50, "2028E": 50},
                    },
                    {
                        "name": "CapEx as % of Revenue",
                        "forecast_values": {"2026E": 6, "2027E": 5, "2028E": 4},
                    },
                    {
                        "name": "Dividend Payout Ratio",
                        "forecast_values": {
                            "2026E": "35% of NI",
                            "2027E": "35% of NI",
                            "2028E": "35% of NI",
                        },
                    },
                ]
            },
            "cash_flow": {},
        },
    }
    return {
        **financial_facts,
        "financial_facts": financial_facts,
        "statement_spec_pack": statement_spec_pack,
        "task2_context_packet": {
            "company_metadata": {
                "ticker": "300502.SZ",
                "company": "新易盛",
                "market": "A-share",
                "currency": "CNY",
                "reporting_unit": "亿元",
                "fiscal_year_end": "12-31",
            }
        },
    }


def _write_model_source_files(model_dir, payload: dict) -> None:
    financial_facts = (
        payload.get("financial_facts")
        if isinstance(payload.get("financial_facts"), dict)
        else payload
    )
    context_packet = payload.get("task2_context_packet")
    if not isinstance(context_packet, dict):
        context_packet = {
            "company_metadata": {
                "company": financial_facts.get("company", "Example Co"),
                "ticker": financial_facts.get("ticker", "EXM"),
                "market": financial_facts.get("market", "US"),
                "currency": financial_facts.get("currency", "USD"),
                "reporting_unit": financial_facts.get("unit", "millions"),
            },
            "period_plan": {"forecast_horizon": "FY2024E-FY2025E"},
            "canonical_row_keys": {
                statement_type: list(tools.STATEMENT_CANONICAL_KEYS[statement_type])
                for statement_type in tools.STATEMENT_JSON_ALLOWED_TYPES
            },
        }
    statement_pack = payload.get("statement_spec_pack")
    if not isinstance(statement_pack, dict):
        statement_pack = {
            "status": "PASS",
            "critical_count": 0,
            "warning_count": 0,
            "builder_blocked": False,
            "statement_specs": {},
        }
    model_dir.mkdir(parents=True, exist_ok=True)
    (model_dir / "financial_facts.json").write_text(
        json.dumps(financial_facts, ensure_ascii=False),
        encoding="utf-8",
    )
    (model_dir / "statement_spec_pack.json").write_text(
        json.dumps(statement_pack, ensure_ascii=False),
        encoding="utf-8",
    )
    (model_dir / "task2_context_packet.json").write_text(
        json.dumps(context_packet, ensure_ascii=False),
        encoding="utf-8",
    )
    revenue_build_spec = payload.get("revenue_build_spec")
    if not isinstance(revenue_build_spec, dict):
        income_spec = (statement_pack.get("statement_specs") or {}).get(
            "income_statement"
        )
        if isinstance(income_spec, dict):
            revenue_build_spec = income_spec.get("revenue_build_spec")
    if isinstance(revenue_build_spec, dict):
        (model_dir / "revenue_build_spec.json").write_text(
            json.dumps(revenue_build_spec, ensure_ascii=False),
            encoding="utf-8",
        )


def _write_task1_fixture(run_dir):
    (run_dir / "01_company_research").mkdir(parents=True, exist_ok=True)
    (run_dir / "01_company_research" / "company_research.md").write_text(
        "# Example Co\n",
        encoding="utf-8",
    )
    (run_dir / "01_company_research" / "business_driver_map.json").write_text(
        json.dumps({"company": "Example Co", "ticker": "EXM"}),
        encoding="utf-8",
    )
    (run_dir / "01_company_research" / "source_log.json").write_text(
        json.dumps({"sources": ["annual report"]}),
        encoding="utf-8",
    )


def _agent_prompt(name: str) -> str:
    return (PROJECT_ROOT / "agents" / name).read_text(encoding="utf-8")


def _skill_text(name: str) -> str:
    return (PROJECT_ROOT / "skills" / name / "SKILL.md").read_text(encoding="utf-8")


def _clear_env(monkeypatch):
    for env_name in [
        "MODEL_NAME",
        "MODEL_GATEWAY_BASE_URL",
        "MODEL_GATEWAY_API_KEY",
        "MODEL_RELAY_BASE_URL",
        "MODEL_RELAY_API_KEY",
        "MODEL_BASE_URL",
        "MODEL_API_KEY",
        "MODEL_THINKING",
        "MODEL_MAX_TOKENS",
        "DASHSCOPE_API_KEY",
        "ALIBABA_API_KEY",
        "IFIND_MCP_AUTHORIZATION",
        "IFIND_MCP_TOKEN",
        "MX_DS_MCP_API_KEY",
        "MX_DS_MCP_EM_API_KEY",
        "EASTMONEY_MX_DS_MCP_API_KEY",
        "MX_DS_MCP_URL",
        "MX_DS_MCP_TRANSPORT",
        "MX_DS_MCP_MCP_URL",
        "MX_DS_MCP_MCP_TRANSPORT",
        "AGENT_FILE_STORAGE_ROOT",
        "SINGLE_STOCK_COVERAGE_DISABLE_MCP",
        "SINGLE_STOCK_COVERAGE_TEST_MODE",
        "SINGLE_STOCK_COVERAGE_OUTPUT_TIMESTAMP",
    ]:
        monkeypatch.delenv(env_name, raising=False)


def test_default_config_resolves_from_project_root(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    monkeypatch.setattr(config_module, "WORKSPACE_ENV_PATH", tmp_path / "missing.env")
    monkeypatch.chdir(tmp_path)

    cfg = load_config()

    assert PROJECT_ROOT.name == "single-stock-coverage"
    assert WORKSPACE_ROOT.name == "financialServicesModified"
    assert cfg.model.default == "qwen-3.7-max"
    assert cfg.output.dir == "./out/coverage"
    assert "ifind-stock" in cfg.mcp
    assert "mx-ds-mcp" in cfg.mcp


def test_workspace_env_and_process_env_override_ifind_auth(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    workspace_env = tmp_path / ".env"
    workspace_env.write_text("IFIND_MCP_TOKEN=from-workspace-env\n", encoding="utf-8")
    monkeypatch.setattr(config_module, "WORKSPACE_ENV_PATH", workspace_env)

    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        """
mcp:
  ifind-stock:
    url: "https://example.test/stock"
    transport: "streamable_http"
""",
        encoding="utf-8",
    )

    cfg = load_config(str(config_path))
    server_configs = enabled_mcp_server_configs(cfg)

    assert server_configs["ifind-stock"]["headers"] == {
        "Authorization": "Bearer from-workspace-env"
    }

    monkeypatch.setenv("IFIND_MCP_AUTHORIZATION", "from-process-env")
    cfg = load_config(str(config_path))
    server_configs = enabled_mcp_server_configs(cfg)

    assert server_configs["ifind-stock"]["headers"] == {
        "Authorization": "from-process-env"
    }


def test_mx_ds_mcp_config_uses_env_header_and_nested_servers(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    workspace_env = tmp_path / ".env"
    workspace_env.write_text("MX_DS_MCP_API_KEY=from-workspace-env\n", encoding="utf-8")
    monkeypatch.setattr(config_module, "WORKSPACE_ENV_PATH", workspace_env)

    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        """
mcp:
  servers:
    ifind-stock:
      url: "https://example.test/stock"
      transport: "streamable_http"
    mx-ds-mcp:
      url: "https://mxapi.eastmoney.com/mxds/mcp"
      transport: "streamable-http"
      connectTimeout: 10
      timeout: 120
      headers:
        em_api_key: "${MX_DS_MCP_API_KEY}"
""",
        encoding="utf-8",
    )

    cfg = load_config(str(config_path))
    assert cfg.mcp["mx-ds-mcp"].transport == "streamable_http"
    assert cfg.mcp["mx-ds-mcp"].connect_timeout == 10

    server_configs = enabled_mcp_server_configs(cfg)
    assert server_configs["mx-ds-mcp"] == {
        "url": "https://mxapi.eastmoney.com/mxds/mcp",
        "transport": "streamable_http",
        "headers": {"em_api_key": "from-workspace-env"},
        "timeout": 120,
    }
    assert set(enabled_mcp_server_configs(cfg, server_names={"mx-ds-mcp"})) == {
        "mx-ds-mcp"
    }

    monkeypatch.setenv("MX_DS_MCP_API_KEY", "from-process-env")
    cfg = load_config(str(config_path))
    server_configs = enabled_mcp_server_configs(cfg)
    assert server_configs["mx-ds-mcp"]["headers"] == {"em_api_key": "from-process-env"}


def test_mx_ds_mcp_url_and_transport_env_aliases(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    monkeypatch.setattr(config_module, "WORKSPACE_ENV_PATH", tmp_path / "missing.env")
    monkeypatch.setenv("MX_DS_MCP_URL", "https://example.test/mx")
    monkeypatch.setenv("MX_DS_MCP_TRANSPORT", "streamable-http")
    monkeypatch.setenv("MX_DS_MCP_API_KEY", "env-key")

    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        """
mcp:
  mx-ds-mcp:
    url: ""
    transport: "sse"
    headers:
      em_api_key: ""
""",
        encoding="utf-8",
    )

    cfg = load_config(str(config_path))
    server_configs = enabled_mcp_server_configs(cfg)
    assert server_configs["mx-ds-mcp"]["url"] == "https://example.test/mx"
    assert server_configs["mx-ds-mcp"]["transport"] == "streamable_http"
    assert server_configs["mx-ds-mcp"]["headers"] == {"em_api_key": "env-key"}


def test_coverage_run_and_artifact_tools(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    monkeypatch.setenv("SINGLE_STOCK_COVERAGE_OUTPUT_TIMESTAMP", "20260604-120000")

    result_json = tools.create_coverage_run_dir.invoke(
        {
            "company": "测试公司",
            "ticker": "000001.SZ",
            "market": "A-share",
            "task_type": "initiation",
            "triggering_event": "",
        }
    )
    result = json.loads(result_json)

    assert result["run_dir"] == "out/coverage/a-share-000001.sz/runs/20260604-120000"
    manifest_path = tmp_path / result["manifest_path"]
    state_path = tmp_path / result["coverage_state_path"]
    assert manifest_path.exists()
    assert state_path.exists()

    md_path = tools.write_markdown_artifact.invoke(
        {
            "markdown": "# Test\n",
            "filename": "company_research.md",
            "subdir": "01_company_research",
            "ticker": "000001.SZ",
            "market": "A-share",
            "run_dir": result["run_dir"],
        }
    )
    json_path = tools.write_json_artifact.invoke(
        {
            "data_json": '{"ticker": "000001.SZ"}',
            "filename": "business_driver_map.json",
            "subdir": "01_company_research",
            "ticker": "000001.SZ",
            "market": "A-share",
            "run_dir": result["run_dir"],
        }
    )

    assert (tmp_path / md_path).read_text(encoding="utf-8") == "# Test\n"
    assert json.loads((tmp_path / json_path).read_text(encoding="utf-8")) == {
        "ticker": "000001.SZ"
    }

    manifest_result = tools.update_run_manifest.invoke(
        {
            "patch_json": '{"subagents_called": ["task1_company_researcher"]}',
            "ticker": "000001.SZ",
            "market": "A-share",
            "run_dir": result["run_dir"],
        }
    )
    manifest = json.loads((tmp_path / manifest_result).read_text(encoding="utf-8"))
    assert manifest["subagents_called"] == ["task1_company_researcher"]

    oversized_json = json.dumps({"blob": "x" * tools.INLINE_JSON_MAX_BYTES})
    json_failure = json.loads(
        tools.write_json_artifact.invoke(
            {
                "data_json": oversized_json,
                "filename": "too_large.json",
                "subdir": "01_company_research",
                "ticker": "000001.SZ",
                "market": "A-share",
                "run_dir": result["run_dir"],
            }
        )
    )
    assert json_failure["status"] == "FAIL"
    assert json_failure["field"] == "data_json"

    manifest_failure = json.loads(
        tools.update_run_manifest.invoke(
            {
                "patch_json": oversized_json,
                "ticker": "000001.SZ",
                "market": "A-share",
                "run_dir": result["run_dir"],
            }
        )
    )
    assert manifest_failure["status"] == "FAIL"
    assert manifest_failure["field"] == "patch_json"


def test_manifest_rejects_completed_task_without_subagent(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    monkeypatch.setenv("SINGLE_STOCK_COVERAGE_OUTPUT_TIMESTAMP", "20260604-121000")

    result = json.loads(
        tools.create_coverage_run_dir.invoke(
            {
                "company": "测试公司",
                "ticker": "000001.SZ",
                "market": "A-share",
                "task_type": "initiation",
                "triggering_event": "",
            }
        )
    )

    failure = json.loads(
        tools.update_run_manifest.invoke(
            {
                "patch_json": json.dumps(
                    {
                        "subagents_called": [],
                        "tasks": {
                            "task1_company_research": {
                                "status": "completed",
                                "artifacts": [
                                    "01_company_research/company_research.md",
                                    "01_company_research/business_driver_map.json",
                                    "01_company_research/source_log.json",
                                ],
                            }
                        },
                    }
                ),
                "ticker": "000001.SZ",
                "market": "A-share",
                "run_dir": result["run_dir"],
            }
        )
    )

    assert failure["status"] == "FAIL"
    assert failure["reason"] == "invalid_top_level_task_completion"
    assert failure["failures"][0]["task"] == "task1_company_research"
    assert failure["failures"][0]["required_subagent"] == "task1_company_researcher"
    manifest = json.loads((tmp_path / result["manifest_path"]).read_text())
    assert "tasks" not in manifest


def test_manifest_rejects_simplified_task2_completion(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    monkeypatch.setenv("SINGLE_STOCK_COVERAGE_OUTPUT_TIMESTAMP", "20260604-121500")

    result = json.loads(
        tools.create_coverage_run_dir.invoke(
            {
                "company": "测试公司",
                "ticker": "000001.SZ",
                "market": "A-share",
                "task_type": "initiation",
                "triggering_event": "",
            }
        )
    )
    model_dir = tmp_path / result["run_dir"] / "02_financial_model"
    model_dir.mkdir(parents=True)
    (model_dir / "three_statement_model.md").write_text("# Model\n", encoding="utf-8")
    (model_dir / "financial_drivers.json").write_text("{}", encoding="utf-8")

    failure = json.loads(
        tools.update_run_manifest.invoke(
            {
                "patch_json": json.dumps(
                    {
                        "subagents_called": ["task2_financial_modeler"],
                        "tasks": {
                            "task2_financial_model": {
                                "status": "completed",
                                "artifacts": [
                                    "02_financial_model/three_statement_model.md",
                                    "02_financial_model/financial_drivers.json",
                                ],
                            }
                        },
                    }
                ),
                "ticker": "000001.SZ",
                "market": "A-share",
                "run_dir": result["run_dir"],
            }
        )
    )

    assert failure["status"] == "FAIL"
    missing = failure["failures"][0]["missing_artifacts"]
    assert "02_financial_model/financial_facts.json" in missing
    assert "02_financial_model/integrated_model.xlsx" in missing


def test_agent_file_storage_root_controls_coverage_artifacts(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    storage_root = tmp_path / "clean-storage"
    monkeypatch.setenv("AGENT_FILE_STORAGE_ROOT", str(storage_root))
    monkeypatch.setenv("SINGLE_STOCK_COVERAGE_OUTPUT_TIMESTAMP", "20260604-123000")

    result = json.loads(
        tools.create_coverage_run_dir.invoke(
            {
                "company": "Example",
                "ticker": "EXM",
                "market": "US",
                "task_type": "initiation",
            }
        )
    )
    markdown_path = tools.write_markdown_artifact.invoke(
        {
            "markdown": "# Research\n",
            "filename": "company_research.md",
            "subdir": "01_company_research",
            "ticker": "EXM",
            "market": "US",
            "run_dir": result["run_dir"],
        }
    )

    assert file_storage_root() == storage_root
    assert result["run_dir"] == "out/coverage/us-exm/runs/20260604-123000"
    assert (storage_root / result["manifest_path"]).exists()
    assert markdown_path == (
        "out/coverage/us-exm/runs/20260604-123000/"
        "01_company_research/company_research.md"
    )
    assert (storage_root / markdown_path).read_text(encoding="utf-8") == "# Research\n"


def test_dcf_builder_exact_output_dir_writes_task3_artifacts_together(tmp_path):
    from dcf_builder import tools as dcf_tools

    valuation_dir = (
        tmp_path
        / "out"
        / "coverage"
        / "us-exm"
        / "runs"
        / "20260604-123000"
        / "03_valuation"
    )
    dcf_payload = {
        "company": "Example Co",
        "ticker": "EXM",
        "currency": "USD",
        "unit": "millions",
        "projection_periods": 5,
        "market_data": {
            "source": "test source",
            "current_stock_price": 10,
            "debt": 100,
            "cash": 50,
            "shares_outstanding": 100,
            "beta": 1.1,
            "risk_free_rate": 0.03,
            "equity_risk_premium": 0.05,
            "pretax_cost_of_debt": 0.04,
            "tax_rate": 0.25,
        },
        "historicals": [
            {
                "year": 2024,
                "revenue": 900,
                "ebit": 108,
                "da": 27,
                "capex": 35,
                "nwc_change": 5,
                "debt": 100,
                "cash": 50,
                "shares_outstanding": 100,
                "source": "test source",
            },
            {
                "year": 2025,
                "revenue": 1000,
                "ebit": 120,
                "da": 30,
                "capex": 40,
                "nwc_change": 6,
                "debt": 100,
                "cash": 50,
                "shares_outstanding": 100,
                "source": "test source",
            },
        ],
        "scenarios": {
            name: {
                "source": "test source",
                "revenue_growth": growth,
                "ebit_margin": margin,
                "tax_rate": [0.25] * 5,
                "da_pct_revenue": [0.03] * 5,
                "capex_pct_revenue": [0.04] * 5,
                "nwc_pct_delta_revenue": [0.06] * 5,
                "wacc": [wacc] * 5,
                "terminal_growth": [terminal_growth] * 5,
            }
            for name, growth, margin, wacc, terminal_growth in [
                ("Bear", [0.02] * 5, [0.10] * 5, 0.11, 0.01),
                ("Base", [0.05] * 5, [0.12] * 5, 0.10, 0.02),
                ("Bull", [0.08] * 5, [0.15] * 5, 0.09, 0.025),
            ]
        },
    }
    comps_payload = {
        "companies": [
            {
                "company": "Example Co",
                "ticker": "EXM",
                "revenue": 1000,
                "revenue_growth": 0.05,
                "ebitda": 150,
                "net_income": 80,
                "market_cap": 1000,
                "enterprise_value": 1050,
                "source": "test source",
            },
            {
                "company": "Peer Co",
                "ticker": "PEER",
                "revenue": 1200,
                "revenue_growth": 0.04,
                "ebitda": 180,
                "net_income": 90,
                "market_cap": 1100,
                "enterprise_value": 1150,
                "source": "test source",
            },
        ]
    }

    dcf_path = Path(
        dcf_tools.build_dcf_model.invoke(
            {
                "dcf_json": json.dumps(dcf_payload),
                "output_dir": str(valuation_dir),
                "exact_output_dir": True,
            }
        )
    )
    comps_path = Path(
        dcf_tools.build_comps_excel.invoke(
            {
                "data_json": json.dumps(comps_payload),
                "sector": "example",
                "output_dir": str(valuation_dir),
                "exact_output_dir": True,
            }
        )
    )
    validation = json.loads(
        dcf_tools.validate_dcf_model.invoke({"excel_path": str(dcf_path)})
    )

    assert dcf_path == valuation_dir / "dcf_model.xlsx"
    assert comps_path == valuation_dir / "comps.xlsx"
    assert Path(validation["validation_path"]) == valuation_dir / "validation.json"
    assert dcf_path.exists()
    assert comps_path.exists()
    assert not (valuation_dir / "20260604-123000").exists()


def test_statement_json_tools_validate_write_and_read_context(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "coverage" / "us-exm" / "runs" / "20260604-120000"
    (run_dir / "01_company_research").mkdir(parents=True)
    (run_dir / "02_financial_model").mkdir(parents=True)
    (run_dir / "01_company_research" / "company_research.md").write_text(
        "# Example Co\n",
        encoding="utf-8",
    )
    (run_dir / "01_company_research" / "business_driver_map.json").write_text(
        '{"company": "Example Co", "ticker": "EXM"}',
        encoding="utf-8",
    )
    (run_dir / "01_company_research" / "source_log.json").write_text(
        '{"sources": ["annual report"]}',
        encoding="utf-8",
    )
    (run_dir / "02_financial_model" / "financial_facts.json").write_text(
        '{"historicals": [{"period": "FY2025A"}]}',
        encoding="utf-8",
    )
    (run_dir / "02_financial_model" / "task2_context_packet.json").write_text(
        '{"currency": "USD"}',
        encoding="utf-8",
    )

    context = json.loads(
        tools.read_statement_context.invoke(
            {"statement_type": "income_statement", "run_dir": str(run_dir)}
        )
    )
    assert context["statement_type"] == "income_statement"
    assert context["missing_artifacts"] == []
    assert "net_income" in context["canonical_row_keys"]
    assert "company_research" in context["artifacts"]
    assert "company_research" not in json.dumps(context.get("historical_facts", []))
    assert "annual report" not in json.dumps(context.get("artifacts", {}))
    assert len(json.dumps(context, ensure_ascii=False)) < 10000

    validation = json.loads(
        tools.validate_income_statement_json.invoke({"run_dir": str(run_dir)})
    )
    assert validation["status"] == "PASS"
    income_result = json.loads(
        tools.write_income_statement_json.invoke(
            {
                "ticker": "EXM",
                "market": "US",
                "run_dir": str(run_dir),
            }
        )
    )
    assert income_result["status"] == "OK"
    assert (run_dir / "02_financial_model" / "income_statement_spec.json").exists()
    assert (run_dir / "02_financial_model" / "revenue_build_spec.json").exists()

    assert (
        json.loads(tools.validate_balance_sheet_json.invoke({"run_dir": str(run_dir)}))[
            "status"
        ]
        == "PASS"
    )
    tools.write_balance_sheet_json.invoke(
        {
            "ticker": "EXM",
            "market": "US",
            "run_dir": str(run_dir),
        }
    )
    assert (run_dir / "02_financial_model" / "balance_sheet_spec.json").exists()

    assert (
        json.loads(tools.validate_cash_flow_json.invoke({"run_dir": str(run_dir)}))[
            "status"
        ]
        == "PASS"
    )
    tools.write_cash_flow_json.invoke(
        {
            "ticker": "EXM",
            "market": "US",
            "run_dir": str(run_dir),
        }
    )
    assert (run_dir / "02_financial_model" / "cash_flow_statement_spec.json").exists()


def test_statement_tools_handle_nested_financial_facts_modeler_keys(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "coverage" / "a-share-nested" / "runs" / "20260604-120000"
    facts = {
        "company": "Nested Co",
        "ticker": "NST",
        "market": "A-share",
        "currency": "CNY",
        "unit": "亿元",
        "historicals": [
            {
                "period": "FY2025",
                "year": 2025,
                "income_statement": {
                    "revenue": 1581.79,
                    "gross_margin": 0.2906,
                    "ebit": 325.01,
                    "ebitda": 460.05,
                    "income_tax": 57.28,
                    "net_income": 266.09,
                    "source": "IS source",
                },
                "balance_sheet": {
                    "total_assets": 2362.69,
                    "total_liabilities": 988.62,
                    "total_equity": 1374.07,
                    "cash": 317.98,
                    "short_term_debt": 16.39,
                    "long_term_debt": 198.22,
                    "total_debt": 214.61,
                    "PP&E": 1047.26,
                    "source": "BS source",
                },
                "cash_flow": {
                    "operating_cf": 352.69,
                    "capex": 170.63,
                    "DA": 135.05,
                    "dividends_paid": 217.82,
                    "investing_cf": -95.23,
                    "financing_cf": -246.04,
                    "source": "CF source",
                },
            }
        ],
        "sources": ["financial_facts_modeler"],
        "unsourced_items": ["retained earnings detail"],
    }
    _write_model_source_files(run_dir / "02_financial_model", facts)

    for tool_call in (
        tools.write_income_statement_json,
        tools.write_balance_sheet_json,
        tools.write_cash_flow_json,
    ):
        result = json.loads(
            tool_call.invoke(
                {
                    "ticker": "NST",
                    "market": "A-share",
                    "run_dir": str(run_dir),
                }
            )
        )
        assert result["status"] == "OK"

    def historical_value(path, key):
        payload = json.loads(path.read_text(encoding="utf-8"))
        for item in payload["historical_inputs"]:
            if item["canonical_key"] == key:
                return item["value"]
        raise AssertionError(key)

    model_dir = run_dir / "02_financial_model"
    assert (
        historical_value(model_dir / "income_statement_spec.json", "revenue_total")
        == 1581.79
    )
    assert (
        historical_value(model_dir / "balance_sheet_spec.json", "total_assets")
        == 2362.69
    )
    assert (
        historical_value(model_dir / "cash_flow_statement_spec.json", "cfo_total")
        == 352.69
    )


def test_resolve_task2_handoff_reuses_existing_task1_run(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)

    older = tmp_path / "out" / "coverage" / "us-exm" / "runs" / "20260604-110000"
    newer = tmp_path / "out" / "coverage" / "us-exm" / "runs" / "20260604-120000"
    _write_task1_fixture(older)
    _write_task1_fixture(newer)

    explicit = json.loads(
        tools.resolve_task2_handoff.invoke(
            {
                "ticker": "EXM",
                "market": "US",
                "task1_dir": "out/coverage/us-exm/runs/20260604-120000/01_company_research",
            }
        )
    )
    assert explicit["status"] == "OK"
    assert explicit["created_new_run"] is False
    assert explicit["run_dir"] == "out/coverage/us-exm/runs/20260604-120000"
    assert explicit["missing_artifacts"] == []

    latest = json.loads(
        tools.resolve_task2_handoff.invoke({"ticker": "EXM", "market": "US"})
    )
    assert latest["status"] == "OK"
    assert latest["source"] == "latest_task1_run"
    assert latest["run_dir"] == "out/coverage/us-exm/runs/20260604-120000"


def test_verify_task2_artifacts_blocks_wrong_root_statement(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    monkeypatch.setattr(
        tools, "_project_root", lambda: tmp_path / "single-stock-coverage"
    )

    run_dir = tmp_path / "out" / "coverage" / "us-exm" / "runs" / "20260604-120000"
    _write_task1_fixture(run_dir)
    wrong_model_dir = (
        tmp_path
        / "single-stock-coverage"
        / "out"
        / "coverage"
        / "us-exm"
        / "runs"
        / "20260604-120000"
        / "02_financial_model"
    )
    wrong_model_dir.mkdir(parents=True)
    (wrong_model_dir / "balance_sheet_spec.json").write_text(
        json.dumps(_statement_payload("balance_sheet")),
        encoding="utf-8",
    )

    verification = json.loads(
        tools.verify_task2_artifacts.invoke(
            {
                "run_dir": "out/coverage/us-exm/runs/20260604-120000",
                "stage": "statements",
            }
        )
    )
    assert verification["status"] == "FAIL"
    assert any(
        item["category"] == "Wrong Artifact Root" for item in verification["critical"]
    )
    assert verification["wrong_root_artifacts"][0]["canonical_path"].startswith(
        "out/coverage/us-exm/runs/20260604-120000"
    )


def test_write_task2_model_audit_is_compact_and_omits_runtime_logs(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "out" / "coverage" / "us-exm" / "runs" / "20260604-120000"
    run_dir.mkdir(parents=True)

    result = json.loads(
        tools.write_task2_model_audit.invoke(
            {
                "run_dir": str(run_dir),
                "audit_json": json.dumps(
                    {
                        "status": "BLOCKED",
                        "critical": [
                            {
                                "category": "Missing Statement JSON",
                                "issue": "SystemMessage inputs.messages should not leak",
                            }
                        ],
                        "warnings": [{"category": "Source Gap", "issue": "ok"}],
                        "artifacts": {"balance_sheet_spec.json": "missing"},
                        "next_steps": ["Retry bs_modeler with reduced scope"],
                        "task3_handoff_ready": False,
                    }
                ),
            }
        )
    )
    assert result["status"] == "OK"
    audit_text = (tmp_path / result["model_audit_path"]).read_text(encoding="utf-8")
    assert "SystemMessage" not in audit_text
    assert "inputs.messages" not in audit_text
    assert "Missing Statement JSON" in audit_text
    assert len(audit_text) < 5000


def test_statement_validation_allows_supplemental_parent_canonical_key():
    payload = _statement_payload("cash_flow")
    payload["historical_inputs"].append(
        {
            "period": "FY2023A",
            "canonical_key": "asset_impairment",
            "parent_canonical_key": "cfo_total",
            "value": 5,
            "source": "FY2023 annual report note",
            "currency": "USD",
            "unit": "millions",
        }
    )

    validation = tools._validate_statement_payload(payload, "cash_flow")
    assert validation["status"] == "PASS"
    assert not [
        item for item in validation["warnings"] if "asset_impairment" in item["issue"]
    ]


def test_reconcile_statement_specs_writes_pack_and_preserves_warnings(
    monkeypatch, tmp_path
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "coverage" / "us-exm" / "runs" / "20260604-120000"
    payload = _minimal_model_input()
    payload["unsourced"] = ["retained_earnings bridge"]
    _write_model_source_files(run_dir / "02_financial_model", payload)

    for tool_call in (
        tools.write_income_statement_json,
        tools.write_balance_sheet_json,
        tools.write_cash_flow_json,
    ):
        result = json.loads(
            tool_call.invoke(
                {
                    "ticker": "EXM",
                    "market": "US",
                    "run_dir": str(run_dir),
                }
            )
        )
        assert result["status"] == "OK"

    result = json.loads(
        tools.reconcile_statement_specs.invoke(
            {"ticker": "EXM", "market": "US", "run_dir": str(run_dir)}
        )
    )

    pack_path = tmp_path / result["statement_spec_pack_path"]
    assert result["status"] == "PASS"
    assert result["builder_blocked"] is False
    assert result["critical_count"] == 0
    assert result["warning_count"] > 0
    assert pack_path.exists()
    assert json.loads(pack_path.read_text(encoding="utf-8"))["status"] == "PASS"


def test_reconcile_statement_specs_preserves_existing_financial_context(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "out" / "coverage" / "us-exm" / "runs" / "20260604-120000"
    model_dir = run_dir / "02_financial_model"
    model_dir.mkdir(parents=True)
    facts = {
        "company": "Example Co",
        "ticker": "EXM",
        "market": "US",
        "historicals": [{"period": "FY2023A", "year": 2023, "revenue": 123}],
        "sources": ["financial_facts_modeler"],
    }
    context = {"company": "Example Co", "custom_context": "preserve me"}
    (model_dir / "financial_facts.json").write_text(json.dumps(facts), encoding="utf-8")
    (model_dir / "task2_context_packet.json").write_text(
        json.dumps(context),
        encoding="utf-8",
    )

    for tool_call in (
        tools.write_income_statement_json,
        tools.write_balance_sheet_json,
        tools.write_cash_flow_json,
    ):
        result = json.loads(
            tool_call.invoke(
                {
                    "ticker": "EXM",
                    "market": "US",
                    "run_dir": str(run_dir),
                }
            )
        )
        assert result["status"] == "OK"

    result = json.loads(
        tools.reconcile_statement_specs.invoke(
            {"ticker": "EXM", "market": "US", "run_dir": str(run_dir)}
        )
    )
    assert result["status"] == "PASS"
    preserved_facts = json.loads((model_dir / "financial_facts.json").read_text())
    preserved_context = json.loads(
        (model_dir / "task2_context_packet.json").read_text()
    )
    assert preserved_facts["sources"] == ["financial_facts_modeler"]
    assert preserved_facts["historicals"][0]["revenue"] == 123
    assert preserved_context["custom_context"] == "preserve me"
    assert preserved_context["reconciliation_status"] == "PASS"


def test_integrated_three_statement_builder_and_validator(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "coverage" / "us-exm" / "runs" / "20260604-120000"
    run_dir.mkdir(parents=True)
    _write_model_source_files(run_dir / "02_financial_model", _minimal_model_input())

    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    assert result["status"] == "OK"
    assert workbook_path.exists()
    assert result["cached_formula_count"] > 0
    assert result["row_map"]["income_statement"]["revenue_total"] == 8
    assert result["row_map"]["balance_sheet"]["cash_and_equivalents"] == 8
    assert result["row_map"]["cash_flow"]["ending_cash"] == 25
    assert result["period_columns"] == {
        "FY2023A": "C",
        "FY2024E": "D",
        "FY2025E": "E",
    }

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    assert wb.sheetnames == list(tools.THREE_STATEMENT_TABS)
    assert all(name in set(wb.defined_names) for name in tools.REQUIRED_MODEL_NAMES)
    assert wb["DCF Inputs"]["D8"].value.startswith("=")
    assert wb["Checks"]["D9"].value.startswith("=")
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    revenue_total_row = result["row_map"]["revenue_build"]["revenue_total"]
    assert (
        wb_values["Revenue Build"].cell(row=revenue_total_row, column=4).value == 1050
    )
    assert wb_values["Income Statement"]["D8"].value == 1050
    assert wb_values["Cash Flow Statement"]["D25"].value is not None

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(workbook_path),
                "row_map_json": json.dumps(result["row_map"]),
            }
        )
    )
    assert validation["status"] == "PASS"
    assert validation["critical_count"] == 0


def test_integrated_builder_handles_nested_a_share_facts_and_three_year_forecast(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "out" / "coverage" / "a-share-300502.sz" / "runs" / "nested"
    model_dir = run_dir / "02_financial_model"
    payload = _nested_a_share_model_input()
    _write_model_source_files(model_dir, payload)

    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    assert result["status"] == "OK"
    assert result["cached_formula_count"] > 0
    assert result["period_columns"] == {
        "FY2022A": "C",
        "FY2023A": "D",
        "FY2024A": "E",
        "FY2025A": "F",
        "FY2026E": "G",
        "FY2027E": "H",
        "FY2028E": "I",
    }

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    assert wb["Cover"]["B6"].value == "新易盛"
    assert wb["Cover"]["B7"].value == "300502.SZ"
    assert wb["Cover"]["B8"].value == "A-share"
    assert wb["Cover"]["A3"].value == "Currency: CNY | Unit: 亿元"
    assert [wb["Checks"].cell(row=5, column=col).value for col in range(3, 10)] == [
        "FY2022A",
        "FY2023A",
        "FY2024A",
        "FY2025A",
        "FY2026E",
        "FY2027E",
        "FY2028E",
    ]
    assert [
        wb["Income Statement"].cell(row=8, column=col).value for col in range(3, 7)
    ] == [
        33.1057,
        30.9761,
        86.4683,
        248.4185,
    ]
    assert [
        wb["Assumptions"].cell(row=40, column=col).value for col in range(7, 10)
    ] == [190, 180, 160]
    assert [
        wb["Assumptions"].cell(row=42, column=col).value for col in range(7, 10)
    ] == [160, 290, 420]
    assert [
        wb["Assumptions"].cell(row=44, column=col).value for col in range(7, 10)
    ] == [30, 40, 60]
    assert [
        wb["Revenue Build"].cell(row=8, column=1).value,
        wb["Revenue Build"].cell(row=13, column=1).value,
        wb["Revenue Build"].cell(row=18, column=1).value,
    ] == [
        "800G Revenue Revenue",
        "1.6T Revenue Revenue",
        "Other Revenue Revenue",
    ]
    revenue_total_row = result["row_map"]["revenue_build"]["revenue_total"]
    assert (
        wb["Revenue Build"].cell(row=revenue_total_row, column=7).value
        == "=SUM(G8,G13,G18,G23)"
    )
    assert (
        wb["Revenue Build"].cell(row=revenue_total_row, column=8).value
        == "=SUM(H8,H13,H18,H23)"
    )
    assert (
        wb["Revenue Build"].cell(row=revenue_total_row, column=9).value
        == "=SUM(I8,I13,I18,I23)"
    )
    assert (
        wb["Debt & Interest"]["F16"].value
        == '=IF(F12>MAX(F11*5,0.01),"CHECK SHORT-TERM RAW > TOTAL DEBT","OK")'
    )
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    assert "Integrated 3-Statement Model" in wb_values["Cover"]["A1"].value
    assert wb_values["Cover"]["B17"].value == "PASS"
    assert [
        wb_values["Cover"].cell(row=23, column=col).value for col in range(7, 10)
    ] == [
        380,
        510,
        640,
    ]
    assert wb_values["Cover"]["A34"].value == "Workbook Navigation"
    assert wb_values["Cover"]["A50"].value == "Formatting Legend"
    assert [
        wb_values["Revenue Build"].cell(row=revenue_total_row, column=col).value
        for col in range(7, 10)
    ] == [
        380,
        510,
        640,
    ]
    assert [
        wb_values["Income Statement"].cell(row=8, column=col).value
        for col in range(7, 10)
    ] == [
        380,
        510,
        640,
    ]
    assert wb_values["Balance Sheet"]["F20"].value == 0.1493887
    assert wb_values["Debt & Interest"]["F11"].value == 0.1493887
    assert wb_values["Debt & Interest"]["F12"].value == 15.5258
    assert (
        wb_values["Debt & Interest"]["F16"].value == "CHECK SHORT-TERM RAW > TOTAL DEBT"
    )
    assert wb_values["DCF Inputs"]["F14"].value == 0.1493887
    assert wb["DCF Inputs"]["F10"].value.startswith("=IF(")
    assert wb_values["Income Statement"]["G22"].value is not None
    assert wb_values["Checks"]["B4"].value == 0

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(workbook_path),
                "row_map_json": json.dumps(result["row_map"]),
            }
        )
    )
    assert validation["status"] == "PASS"
    assert validation["critical_count"] == 0
    assert "Debt Data Quality" in {item["category"] for item in validation["warnings"]}


def test_integrated_builder_uses_standalone_spec_driven_components(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "coverage" / "us-platform" / "runs" / "dynamic"
    model_dir = run_dir / "02_financial_model"
    payload = _minimal_model_input()
    payload["revenue_build_spec"] = {
        "statement_type": "revenue_build",
        "segments": [
            {
                "segment_id": "subscription",
                "display_name": "Subscription",
                "driver_type": "seat_count_x_arpu",
                "historical": {
                    "FY2023": {"revenue": 600, "cost": 210, "gross_profit": 390}
                },
                "forecast_revenue": [
                    {"period": "FY2024E", "value": 660},
                    {"period": "FY2025E", "value": 720},
                ],
            },
            {
                "segment_id": "services",
                "display_name": "Services",
                "driver_type": "utilization_x_rate",
                "historical": {
                    "FY2023": {"revenue": 250, "cost": 150, "gross_profit": 100}
                },
                "forecast_revenue": [
                    {"period": "FY2024E", "value": 260},
                    {"period": "FY2025E", "value": 275},
                ],
            },
            {
                "segment_id": "marketplace",
                "display_name": "Marketplace",
                "driver_type": "gmv_x_take_rate",
                "historical": {
                    "FY2023": {"revenue": 150, "cost": 40, "gross_profit": 110}
                },
                "forecast_revenue": [
                    {"period": "FY2024E", "value": 130},
                    {"period": "FY2025E", "value": 155},
                ],
            },
        ],
        "total_revenue_reconciliation": {
            "forecast": [
                {"period": "FY2024E", "total": 1050},
                {"period": "FY2025E", "total": 1150},
            ]
        },
    }
    _write_model_source_files(model_dir, payload)

    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    labels = [
        wb["Revenue Build"].cell(row=row, column=1).value
        for row in range(1, wb["Revenue Build"].max_row + 1)
    ]
    assert "Subscription Revenue" in labels
    assert "Services Revenue" in labels
    assert "Marketplace Revenue" in labels
    assert "800G Revenue" not in labels
    assert "1.6T Revenue" not in labels
    revenue_total_row = result["row_map"]["revenue_build"]["revenue_total"]
    assert (
        wb["Revenue Build"].cell(row=revenue_total_row, column=4).value
        == "=SUM(D8,D13,D18,D23)"
    )

    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    assert (
        wb_values["Revenue Build"].cell(row=revenue_total_row, column=4).value == 1050
    )
    assert wb_values["Income Statement"]["D8"].value == 1050
    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(workbook_path),
                "row_map_json": json.dumps(result["row_map"]),
            }
        )
    )
    assert validation["status"] == "PASS"


def test_integrated_builder_consumes_five_component_spec_and_excludes_interim(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "out" / "coverage" / "a-share-601225.sh" / "runs" / "dynamic"
    model_dir = run_dir / "02_financial_model"
    payload = {
        "company": "陕西煤业股份有限公司",
        "ticker": "601225.SH",
        "market": "A-share",
        "currency": "CNY",
        "unit": "million_CNY",
        "fiscal_year_end": "12-31",
        "historicals": [
            {
                "period": "FY2024",
                "year": 2024,
                "revenue": 1000,
                "gross_profit": 360,
                "operating_expenses": 80,
                "da": 50,
                "ebit": 230,
                "pretax_income": 220,
                "tax_expense": 40,
                "net_income": 180,
                "cash": 120,
                "debt": 90,
                "total_assets": 900,
                "total_liabilities": 400,
                "total_equity": 500,
                "retained_earnings": 300,
                "shares": 100,
                "source": "FY2024 annual report",
            },
            {
                "period": "FY2025",
                "year": 2025,
                "revenue": 1100,
                "gross_profit": 410,
                "operating_expenses": 90,
                "da": 55,
                "ebit": 265,
                "pretax_income": 250,
                "tax_expense": 45,
                "net_income": 205,
                "cash": 150,
                "debt": 80,
                "total_assets": 950,
                "total_liabilities": 410,
                "total_equity": 540,
                "retained_earnings": 340,
                "shares": 100,
                "source": "FY2025 annual report",
            },
            {
                "period": "Q1-2026",
                "year": 2026,
                "revenue": 260,
                "gross_profit": 90,
                "net_income": 40,
                "source": "Q1-2026 interim report",
            },
        ],
        "task2_context_packet": {
            "company_metadata": {
                "company": "陕西煤业股份有限公司",
                "ticker": "601225.SH",
                "market": "A-share",
                "currency": "CNY",
                "reporting_unit": "million_CNY",
                "fiscal_year_end": "12-31",
            },
            "model_horizon": {
                "forecast_years": ["FY2026E", "FY2027E", "FY2028E"],
                "latest_interim": "Q1-2026",
            },
        },
        "revenue_build_spec": {
            "statement_type": "revenue_build",
            "segments": [
                {
                    "segment_id": "raw",
                    "display_name": "Raw Product",
                    "historical": {
                        "FY2024": {"revenue": 350, "cost": 180},
                        "FY2025": {"revenue": 380, "cost": 195},
                    },
                },
                {
                    "segment_id": "trading",
                    "display_name": "Trading Product",
                    "historical": {
                        "FY2024": {"revenue": 280, "cost": 270},
                        "FY2025": {"revenue": 300, "cost": 289},
                    },
                },
                {
                    "segment_id": "processed",
                    "display_name": "Processed Product",
                    "historical": {
                        "FY2024": {"revenue": 180, "cost": 90},
                        "FY2025": {"revenue": 200, "cost": 100},
                    },
                },
                {
                    "segment_id": "power",
                    "display_name": "Power Service",
                    "historical": {
                        "FY2024": {"revenue": 120, "cost": 95},
                        "FY2025": {"revenue": 140, "cost": 110},
                    },
                },
                {
                    "segment_id": "other",
                    "display_name": "Other Operations",
                    "historical": {
                        "FY2024": {"revenue": 70, "cost": 50},
                        "FY2025": {"revenue": 80, "cost": 60},
                    },
                },
            ],
        },
    }
    _write_model_source_files(model_dir, payload)

    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    assert result["period_columns"] == {
        "FY2024A": "C",
        "FY2025A": "D",
        "FY2026E": "E",
        "FY2027E": "F",
        "FY2028E": "G",
    }
    assert "FY2026A" not in result["period_columns"]
    labels = [
        wb["Revenue Build"].cell(row=row, column=1).value
        for row in range(1, wb["Revenue Build"].max_row + 1)
    ]
    for expected_label in (
        "Raw Product Revenue",
        "Trading Product Revenue",
        "Processed Product Revenue",
        "Power Service Revenue",
        "Other Operations Revenue",
    ):
        assert expected_label in labels
    assert "800G Revenue" not in labels
    assert "1.6T Revenue" not in labels
    assert any(
        wb["Sources"].cell(row=row, column=1).value == "Q1-2026"
        and wb["Sources"].cell(row=row, column=4).value
        == "Excluded from annual model columns"
        for row in range(1, wb["Sources"].max_row + 1)
    )

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(workbook_path),
                "row_map_json": json.dumps(result["row_map"]),
            }
        )
    )
    assert validation["status"] == "PASS"


def test_integrated_validator_rejects_placeholder_shell_workbook(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "out" / "coverage" / "a-share-300502.sz" / "runs" / "bad"
    model_dir = run_dir / "02_financial_model"
    payload = _nested_a_share_model_input()
    _write_model_source_files(model_dir, payload)

    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    for ws in wb.worksheets:
        ws.delete_cols(9)
    wb["Cover"]["A3"] = "Currency: USD | Unit: millions"
    wb["Cover"]["B6"] = "{'short_name': '新易盛'}"
    wb["Cover"]["B7"] = "TICKER"
    wb["Assumptions"]["A10"] = "Assumption 10"
    for col in range(3, 7):
        wb["Income Statement"].cell(row=8, column=col, value=0)
    wb.save(workbook_path)

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(workbook_path),
                "row_map_json": json.dumps(result["row_map"]),
            }
        )
    )
    categories = {item["category"] for item in validation["critical"]}
    assert validation["status"] == "FAIL"
    assert "Default Metadata" in categories
    assert "Missing Forecast Period" in categories
    assert "Placeholder Assumption" in categories
    assert (
        "Historical Values Missing" in categories
        or "Historical Value Mismatch" in categories
    )


def test_integrated_validator_rejects_zero_workbook_debt_against_source(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "out" / "coverage" / "a-share-300502.sz" / "runs" / "debt-bad"
    model_dir = run_dir / "02_financial_model"
    payload = _nested_a_share_model_input()
    _write_model_source_files(model_dir, payload)

    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    wb["Balance Sheet"]["F20"] = 0
    wb["Debt & Interest"]["F11"] = 0
    wb.save(workbook_path)

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(workbook_path),
                "row_map_json": json.dumps(result["row_map"]),
            }
        )
    )
    categories = {item["category"] for item in validation["critical"]}
    assert validation["status"] == "FAIL"
    assert "Debt Source Mismatch" in categories
    assert "Debt Schedule Mismatch" in categories


def test_integrated_validator_rejects_missing_formula_cache(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "coverage" / "us-exm" / "runs" / "cacheless"
    run_dir.mkdir(parents=True)
    _write_model_source_files(run_dir / "02_financial_model", _minimal_model_input())
    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    wb.save(workbook_path)

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(workbook_path),
                "row_map_json": json.dumps(result["row_map"]),
            }
        )
    )
    categories = {item["category"] for item in validation["critical"]}
    assert validation["status"] == "FAIL"
    assert "Formula Cache Missing" in categories


def test_integrated_three_statement_validator_flags_missing_tab(monkeypatch, tmp_path):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "coverage" / "us-exm" / "runs" / "20260604-120000"
    run_dir.mkdir(parents=True)
    _write_model_source_files(run_dir / "02_financial_model", _minimal_model_input())
    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path)
    del wb["Checks"]
    wb.save(workbook_path)

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {"excel_path": str(workbook_path)}
        )
    )
    assert validation["status"] == "FAIL"
    assert validation["critical"][0]["category"] == "Missing Required Tab"


def test_integrated_three_statement_validator_flags_hardcode_and_cash_break(
    monkeypatch, tmp_path
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    run_dir = tmp_path / "coverage" / "us-exm" / "runs" / "20260604-120000"
    run_dir.mkdir(parents=True)
    _write_model_source_files(run_dir / "02_financial_model", _minimal_model_input())
    result = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(run_dir)})
    )
    workbook_path = tmp_path / result["workbook_path"]

    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    wb["Income Statement"]["D8"] = 1234
    wb["Balance Sheet"]["D8"] = 555
    wb.save(workbook_path)

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(workbook_path),
                "row_map_json": json.dumps(result["row_map"]),
            }
        )
    )
    categories = {item["category"] for item in validation["critical"]}
    assert validation["status"] == "FAIL"
    assert "Projection Hardcode" in categories
    assert "Cash Tie-Out" in categories


def test_task2_artifact_flow_defaults_to_out_coverage_after_task1_fixture(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)
    monkeypatch.setenv("SINGLE_STOCK_COVERAGE_OUTPUT_TIMESTAMP", "20260607-101500")

    legacy_coverage = tmp_path / "coverage" / "sz-300516.sz" / "runs" / "old-task1"
    legacy_coverage.mkdir(parents=True)
    (legacy_coverage / "run_manifest.json").write_text("{}", encoding="utf-8")

    run = json.loads(
        tools.create_coverage_run_dir.invoke(
            {
                "company": "测试公司",
                "ticker": "300516.SZ",
                "market": "SZ",
                "task_type": "model_update",
                "triggering_event": "Task1 fixture from root coverage regression",
            }
        )
    )
    assert run["run_dir"] == "out/coverage/sz-300516.sz/runs/20260607-101500"
    run_dir = tmp_path / run["run_dir"]

    tools.write_markdown_artifact.invoke(
        {
            "markdown": "# 测试公司\n\nTask1 company research fixture.\n",
            "filename": "company_research.md",
            "subdir": "01_company_research",
            "ticker": "300516.SZ",
            "market": "SZ",
            "run_dir": run["run_dir"],
        }
    )
    tools.write_json_artifact.invoke(
        {
            "data_json": json.dumps(
                {
                    "company": "测试公司",
                    "ticker": "300516.SZ",
                    "drivers": ["revenue growth"],
                }
            ),
            "filename": "business_driver_map.json",
            "subdir": "01_company_research",
            "ticker": "300516.SZ",
            "market": "SZ",
            "run_dir": run["run_dir"],
        }
    )
    tools.write_json_artifact.invoke(
        {
            "data_json": json.dumps({"sources": ["Task1 fixture source"]}),
            "filename": "source_log.json",
            "subdir": "01_company_research",
            "ticker": "300516.SZ",
            "market": "SZ",
            "run_dir": run["run_dir"],
        }
    )
    _write_model_source_files(run_dir / "02_financial_model", _minimal_model_input())

    for tool_call in (
        tools.write_income_statement_json,
        tools.write_balance_sheet_json,
        tools.write_cash_flow_json,
    ):
        result = json.loads(
            tool_call.invoke(
                {
                    "ticker": "300516.SZ",
                    "market": "SZ",
                    "run_dir": run["run_dir"],
                }
            )
        )
        assert result["status"] == "OK"

    pack = json.loads(
        tools.reconcile_statement_specs.invoke(
            {"ticker": "300516.SZ", "market": "SZ", "run_dir": run["run_dir"]}
        )
    )
    assert pack["status"] == "PASS"
    assert pack["statement_spec_pack_path"].startswith(run["run_dir"])
    assert pack["financial_facts_path"].startswith(run["run_dir"])
    assert pack["task2_context_packet_path"].startswith(run["run_dir"])
    facts = json.loads((tmp_path / pack["financial_facts_path"]).read_text())
    context = json.loads((tmp_path / pack["task2_context_packet_path"]).read_text())
    assert facts["historicals"][0]["revenue"] == 1000
    assert "income_statement" in context["canonical_row_keys"]

    build = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": run["run_dir"]})
    )
    assert build["status"] == "OK"
    assert build["workbook_path"].startswith(run["run_dir"])
    assert (tmp_path / build["workbook_path"]).exists()

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {
                "excel_path": str(tmp_path / build["workbook_path"]),
                "row_map_json": json.dumps(build["row_map"]),
            }
        )
    )
    assert validation["status"] == "PASS"

    audit_path = tools.write_markdown_artifact.invoke(
        {
            "markdown": "# Model Audit\n\nOverall: Clean\n",
            "filename": "model_audit.md",
            "subdir": "02_financial_model",
            "ticker": "300516.SZ",
            "market": "SZ",
            "run_dir": run["run_dir"],
        }
    )
    manifest_path = tools.update_run_manifest.invoke(
        {
            "patch_json": json.dumps(
                {
                    "subagents_called": [
                        "is_modeler",
                        "bs_modeler",
                        "cf_modeler",
                        "workbook_builder",
                    ],
                    "output_artifacts": [
                        pack["statement_spec_pack_path"],
                        build["workbook_path"],
                        audit_path,
                    ],
                    "task3_handoff_ready": True,
                }
            ),
            "ticker": "300516.SZ",
            "market": "SZ",
            "run_dir": run["run_dir"],
        }
    )

    assert manifest_path.startswith("out/coverage/")
    assert audit_path.startswith(run["run_dir"])
    assert (run_dir / "02_financial_model" / "integrated_model.xlsx").exists()
    assert (tmp_path / "coverage" / "sz-300516.sz" / "runs" / "old-task1").exists()
    assert not (legacy_coverage / "out").exists()


def test_model_update_executor_tool_copies_prior_workbook_and_validates(
    monkeypatch,
    tmp_path,
):
    _clear_env(monkeypatch)
    tools._ACTIVE_RUNS.clear()
    monkeypatch.setattr(tools, "_workspace_root", lambda: tmp_path)

    prior_run = tmp_path / "out" / "coverage" / "us-exm" / "runs" / "prior"
    prior_run.mkdir(parents=True)
    _write_model_source_files(prior_run / "02_financial_model", _minimal_model_input())
    prior_build = json.loads(
        tools.build_integrated_three_statement_model.invoke({"run_dir": str(prior_run)})
    )
    prior_workbook = tmp_path / prior_build["workbook_path"]
    assert prior_workbook.exists()

    run_dir = tmp_path / "out" / "coverage" / "us-exm" / "runs" / "update"
    _write_model_source_files(run_dir / "02_financial_model", _minimal_model_input())
    for tool_call in (
        tools.write_income_statement_json,
        tools.write_balance_sheet_json,
        tools.write_cash_flow_json,
    ):
        result = json.loads(
            tool_call.invoke(
                {
                    "ticker": "EXM",
                    "market": "US",
                    "run_dir": str(run_dir),
                }
            )
        )
        assert result["status"] == "OK"

    pack = json.loads(
        tools.reconcile_statement_specs.invoke(
            {"ticker": "EXM", "market": "US", "run_dir": str(run_dir)}
        )
    )
    assert pack["status"] == "PASS"
    update = json.loads(
        tools.update_integrated_three_statement_model.invoke(
            {
                "prior_workbook_path": str(prior_workbook),
                "run_dir": str(run_dir),
                "update_scope_json": json.dumps({"trigger": "earnings_update"}),
            }
        )
    )

    assert update["status"] == "OK"
    assert update["workbook_path"].endswith(
        "out/coverage/us-exm/runs/update/02_financial_model/integrated_model.xlsx"
    )
    assert "Income Statement!C8" in update["updated_cells"]
    assert (tmp_path / update["workbook_path"]).exists()

    validation = json.loads(
        tools.validate_integrated_three_statement_model.invoke(
            {"excel_path": str(tmp_path / update["workbook_path"])}
        )
    )
    assert validation["status"] == "PASS"


def test_agent_registry_exposes_task2_parallel_statement_context():
    registry = load_agent_registry()

    assert mcp_tool_group_names(registry) == (
        "mcp_tools",
        "ifind_mcp_tools",
        "mx_ds_mcp_tools",
    )
    all_server_names = ["ifind-stock", "ifind-news", "mx-ds-mcp"]
    assert mcp_tool_group_server_names(
        registry,
        "ifind_mcp_tools",
        all_server_names,
    ) == {"ifind-stock", "ifind-news"}
    assert mcp_tool_group_server_names(
        registry,
        "mx_ds_mcp_tools",
        all_server_names,
    ) == {"mx-ds-mcp"}

    root = describe_agent(registry, "single_stock_coverage")
    assert root["tool_groups"] == ["coverage_orchestration_tools"]
    assert root["tools"]["coverage_orchestration_tools"] == [
        "create_coverage_run_dir",
        "update_run_manifest",
        "write_coverage_state",
    ]
    assert root["excluded_builtin_tools"] == [
        "write_file",
        "edit_file",
        "execute",
    ]

    task1 = describe_agent(registry, "task1_company_researcher")
    assert task1["tool_groups"] == [
        "ifind_mcp_tools",
        "mx_ds_mcp_tools",
        "coverage_artifact_tools",
    ]
    assert task1["tools"]["ifind_mcp_tools"] == ["<runtime MCP tools from ifind-*>"]
    assert task1["tools"]["mx_ds_mcp_tools"] == ["<runtime MCP tools from mx-ds-mcp>"]
    assert task1["tools"]["coverage_artifact_tools"] == [
        "create_coverage_run_dir",
        "write_markdown_artifact",
        "write_json_artifact",
        "update_run_manifest",
        "write_coverage_state",
    ]
    assert task1["skills"] == {"single_stock_coverage": ["company-research"]}

    task2 = describe_agent(registry, "task2_financial_modeler")
    assert task2["parent"] == "single_stock_coverage"
    assert task2["level"] == 1
    assert task2["tool_groups"] == [
        "task2_check_tools",
        "run_manifest_tools",
    ]
    assert task2["tools"]["task2_check_tools"] == [
        "resolve_task2_handoff",
        "verify_task2_artifacts",
        "reconcile_statement_specs",
        "write_task2_model_audit",
    ]
    assert task2["tools"]["run_manifest_tools"] == ["update_run_manifest"]
    assert task2["skills"] == {
        "single_stock_coverage": [
            "model-update",
            "statement-reconciliation-checks",
        ]
    }
    assert task2["subagents"] == [
        "financial_facts_modeler",
        "is_modeler",
        "bs_modeler",
        "cf_modeler",
        "model_update_executor",
        "workbook_builder",
    ]
    assert not agent_uses_tool_group(
        registry, "task2_financial_modeler", "mcp_tools", recursive=False
    )
    assert not agent_uses_tool_group(registry, "task2_financial_modeler", "mcp_tools")
    assert agent_uses_tool_group(
        registry,
        "task2_financial_modeler",
        "ifind_mcp_tools",
    )
    assert agent_uses_tool_group(
        registry,
        "task2_financial_modeler",
        "mx_ds_mcp_tools",
    )
    assert not agent_uses_tool_group(
        registry, "task2_financial_modeler", "coverage_artifact_tools"
    )

    financial_facts_modeler = describe_agent(registry, "financial_facts_modeler")
    assert financial_facts_modeler["parent"] == "task2_financial_modeler"
    assert financial_facts_modeler["tool_groups"] == [
        "ifind_mcp_tools",
        "mx_ds_mcp_tools",
        "task2_financial_fact_artifact_tools",
    ]
    assert financial_facts_modeler["tools"]["task2_financial_fact_artifact_tools"] == [
        "write_json_artifact"
    ]
    assert financial_facts_modeler["excluded_builtin_tools"] == ["task"]
    assert financial_facts_modeler["skills"] == {
        "single_stock_coverage": ["financial-data-normalization", "model-update"]
    }

    is_modeler = describe_agent(registry, "is_modeler")
    assert is_modeler["parent"] == "task2_financial_modeler"
    assert is_modeler["level"] == 2
    assert is_modeler["tool_groups"] == ["statement_modeling_tools"]
    assert is_modeler["tools"]["statement_modeling_tools"] == [
        "read_statement_context",
        "validate_income_statement_json",
        "write_income_statement_json",
        "validate_balance_sheet_json",
        "write_balance_sheet_json",
        "validate_cash_flow_json",
        "write_cash_flow_json",
    ]
    assert is_modeler["excluded_builtin_tools"] == ["task"]
    assert is_modeler["skills"] == {
        "single_stock_coverage": [
            "financial-data-normalization",
            "income-statement-model",
            "statement-json-checks",
        ]
    }
    assert "02_financial_model/income_statement_spec.json" in is_modeler["outputs"]

    bs_modeler = describe_agent(registry, "bs_modeler")
    assert bs_modeler["parent"] == "task2_financial_modeler"
    assert bs_modeler["level"] == 2
    assert bs_modeler["tool_groups"] == ["statement_modeling_tools"]
    assert bs_modeler["excluded_builtin_tools"] == ["task"]
    assert bs_modeler["skills"] == {
        "single_stock_coverage": [
            "financial-data-normalization",
            "balance-sheet-model",
            "statement-json-checks",
        ]
    }
    assert bs_modeler["outputs"] == ["02_financial_model/balance_sheet_spec.json"]
    assert not agent_uses_tool_group(
        registry, "bs_modeler", "mcp_tools", recursive=False
    )
    assert not agent_uses_tool_group(
        registry,
        "bs_modeler",
        "ifind_mcp_tools",
        recursive=False,
    )
    assert not agent_uses_tool_group(
        registry,
        "bs_modeler",
        "mx_ds_mcp_tools",
        recursive=False,
    )

    cf_modeler = describe_agent(registry, "cf_modeler")
    assert cf_modeler["parent"] == "task2_financial_modeler"
    assert cf_modeler["level"] == 2
    assert cf_modeler["tool_groups"] == ["statement_modeling_tools"]
    assert cf_modeler["excluded_builtin_tools"] == ["task"]
    assert cf_modeler["skills"] == {
        "single_stock_coverage": [
            "financial-data-normalization",
            "cash-flow-model",
            "statement-json-checks",
        ]
    }
    assert cf_modeler["outputs"] == ["02_financial_model/cash_flow_statement_spec.json"]

    workbook_builder = describe_agent(registry, "workbook_builder")
    assert workbook_builder["parent"] == "task2_financial_modeler"
    assert workbook_builder["tool_groups"] == [
        "workbook_authoring_tools",
        "task2_audit_artifact_tools",
    ]
    assert workbook_builder["tools"]["workbook_authoring_tools"] == [
        "build_integrated_three_statement_model",
        "validate_integrated_three_statement_model",
    ]
    assert workbook_builder["tools"]["task2_audit_artifact_tools"] == [
        "write_markdown_artifact"
    ]
    assert workbook_builder["excluded_builtin_tools"] == ["task"]
    assert workbook_builder["skills"] == {
        "single_stock_coverage": ["three-statement-model", "xlsx-author", "audit-xls"]
    }

    model_update_executor = describe_agent(registry, "model_update_executor")
    assert model_update_executor["parent"] == "task2_financial_modeler"
    assert model_update_executor["tool_groups"] == [
        "workbook_update_tools",
        "task2_audit_artifact_tools",
    ]
    assert model_update_executor["tools"]["workbook_update_tools"] == [
        "update_integrated_three_statement_model",
        "validate_integrated_three_statement_model",
    ]
    assert model_update_executor["tools"]["task2_audit_artifact_tools"] == [
        "write_markdown_artifact"
    ]
    assert model_update_executor["excluded_builtin_tools"] == ["task"]
    assert not agent_uses_tool_group(
        registry,
        "model_update_executor",
        "mcp_tools",
        recursive=False,
    )
    assert not agent_uses_tool_group(
        registry,
        "model_update_executor",
        "ifind_mcp_tools",
        recursive=False,
    )
    assert not agent_uses_tool_group(
        registry,
        "model_update_executor",
        "mx_ds_mcp_tools",
        recursive=False,
    )
    assert model_update_executor["skills"] == {
        "single_stock_coverage": [
            "model-update",
            "three-statement-model",
            "xlsx-author",
            "audit-xls",
        ]
    }

    valuation = describe_agent(registry, "task3_valuation_analyst")
    assert valuation["tool_groups"] == [
        "ifind_mcp_tools",
        "mx_ds_mcp_tools",
        "coverage_artifact_tools",
    ]
    assert valuation["subagents"] == ["assumption_generator", "dcf_execution"]

    assumption_generator = describe_agent(registry, "assumption_generator")
    assert assumption_generator["tool_groups"] == [
        "ifind_mcp_tools",
        "mx_ds_mcp_tools",
        "coverage_artifact_tools",
    ]


def test_statement_json_tool_groups_resolve_runtime_tools():
    resolver = ToolGroupResolver(mcp_tools=[])

    assert [
        tool.name for tool in resolver.resolve(("coverage_orchestration_tools",))
    ] == [
        "create_coverage_run_dir",
        "update_run_manifest",
        "write_coverage_state",
    ]
    assert [tool.name for tool in resolver.resolve(("statement_modeling_tools",))] == [
        "read_statement_context",
        "validate_income_statement_json",
        "write_income_statement_json",
        "validate_balance_sheet_json",
        "write_balance_sheet_json",
        "validate_cash_flow_json",
        "write_cash_flow_json",
    ]
    assert [tool.name for tool in resolver.resolve(("task2_check_tools",))] == [
        "resolve_task2_handoff",
        "verify_task2_artifacts",
        "reconcile_statement_specs",
        "write_task2_model_audit",
    ]
    assert [tool.name for tool in resolver.resolve(("run_manifest_tools",))] == [
        "update_run_manifest",
    ]
    assert [
        tool.name for tool in resolver.resolve(("task2_financial_fact_artifact_tools",))
    ] == [
        "write_json_artifact",
    ]
    assert [
        tool.name for tool in resolver.resolve(("task2_audit_artifact_tools",))
    ] == [
        "write_markdown_artifact",
    ]
    assert [tool.name for tool in resolver.resolve(("workbook_authoring_tools",))] == [
        "build_integrated_three_statement_model",
        "validate_integrated_three_statement_model",
    ]
    assert [tool.name for tool in resolver.resolve(("workbook_update_tools",))] == [
        "update_integrated_three_statement_model",
        "validate_integrated_three_statement_model",
    ]
    resolved_tools = {
        tool.name: tool
        for tool in resolver.resolve(
            (
                "statement_modeling_tools",
                "workbook_authoring_tools",
                "workbook_update_tools",
            )
        )
    }
    for name in (
        "validate_income_statement_json",
        "write_income_statement_json",
        "validate_balance_sheet_json",
        "write_balance_sheet_json",
        "validate_cash_flow_json",
        "write_cash_flow_json",
    ):
        assert "statement_json" not in resolved_tools[name].args
    assert (
        "model_input_json"
        not in resolved_tools["build_integrated_three_statement_model"].args
    )
    assert (
        "model_input_json"
        not in resolved_tools["update_integrated_three_statement_model"].args
    )
    assert (
        "statement_spec_pack_json"
        not in resolved_tools["update_integrated_three_statement_model"].args
    )


def test_mcp_tool_groups_resolve_runtime_tools_by_provider():
    ifind_tool = SimpleNamespace(name="ifind_quote")
    mx_tool = SimpleNamespace(name="mx_ds_query")
    resolver = ToolGroupResolver(
        mcp_tool_groups={
            "ifind_mcp_tools": [ifind_tool],
            "mx_ds_mcp_tools": [mx_tool],
        }
    )

    assert resolver.resolve(("ifind_mcp_tools",)) == [ifind_tool]
    assert resolver.resolve(("mx_ds_mcp_tools",)) == [mx_tool]
    assert resolver.resolve(("mcp_tools",)) == [ifind_tool, mx_tool]


def test_root_coverage_prompt_nests_under_artifact_root():
    prompt = _agent_prompt("single-stock-coverage.md")
    assert "artifact root / output directory" in prompt
    assert "the whole tree shares one source" in prompt


def test_task2_prompts_are_json_first_with_parent_gates():
    parent = _agent_prompt("task2-financial-modeler.md")
    assert "Do not call MCP tools" in parent
    assert "Do not build, open, edit, update, or save `integrated_model.xlsx`" in parent
    assert "financial_facts_modeler" in parent
    assert "resolve_task2_handoff" in parent
    assert "verify_task2_artifacts" in parent
    assert "write_task2_model_audit" in parent
    assert "reconcile_statement_specs" in parent
    assert (
        "must use exactly the `run_dir` returned by `resolve_task2_handoff`" in parent
    )
    assert "assign `workbook_builder`" in parent.lower()
    assert "assign `model_update_executor`" in parent.lower()
    assert "build_integrated_three_statement_model" not in parent
    assert "validate_integrated_three_statement_model" not in parent

    facts_prompt = _agent_prompt("task2-financial-facts-modeler.md")
    assert "Use the canonical `run_dir` passed by the Task 2 parent" in facts_prompt
    assert "Do not create a coverage run" in facts_prompt
    assert "do not use generic filesystem write/edit tools" in facts_prompt
    assert "passing the parent-provided `run_dir`" in facts_prompt
    assert "parent owns artifact verification" in facts_prompt

    workbook_prompt = _agent_prompt("task2-workbook-builder.md")
    assert "only Task 2 agent allowed to create, open, edit, or save" in workbook_prompt
    assert "Use the canonical `run_dir` passed by the Task 2 parent" in workbook_prompt
    assert "do not use generic filesystem write/edit tools" in workbook_prompt
    assert "build_integrated_three_statement_model" in workbook_prompt
    assert "validate_integrated_three_statement_model" in workbook_prompt
    assert "model_input_json" not in workbook_prompt
    assert "model_audit.md" in workbook_prompt

    update_prompt = _agent_prompt("task2-model-update-executor.md")
    assert "Do not call MCP tools" in update_prompt
    assert "Data retrieval belongs only to `financial_facts_modeler`" in update_prompt
    assert "Use the canonical `run_dir` passed by the Task 2 parent" in update_prompt
    assert "do not use generic filesystem write/edit tools" in update_prompt
    assert "update_integrated_three_statement_model" in update_prompt
    assert "validate_integrated_three_statement_model" in update_prompt
    assert "model_input_json" not in update_prompt
    assert "statement_spec_pack_json" not in update_prompt

    prompt_expectations = {
        "task2-is-modeler.md": (
            "income_statement",
            "validate_income_statement_json",
            "write_income_statement_json",
        ),
        "task2-bs-modeler.md": (
            "balance_sheet",
            "validate_balance_sheet_json",
            "write_balance_sheet_json",
        ),
        "task2-cf-modeler.md": (
            "cash_flow",
            "validate_cash_flow_json",
            "write_cash_flow_json",
        ),
    }
    for prompt_name, (
        statement_type,
        validate_tool,
        write_tool,
    ) in prompt_expectations.items():
        prompt = _agent_prompt(prompt_name)
        assert f'statement_type="{statement_type}"' in prompt
        assert "Do not create, open, edit, or save `integrated_model.xlsx`" in prompt
        assert "Do not read sibling statement JSON" in prompt
        assert "financial-data-normalization" in prompt
        assert "statement-json-checks" in prompt
        assert "Do not construct or pass a full JSON payload" in prompt
        assert "period`, `canonical_key`, `value`, `source`" in prompt
        assert validate_tool in prompt
        assert write_tool in prompt
        assert "row_map" not in prompt
        assert "populate the `" not in prompt
        assert "xlsx-author" not in prompt


def test_statement_skills_emphasize_checks_and_reconciliation_gates():
    for skill_name in [
        "income-statement-model",
        "balance-sheet-model",
        "cash-flow-model",
        "statement-json-checks",
    ]:
        text = _skill_text(skill_name)
        assert "Do not create, open, edit, or save `integrated_model.xlsx`" in text or (
            skill_name == "statement-json-checks"
        )
        assert "Critical" in text
        assert "source coverage" in text.lower()
        assert "canonical" in text.lower()

    reconciliation = _skill_text("statement-reconciliation-checks")
    assert "Critical findings block assignment to `workbook_builder`" in reconciliation
    assert "Cash Flow Statement `ending_cash`" in reconciliation
    assert "Income Statement `net_income`" in reconciliation
    assert "model_audit.md" in reconciliation


def test_root_langgraph_registers_single_stock_debug_entries():
    langgraph_config = json.loads((WORKSPACE_ROOT / "langgraph.json").read_text())

    assert (
        langgraph_config["graphs"]["single_stock_coverage_task2_bs_modeler"]
        == "./single-stock-coverage/src/single_stock_coverage_agent/graph.py:task2_bs_modeler_graph"
    )
    assert (
        langgraph_config["graphs"][
            "single_stock_coverage_task2_financial_facts_modeler"
        ]
        == "./single-stock-coverage/src/single_stock_coverage_agent/graph.py:task2_financial_facts_modeler_graph"
    )
    assert (
        langgraph_config["graphs"]["single_stock_coverage_task2_model_update_executor"]
        == "./single-stock-coverage/src/single_stock_coverage_agent/graph.py:task2_model_update_executor_graph"
    )
    assert (
        langgraph_config["graphs"]["single_stock_coverage_task2_workbook_builder"]
        == "./single-stock-coverage/src/single_stock_coverage_agent/graph.py:task2_workbook_builder_graph"
    )
    assert (
        langgraph_config["graphs"]["single_stock_coverage_task1_company_researcher"]
        == "./single-stock-coverage/src/single_stock_coverage_agent/graph.py:task1_company_researcher_graph"
    )


def test_graph_factories_import_in_test_mode(monkeypatch):
    _clear_env(monkeypatch)
    monkeypatch.setenv("SINGLE_STOCK_COVERAGE_TEST_MODE", "1")
    import single_stock_coverage_agent.graph as graph_module

    graph_module = importlib.reload(graph_module)

    root_graph = asyncio.run(graph_module.graph())
    assert root_graph["name"] == "single_stock_coverage"
    assert root_graph["test_mode"] is True
    assert root_graph["agent_config"]["tool_groups"] == ["coverage_orchestration_tools"]
    assert root_graph["agent_config"]["excluded_builtin_tools"] == [
        "write_file",
        "edit_file",
        "execute",
    ]
    assert root_graph["agent_config"]["subagents"] == [
        "task1_company_researcher",
        "task2_financial_modeler",
        "task3_valuation_analyst",
        "task4_chart_pack_generator",
        "task5_report_assembler",
    ]
    assert root_graph["backend_type"] == "localshell"
    assert root_graph["backend_map"]["single_stock_coverage"] == "localshell"
    assert root_graph["backend_map"]["task4_chart_pack_generator"] == "localshell"
    assert root_graph["backend_map"]["dcf_execution"] == "localshell"
    assert root_graph["backend_map"]["workbook_builder"] == "localshell"
    assert root_graph["backend_map"]["model_update_executor"] == "localshell"
    assert root_graph["backend_map"]["task3_valuation_analyst"] == "filesystem"

    bs_graph = asyncio.run(graph_module.task2_bs_modeler_graph())
    assert bs_graph["name"] == "bs_modeler"
    assert bs_graph["agent_config"]["parent"] == "task2_financial_modeler"
    assert bs_graph["agent_config"]["tool_groups"] == [
        "statement_modeling_tools",
    ]
    assert bs_graph["agent_config"]["excluded_builtin_tools"] == ["task"]
    assert bs_graph["agent_config"]["skills"] == {
        "single_stock_coverage": [
            "financial-data-normalization",
            "balance-sheet-model",
            "statement-json-checks",
        ]
    }

    facts_graph = asyncio.run(graph_module.task2_financial_facts_modeler_graph())
    assert facts_graph["name"] == "financial_facts_modeler"
    assert facts_graph["agent_config"]["tool_groups"] == [
        "ifind_mcp_tools",
        "mx_ds_mcp_tools",
        "task2_financial_fact_artifact_tools",
    ]
    assert facts_graph["agent_config"]["excluded_builtin_tools"] == ["task"]

    workbook_graph = asyncio.run(graph_module.task2_workbook_builder_graph())
    assert workbook_graph["name"] == "workbook_builder"
    assert workbook_graph["agent_config"]["tool_groups"] == [
        "workbook_authoring_tools",
        "task2_audit_artifact_tools",
    ]
    assert workbook_graph["agent_config"]["excluded_builtin_tools"] == ["task"]

    update_graph = asyncio.run(graph_module.task2_model_update_executor_graph())
    assert update_graph["name"] == "model_update_executor"
    assert update_graph["agent_config"]["tool_groups"] == [
        "workbook_update_tools",
        "task2_audit_artifact_tools",
    ]
    assert update_graph["agent_config"]["excluded_builtin_tools"] == ["task"]
