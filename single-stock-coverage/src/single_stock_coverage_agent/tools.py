"""Local artifact tools for the Single Stock Coverage agent."""

from __future__ import annotations

import io
import json
import math
import os
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from financial_agent_runtime import (
    artifact_exists,
    backend_is_daytona,
    copy_artifact,
    ensure_artifact_dir,
    list_artifact_dir,
    materialize_file_artifact,
    read_bytes_artifact,
    read_text_artifact,
    write_text_artifact,
)
from langchain_core.tools import tool

from single_stock_coverage_agent.config import file_storage_root


DEFAULT_OUTPUT_DIR = "./out/coverage"
INLINE_JSON_MAX_BYTES = 32768

_ACTIVE_RUNS: dict[str, Path] = {}

TASK1_REQUIRED_ARTIFACTS: tuple[str, ...] = (
    "company_research.md",
    "business_driver_map.json",
    "source_log.json",
)

TASK2_STATEMENT_ARTIFACTS: dict[str, str] = {
    "income_statement": "income_statement_spec.json",
    "balance_sheet": "balance_sheet_spec.json",
    "cash_flow": "cash_flow_statement_spec.json",
}

TASK2_MODEL_ARTIFACTS: tuple[str, ...] = (
    "financial_facts.json",
    "task2_context_packet.json",
    "statement_spec_pack.json",
    "integrated_model.xlsx",
    "model_audit.md",
)

STATEMENT_JSON_REQUIRED_FIELDS: tuple[str, ...] = (
    "statement_type",
    "canonical_row_keys",
    "line_items",
    "historical_inputs",
    "forecast_logic",
    "assumption_requirements",
    "cross_statement_dependencies",
    "source_coverage",
    "unsourced_items",
    "validation_status",
)

STATEMENT_JSON_ALLOWED_TYPES: tuple[str, ...] = (
    "income_statement",
    "balance_sheet",
    "cash_flow",
)

STATEMENT_JSON_OUTPUTS: dict[str, str] = {
    "income_statement": "income_statement_spec.json",
    "balance_sheet": "balance_sheet_spec.json",
    "cash_flow": "cash_flow_statement_spec.json",
}

STATEMENT_CANONICAL_KEYS: dict[str, tuple[str, ...]] = {
    "income_statement": (
        "revenue_total",
        "gross_profit",
        "ebit",
        "ebitda",
        "interest_expense",
        "pretax_income",
        "tax_expense",
        "net_income",
        "da_total",
    ),
    "balance_sheet": (
        "cash_and_equivalents",
        "total_current_assets",
        "total_assets",
        "total_current_liabilities",
        "total_debt",
        "retained_earnings",
        "total_equity",
        "total_liabilities_and_equity",
    ),
    "cash_flow": (
        "net_income_cf",
        "da_addback",
        "nwc_change",
        "cfo_total",
        "capex",
        "cfi_total",
        "debt_proceeds_repayments",
        "dividends",
        "cff_total",
        "beginning_cash",
        "ending_cash",
    ),
}

STATEMENT_DEPENDENCY_HINTS: dict[str, tuple[str, ...]] = {
    "income_statement": (
        "revenue_build.total_revenue",
        "debt_interest.interest_expense",
        "share_count.diluted_shares",
    ),
    "balance_sheet": (
        "cash_flow.ending_cash",
        "income_statement.net_income",
        "share_count.dividends",
    ),
    "cash_flow": (
        "income_statement.net_income",
        "ppe_da.da_total",
        "balance_sheet.cash_and_equivalents",
    ),
}

THREE_STATEMENT_TABS: tuple[str, ...] = (
    "Cover",
    "Sources",
    "Assumptions",
    "Revenue Build",
    "Income Statement",
    "Balance Sheet",
    "Cash Flow Statement",
    "Working Capital",
    "PP&E & D&A",
    "Debt & Interest",
    "Share Count",
    "DCF Inputs",
    "Checks",
)

REQUIRED_MODEL_NAMES: tuple[str, ...] = (
    "ScenarioSelector",
    "RevDriverBlock",
    "MarginDriverBlock",
    "NWCDriverBlock",
    "CapExDriverBlock",
    "DebtDriverBlock",
)

DEFAULT_THREE_STATEMENT_ROW_MAP: dict[str, dict[str, int]] = {
    "revenue_build": {
        "core_revenue": 10,
        "revenue_total": 12,
        "revenue_growth": 13,
    },
    "income_statement": {
        "revenue_total": 8,
        "cogs": 9,
        "gross_profit": 10,
        "operating_expenses": 12,
        "da_total": 13,
        "ebit": 14,
        "ebitda": 15,
        "interest_expense": 17,
        "interest_income": 18,
        "net_finance_expense": 19,
        "pretax_income": 20,
        "tax_expense": 21,
        "net_income": 22,
        "diluted_shares": 24,
        "eps_diluted": 25,
    },
    "balance_sheet": {
        "cash_and_equivalents": 8,
        "accounts_receivable": 9,
        "inventory": 10,
        "total_current_assets": 11,
        "net_ppe": 13,
        "other_assets": 14,
        "total_assets": 15,
        "accounts_payable": 18,
        "total_current_liabilities": 19,
        "total_debt": 20,
        "total_liabilities": 21,
        "common_stock_apic": 24,
        "retained_earnings": 25,
        "total_equity": 26,
        "total_liabilities_and_equity": 27,
    },
    "cash_flow": {
        "net_income_cf": 8,
        "da_addback": 9,
        "nwc_change": 10,
        "cfo_total": 11,
        "capex": 14,
        "cfi_total": 15,
        "debt_issuance": 18,
        "debt_proceeds_repayments": 19,
        "dividends": 20,
        "cff_total": 21,
        "net_change_cash": 23,
        "beginning_cash": 24,
        "ending_cash": 25,
    },
    "working_capital": {
        "accounts_receivable": 8,
        "inventory": 9,
        "accounts_payable": 10,
        "net_working_capital": 11,
        "nwc_change": 12,
    },
    "ppe_da": {
        "beginning_ppe": 8,
        "capex": 9,
        "da_total": 10,
        "ending_ppe": 11,
    },
    "debt_interest": {
        "beginning_debt": 8,
        "debt_issuance": 9,
        "debt_repayment": 10,
        "ending_debt": 11,
        "short_term_debt_raw": 12,
        "interest_expense": 13,
        "interest_income": 14,
        "net_finance_expense": 15,
        "debt_data_quality_flag": 16,
    },
    "share_count": {
        "beginning_diluted_shares": 8,
        "share_issuance": 9,
        "buybacks": 10,
        "ending_diluted_shares": 11,
        "dividends": 12,
    },
    "dcf_inputs": {
        "revenue_total": 8,
        "ebit": 9,
        "tax_rate": 10,
        "da_total": 11,
        "capex": 12,
        "nwc_change": 13,
        "total_debt": 14,
        "cash_and_equivalents": 15,
        "diluted_shares": 16,
        "scenario_label": 17,
    },
    "checks": {
        "failing_check_count": 4,
        "first_check_row": 8,
    },
}


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _workspace_root() -> Path:
    return file_storage_root()


def _resolve_workspace_path(path: str | Path) -> Path:
    candidate = Path(path)
    if candidate.is_absolute():
        return _canonicalize_workspace_path(candidate)
    return _canonicalize_workspace_path(_workspace_root() / candidate)


def _ensure_dir(path: Path) -> None:
    ensure_artifact_dir(path)


def _write_text(path: Path, text: str) -> None:
    write_text_artifact(path, text, encoding="utf-8")


def _save_workbook(wb: Any, path: Path) -> None:
    materialize_file_artifact(path, lambda target: wb.save(target))


def _save_workbook_with_formula_cache(
    wb: Any,
    path: Path,
    formula_cache: dict[str, dict[str, Any]],
) -> int:
    def _produce(target: Path) -> int:
        wb.save(target)
        return _patch_xlsx_formula_caches(target, formula_cache)

    return materialize_file_artifact(path, _produce)


def _copy_file_artifact(source_path: Path, artifact_path: Path) -> None:
    copy_artifact(source_path, artifact_path)


def _load_workbook(path: Path, **kwargs: Any):
    import openpyxl

    if backend_is_daytona():
        return openpyxl.load_workbook(io.BytesIO(read_bytes_artifact(path)), **kwargs)
    return openpyxl.load_workbook(path, **kwargs)


def _canonicalize_workspace_path(path: Path) -> Path:
    """Map project-local out/ paths to the workspace-level artifact root."""
    project_out = _project_root() / "out"
    try:
        return _workspace_root() / "out" / path.resolve().relative_to(project_out)
    except ValueError:
        return path


def _project_wrong_root_path(path: Path) -> Path | None:
    try:
        rel = path.resolve().relative_to(_workspace_root())
    except ValueError:
        return None
    if not rel.parts or rel.parts[0] != "out":
        return None
    return _project_root() / rel


def _slugify(text: str, fallback: str = "unknown") -> str:
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "-", str(text).strip()).strip("-")
    return normalized.lower() or fallback


def _safe_market(market: str) -> str:
    return _slugify(market or "market", "market")


def _safe_ticker(ticker: str) -> str:
    return _slugify(ticker or "ticker", "ticker")


def _coverage_root(output_dir: str = DEFAULT_OUTPUT_DIR) -> Path:
    return _resolve_workspace_path(output_dir)


def _coverage_dir(market: str, ticker: str, output_dir: str = DEFAULT_OUTPUT_DIR) -> Path:
    return _coverage_root(output_dir) / f"{_safe_market(market)}-{_safe_ticker(ticker)}"


def _relative_to_workspace(path: Path) -> str:
    if backend_is_daytona():
        try:
            return str(path.relative_to(_workspace_root()))
        except ValueError:
            return str(path)
    try:
        return str(path.resolve().relative_to(_workspace_root()))
    except ValueError:
        return str(path.resolve())


def _json_loads(data_json: str, field_name: str) -> Any:
    try:
        return json.loads(data_json)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{field_name} must be valid JSON: {exc}") from exc


def _field(data: dict[str, Any], *names: str, default: Any = None) -> Any:
    for name in names:
        value = data.get(name)
        if value not in (None, ""):
            return value
    return default


def _as_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    cleaned = str(value).replace(",", "").replace("%", "").strip()
    if not cleaned or cleaned.upper() in {"[UNSOURCED]", "UNSOURCED"}:
        return default
    try:
        return float(cleaned)
    except ValueError:
        return default


def _as_decimal(value: Any, default: float = 0.0) -> float:
    parsed = _as_float(value, default)
    return parsed / 100.0 if abs(parsed) > 1.5 else parsed


def _period_year(value: Any, fallback: int) -> int:
    match = re.search(r"(19|20)\d{2}", str(value or ""))
    return int(match.group(0)) if match else fallback


def _column_letter(col_index: int) -> str:
    letters = ""
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _sheet_ref(sheet_name: str, cell_ref: str) -> str:
    if " " in sheet_name or "&" in sheet_name:
        return f"'{sheet_name}'!{cell_ref}"
    return f"{sheet_name}!{cell_ref}"


def _formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _column_index(col_letter: str) -> int:
    result = 0
    for char in str(col_letter).upper():
        if "A" <= char <= "Z":
            result = result * 26 + ord(char) - ord("A") + 1
    return result


def _normal_formula_cache_value(value: Any) -> Any:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isfinite(float(value)):
            return float(value)
        return None
    if value is None:
        return None
    return str(value)


def _evaluate_workbook_formula_caches(wb: Any) -> dict[str, dict[str, Any]]:
    """Evaluate the narrow formula surface emitted by the Task 2 builder.

    This is intentionally small: it supports only the functions and reference
    shapes generated by build_integrated_three_statement_model. It is used to
    materialize formula caches so spreadsheet previews and data_only readers do
    not see blank formula cells before Excel recalculates the model.
    """

    value_cache: dict[tuple[str, str], Any] = {}
    formula_cache: dict[str, dict[str, Any]] = {}
    string_literals: list[str] = []

    def cell_coord(col: str, row: int) -> str:
        return f"{col.upper()}{row}"

    def numeric(value: Any) -> float:
        if value is None or value == "":
            return 0.0
        if isinstance(value, bool):
            return 1.0 if value else 0.0
        if isinstance(value, (int, float)):
            return float(value)
        return _as_float(value, 0.0)

    def resolve_sheet(current_sheet: str, sheet_token: str) -> str:
        return current_sheet if sheet_token == "__CURRENT__" else sheet_token

    def evaluate_cell(sheet_name: str, coord: str, stack: set[tuple[str, str]]) -> Any:
        key = (sheet_name, coord.upper())
        if key in value_cache:
            return value_cache[key]
        if key in stack or sheet_name not in wb.sheetnames:
            return 0.0
        cell = wb[sheet_name][coord.upper()]
        value = cell.value
        if _formula(value):
            result = evaluate_formula(sheet_name, value, stack | {key})
            value_cache[key] = result
            return result
        return value

    def ref(current_sheet: str, sheet_token: str, col: str, row: int) -> Any:
        sheet_name = resolve_sheet(current_sheet, sheet_token)
        value = evaluate_cell(sheet_name, cell_coord(col, row), set())
        return 0.0 if value is None else value

    def range_values(
        current_sheet: str,
        sheet_token: str,
        start_col: str,
        start_row: int,
        end_col: str,
        end_row: int,
    ) -> list[Any]:
        sheet_name = resolve_sheet(current_sheet, sheet_token)
        values: list[Any] = []
        for row_idx in range(min(start_row, end_row), max(start_row, end_row) + 1):
            for col_idx in range(
                min(_column_index(start_col), _column_index(end_col)),
                max(_column_index(start_col), _column_index(end_col)) + 1,
            ):
                values.append(
                    evaluate_cell(sheet_name, f"{_column_letter(col_idx)}{row_idx}", set())
                )
        return values

    def flatten(values: tuple[Any, ...]) -> list[Any]:
        flattened: list[Any] = []
        for value in values:
            if isinstance(value, list):
                flattened.extend(value)
            else:
                flattened.append(value)
        return flattened

    def excel_sum(*values: Any) -> float:
        return sum(numeric(value) for value in flatten(values))

    def excel_max(*values: Any) -> float:
        flattened = flatten(values)
        if not flattened:
            return 0.0
        return max(numeric(value) for value in flattened)

    def excel_if(condition: Any, true_value: Any, false_value: Any) -> Any:
        return true_value if bool(condition) else false_value

    def excel_countif(values: Any, criteria: Any) -> int:
        flattened = flatten((values,))
        criteria_text = str(criteria)
        if criteria_text == "<>0":
            return sum(1 for value in flattened if abs(numeric(value)) > 1e-12)
        return sum(1 for value in flattened if str(value) == criteria_text)

    def mask_strings(expr: str) -> str:
        string_literals.clear()

        def repl(match: re.Match[str]) -> str:
            string_literals.append(match.group(0))
            return f"__strlit{len(string_literals) - 1}__"

        return re.sub(r'"[^"]*"', repl, expr)

    def restore_strings(expr: str) -> str:
        for idx, literal in enumerate(string_literals):
            expr = expr.replace(f"__strlit{idx}__", literal)
        return expr

    def translate_references(current_sheet: str, expr: str) -> str:
        del current_sheet
        expr = mask_strings(expr)

        expr = re.sub(
            r"'([^']+)'!\$?([A-Z]{1,3})\$?(\d+)",
            lambda match: (
                f'REF("{match.group(1)}","{match.group(2).upper()}",'
                f"{int(match.group(3))})"
            ),
            expr,
        )
        expr = re.sub(
            r"\b([A-Za-z_][A-Za-z0-9_]*)!\$?([A-Z]{1,3})\$?(\d+)",
            lambda match: (
                f'REF("{match.group(1)}","{match.group(2).upper()}",'
                f"{int(match.group(3))})"
            ),
            expr,
        )
        expr = re.sub(
            r"(?<![A-Za-z0-9_])\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)",
            lambda match: (
                f'RANGE("__CURRENT__","{match.group(1).upper()}",{int(match.group(2))},'
                f'"{match.group(3).upper()}",{int(match.group(4))})'
            ),
            expr,
        )
        expr = re.sub(
            r"(?<![A-Za-z0-9_\"])\$?([A-Z]{1,3})\$?(\d+)(?![A-Za-z0-9_\"])",
            lambda match: (
                f'REF("__CURRENT__","{match.group(1).upper()}",{int(match.group(2))})'
            ),
            expr,
        )
        expr = expr.replace("<>", "!=")
        expr = re.sub(r"(?<![<>=!])=(?![=])", "==", expr)
        return restore_strings(expr)

    def split_formula_args(expr: str) -> list[str]:
        args: list[str] = []
        depth = 0
        in_string = False
        start = 0
        for idx, char in enumerate(expr):
            if char == '"':
                in_string = not in_string
            elif not in_string and char == "(":
                depth += 1
            elif not in_string and char == ")":
                depth -= 1
            elif not in_string and char == "," and depth == 0:
                args.append(expr[start:idx].strip())
                start = idx + 1
        args.append(expr[start:].strip())
        return args

    def evaluate_expression(sheet_name: str, expr: str) -> Any:
        translated = translate_references(sheet_name, expr)
        namespace = {
            "COUNTIF": excel_countif,
            "IF": excel_if,
            "MAX": excel_max,
            "REF": lambda sheet_token, col, row: ref(sheet_name, sheet_token, col, row),
            "RANGE": lambda sheet_token, start_col, start_row, end_col, end_row: range_values(
                sheet_name,
                sheet_token,
                start_col,
                int(start_row),
                end_col,
                int(end_row),
            ),
            "SUM": excel_sum,
        }
        return eval(translated, {"__builtins__": {}}, namespace)

    def evaluate_formula(sheet_name: str, formula: str, stack: set[tuple[str, str]]) -> Any:
        expr = formula[1:] if formula.startswith("=") else formula
        stripped = expr.strip()
        try:
            if stripped.upper().startswith("IF(") and stripped.endswith(")"):
                args = split_formula_args(stripped[3:-1])
                if len(args) == 3:
                    selected = args[1] if bool(evaluate_expression(sheet_name, args[0])) else args[2]
                    result = evaluate_formula(sheet_name, f"={selected}", stack)
                else:
                    result = evaluate_expression(sheet_name, expr)
            else:
                result = evaluate_expression(sheet_name, expr)
        except Exception:
            return None
        return _normal_formula_cache_value(result)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _formula(cell.value):
                    value = evaluate_cell(ws.title, cell.coordinate, set())
                    normal = _normal_formula_cache_value(value)
                    if normal is not None:
                        formula_cache.setdefault(ws.title, {})[cell.coordinate] = normal
    return formula_cache


def _patch_xlsx_formula_caches(
    workbook_path: Path,
    formula_cache: dict[str, dict[str, Any]],
) -> int:
    if not formula_cache:
        return 0

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    ET.register_namespace("", main_ns)
    ET.register_namespace("r", rel_ns)

    tmp_path = workbook_path.with_name(f"{workbook_path.stem}.formula-cache.tmp.xlsx")
    patched = 0
    with zipfile.ZipFile(workbook_path, "r") as zin:
        workbook_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall(f"{{{package_rel_ns}}}Relationship")
            if rel.attrib.get("Id") and rel.attrib.get("Target")
        }
        sheet_paths: dict[str, str] = {}
        for sheet in workbook_root.findall(f"{{{main_ns}}}sheets/{{{main_ns}}}sheet"):
            name = sheet.attrib.get("name")
            rel_id = sheet.attrib.get(f"{{{rel_ns}}}id")
            target = rel_targets.get(str(rel_id))
            if not name or not target:
                continue
            normalized = target.lstrip("/")
            if not normalized.startswith("xl/"):
                normalized = f"xl/{normalized}"
            sheet_paths[normalized] = name

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                sheet_name = sheet_paths.get(info.filename)
                if sheet_name and sheet_name in formula_cache:
                    root = ET.fromstring(data)
                    cells = root.findall(f".//{{{main_ns}}}c")
                    cache_by_ref = formula_cache[sheet_name]
                    for cell in cells:
                        ref = cell.attrib.get("r")
                        if not ref or ref not in cache_by_ref:
                            continue
                        if cell.find(f"{{{main_ns}}}f") is None:
                            continue
                        value = cache_by_ref[ref]
                        v_node = cell.find(f"{{{main_ns}}}v")
                        if v_node is None:
                            v_node = ET.SubElement(cell, f"{{{main_ns}}}v")
                        if isinstance(value, bool):
                            cell.attrib["t"] = "b"
                            v_node.text = "1" if value else "0"
                        elif isinstance(value, (int, float)) and not isinstance(value, bool):
                            cell.attrib.pop("t", None)
                            v_node.text = f"{float(value):.15g}"
                        else:
                            cell.attrib["t"] = "str"
                            v_node.text = str(value)
                        patched += 1
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                zout.writestr(info, data)
    shutil.move(str(tmp_path), str(workbook_path))
    return patched


def _normalize_period_label(value: Any, *, is_actual: bool) -> str:
    text = str(value or "").strip()
    year = _period_year(text, 0)
    if year <= 0:
        return text
    suffix = "A" if is_actual else "E"
    if text.upper().startswith("FY"):
        if text.upper().endswith(("A", "E")):
            return text
        return f"FY{year}{suffix}"
    if text.upper().endswith(("A", "E")):
        return f"FY{text}"
    return f"FY{year}{suffix}"


def _is_interim_period(value: Any) -> bool:
    text = str(value or "").upper().strip()
    if not text:
        return False
    if text.startswith("FY"):
        return False
    return bool(
        re.search(
            r"\bQ[1-4]\b|\b[1-4]Q\b|QUARTER|INTERIM|YTD|9M|1H|2H|H1|H2|半年度|季度|一季|三季",
            text,
        )
    )


def _nested_field(data: dict[str, Any], *paths: str, default: Any = None) -> Any:
    for path in paths:
        current: Any = data
        for part in path.split("."):
            if not isinstance(current, dict):
                current = None
                break
            current = current.get(part)
        if current not in (None, ""):
            return current
    return default


def _statement_spec_pack(payload: dict[str, Any]) -> dict[str, Any]:
    pack = payload.get("statement_spec_pack")
    if isinstance(pack, dict):
        return pack
    if isinstance(payload.get("statement_specs"), dict):
        return payload
    return {}


def _statement_specs(payload: dict[str, Any]) -> dict[str, Any]:
    pack = _statement_spec_pack(payload)
    specs = pack.get("statement_specs")
    return specs if isinstance(specs, dict) else {}


def _income_statement_spec(payload: dict[str, Any]) -> dict[str, Any]:
    spec = _statement_specs(payload).get("income_statement")
    return spec if isinstance(spec, dict) else {}


def _balance_sheet_spec(payload: dict[str, Any]) -> dict[str, Any]:
    spec = _statement_specs(payload).get("balance_sheet")
    return spec if isinstance(spec, dict) else {}


def _cash_flow_spec(payload: dict[str, Any]) -> dict[str, Any]:
    spec = _statement_specs(payload).get("cash_flow")
    return spec if isinstance(spec, dict) else {}


def _revenue_build_spec(payload: dict[str, Any]) -> dict[str, Any]:
    direct = payload.get("revenue_build_spec")
    if isinstance(direct, dict):
        return direct
    spec = _income_statement_spec(payload).get("revenue_build_spec")
    return spec if isinstance(spec, dict) else {}


def _period_key(value: Any) -> str:
    year = _period_year(value, 0)
    return str(year) if year > 0 else str(value or "").strip()


def _historical_input_lookup(payload: dict[str, Any]) -> dict[str, dict[str, float]]:
    lookup: dict[str, dict[str, float]] = {}
    specs = _statement_specs(payload)
    for spec in specs.values():
        if not isinstance(spec, dict):
            continue
        raw = spec.get("historical_inputs")
        if not isinstance(raw, list):
            continue
        for item in raw:
            if not isinstance(item, dict):
                continue
            key = str(item.get("canonical_key") or "")
            if not key:
                continue
            period = _period_key(item.get("period"))
            if not period:
                continue
            lookup.setdefault(period, {})[key] = _as_float(item.get("value"), 0.0)
    return lookup


def _shares_by_year(payload: dict[str, Any]) -> dict[str, float]:
    facts = payload.get("financial_facts") if isinstance(payload.get("financial_facts"), dict) else payload
    raw = facts.get("shares") if isinstance(facts, dict) else None
    if not isinstance(raw, dict):
        return {}
    shares: dict[str, float] = {}
    for year, item in raw.items():
        if not isinstance(item, dict):
            continue
        value = _field(
            item,
            "diluted_shares",
            "shares",
            "shares_outstanding",
            "total_shares_期末",
        )
        parsed = _as_float(value, 0.0)
        if parsed:
            shares[_period_key(year)] = parsed
    return shares


def _normalize_historical_record(
    record: dict[str, Any],
    payload: dict[str, Any],
    spec_lookup: dict[str, dict[str, float]],
    share_lookup: dict[str, float],
) -> dict[str, Any]:
    year_key = _period_key(record.get("year") or record.get("period"))
    spec_values = spec_lookup.get(year_key, {})
    revenue = _as_float(
        _nested_field(record, "revenue", "revenue_total", "is.revenue", "income_statement.revenue"),
        spec_values.get("revenue_total", 0.0),
    )
    gross_profit = _as_float(
        _nested_field(record, "gross_profit", "is.gross_profit", "income_statement.gross_profit"),
        spec_values.get("gross_profit", 0.0),
    )
    if not gross_profit:
        gross_margin = _as_decimal(
            _nested_field(record, "gross_margin", "is.gross_margin", "income_statement.gross_margin"),
            0.0,
        )
        if revenue and gross_margin:
            gross_profit = revenue * gross_margin
    cogs = _as_float(
        _nested_field(record, "cogs", "is.cogs", "income_statement.cogs"),
        revenue - gross_profit if revenue or gross_profit else spec_values.get("cogs", 0.0),
    )
    selling = _as_float(_nested_field(record, "selling_expenses", "is.selling_expenses", "income_statement.selling_expenses"), 0.0)
    admin = _as_float(_nested_field(record, "admin_expenses", "is.admin_expenses", "income_statement.admin_expenses"), 0.0)
    rd = _as_float(_nested_field(record, "rd_expenses", "is.rd_expenses", "income_statement.rd_expenses"), 0.0)
    da = _as_float(
        _nested_field(record, "da", "da_total", "is.da_total", "income_statement.da_total", "cash_flow.DA", "cash_flow.da", "derived.da_estimated"),
        spec_values.get("da_total", 0.0),
    )
    ebit = _as_float(
        _nested_field(record, "ebit", "operating_income", "is.ebit", "is.operating_profit", "income_statement.ebit", "income_statement.operating_profit"),
        spec_values.get("ebit", 0.0),
    )
    operating_expenses = _as_float(
        _nested_field(record, "operating_expenses", "opex", "is.operating_expenses", "income_statement.operating_expenses"),
        selling + admin + rd if any((selling, admin, rd)) else max(gross_profit - ebit - da, 0.0),
    )
    tax_expense = _as_float(
        _nested_field(record, "tax_expense", "tax", "is.income_tax", "income_statement.income_tax"),
        spec_values.get("tax_expense", 0.0),
    )
    net_income = _as_float(
        _nested_field(record, "net_income", "is.net_income", "is.net_income_to_parent", "income_statement.net_income", "income_statement.net_income_attr_parent"),
        spec_values.get("net_income", 0.0),
    )
    cash = _as_float(
        _nested_field(record, "cash", "cash_and_equivalents", "bs.cash", "balance_sheet.cash"),
        spec_values.get("cash_and_equivalents", 0.0),
    )
    ar = _as_float(
        _nested_field(record, "accounts_receivable", "ar", "bs.accounts_receivable", "balance_sheet.accounts_receivable"),
        spec_values.get("accounts_receivable", 0.0),
    )
    inventory = _as_float(
        _nested_field(record, "inventory", "bs.inventory", "balance_sheet.inventory"),
        spec_values.get("inventory", 0.0),
    )
    ap = _as_float(
        _nested_field(record, "accounts_payable", "ap", "bs.accounts_payable", "balance_sheet.accounts_payable"),
        spec_values.get("accounts_payable", 0.0),
    )
    total_current_assets = _as_float(
        _nested_field(record, "total_current_assets", "current_assets", "bs.current_assets", "balance_sheet.current_assets", "balance_sheet.total_current_assets"),
        spec_values.get("total_current_assets", 0.0),
    )
    ppe = _as_float(
        _nested_field(record, "ppe", "net_ppe", "ppe_net", "bs.ppe_net", "balance_sheet.ppe_net", "balance_sheet.PP&E"),
        spec_values.get("ppe_net", spec_values.get("net_ppe", 0.0)),
    )
    total_assets = _as_float(
        _nested_field(record, "total_assets", "bs.total_assets", "balance_sheet.total_assets"),
        spec_values.get("total_assets", 0.0),
    )
    total_liabilities = _as_float(
        _nested_field(record, "total_liabilities", "bs.total_liabilities", "balance_sheet.total_liabilities"),
        spec_values.get("total_liabilities", 0.0),
    )
    current_liabilities = _as_float(
        _nested_field(record, "total_current_liabilities", "current_liabilities", "bs.current_liabilities", "balance_sheet.current_liabilities", "balance_sheet.total_current_liabilities"),
        spec_values.get("total_current_liabilities", 0.0),
    )
    total_equity = _as_float(
        _nested_field(record, "total_equity", "bs.total_equity", "balance_sheet.total_equity"),
        spec_values.get("total_equity", 0.0),
    )
    explicit_debt = _nested_field(
        record,
        "interest_bearing_debt",
        "debt",
        "total_debt",
        "bs.interest_bearing_debt",
        "bs.interest_bearing_debt_亿元",
        "bs.total_debt",
        "balance_sheet.interest_bearing_debt",
        "balance_sheet.total_debt",
    )
    debt = _as_float(explicit_debt, spec_values.get("total_debt", 0.0))
    short_term_debt_raw = _as_float(
        _nested_field(record, "short_term_debt_raw", "bs.short_term_debt_raw", "bs.short_term_debt", "balance_sheet.short_term_debt_raw", "balance_sheet.short_term_debt"),
        spec_values.get("short_term_debt", 0.0),
    )
    long_term_debt_raw = _as_float(
        _nested_field(record, "long_term_debt_raw", "bs.long_term_debt_raw", "bs.long_term_debt", "balance_sheet.long_term_debt_raw", "balance_sheet.long_term_debt"),
        spec_values.get("long_term_debt", 0.0),
    )
    if explicit_debt in (None, "") and not debt:
        debt = _as_float(_nested_field(record, "bs.short_term_debt", "balance_sheet.short_term_debt"), 0.0) + _as_float(
            _nested_field(record, "bs.long_term_debt", "balance_sheet.long_term_debt"),
            0.0,
        )
    capex = _as_float(
        _nested_field(record, "capex", "cf.capex", "cash_flow.capex"),
        spec_values.get("capex", 0.0),
    )
    retained_earnings = _as_float(
        _nested_field(record, "retained_earnings", "bs.retained_earnings", "balance_sheet.retained_earnings"),
        spec_values.get("retained_earnings", total_equity),
    )
    shares = _as_float(
        _nested_field(record, "shares", "diluted_shares", "shares_outstanding"),
        share_lookup.get(year_key, spec_values.get("diluted_shares", 0.0)),
    )
    other_assets = max(total_assets - cash - ar - inventory - ppe, 0.0) if total_assets else 0.0
    source = _source_for(record)
    return {
        **record,
        "year": _period_year(record.get("year") or record.get("period"), 2020),
        "period": _normalize_period_label(record.get("period") or record.get("year"), is_actual=True),
        "revenue": revenue,
        "cogs": cogs,
        "gross_profit": gross_profit,
        "operating_expenses": operating_expenses,
        "da": da,
        "ebit": ebit,
        "ebitda": _as_float(_nested_field(record, "ebitda", "is.ebitda", "income_statement.ebitda"), ebit + da),
        "interest_expense": _as_float(
            _nested_field(record, "interest_expense", "is.interest_expense", "income_statement.interest_expense"),
            spec_values.get("interest_expense", 0.0),
        ),
        "interest_income": _as_float(
            _nested_field(record, "interest_income", "is.interest_income", "income_statement.interest_income"),
            spec_values.get("interest_income", 0.0),
        ),
        "net_finance_expense": _as_float(
            _nested_field(record, "net_finance_expense", "finance_expenses", "is.finance_expenses", "income_statement.finance_expenses"),
            spec_values.get("finance_expenses", 0.0),
        ),
        "pretax_income": _as_float(
            _nested_field(record, "pretax_income", "ebt", "is.pretax_income", "income_statement.pretax_income"),
            spec_values.get("pretax_income", ebit),
        ),
        "tax_expense": tax_expense,
        "net_income": net_income,
        "cash": cash,
        "ar": ar,
        "inventory": inventory,
        "ap": ap,
        "total_current_assets": total_current_assets,
        "ppe": ppe,
        "other_assets": other_assets,
        "total_assets": total_assets,
        "current_liabilities": current_liabilities,
        "total_liabilities": total_liabilities,
        "total_equity": total_equity,
        "debt": debt,
        "short_term_debt_raw": short_term_debt_raw,
        "long_term_debt_raw": long_term_debt_raw,
        "debt_to_market_equity": _as_float(
            _nested_field(record, "debt_to_market_equity", "bs.debt_to_market_equity"),
            0.0,
        ),
        "retained_earnings": retained_earnings,
        "capex": capex,
        "cfo_total": _as_float(_nested_field(record, "cfo_total", "cash_flow.operating_cf"), 0.0),
        "cfi_total": _as_float(_nested_field(record, "cfi_total", "cash_flow.investing_cf"), 0.0),
        "cff_total": _as_float(_nested_field(record, "cff_total", "cash_flow.financing_cf"), 0.0),
        "dividends": _as_float(_nested_field(record, "dividends", "cash_flow.dividends_paid"), 0.0),
        "shares": shares,
        "source": source,
    }


def _merge_payload_model_input(payload: dict[str, Any]) -> dict[str, Any]:
    facts = payload.get("financial_facts")
    if isinstance(facts, dict):
        merged = dict(facts)
        merged.update(payload)
        return merged
    return payload


def _historical_records(payload: dict[str, Any]) -> list[dict[str, Any]]:
    facts = payload.get("financial_facts") if isinstance(payload.get("financial_facts"), dict) else payload
    records = payload.get("historicals") or facts.get("historicals") or payload.get("historical_financials") or []
    if not isinstance(records, list) or not records:
        raise ValueError("Task 2 model sources must include a non-empty historicals list")
    normalized: list[dict[str, Any]] = []
    spec_lookup = _historical_input_lookup(payload)
    share_lookup = _shares_by_year(payload)
    for idx, record in enumerate(records):
        if not isinstance(record, dict):
            raise ValueError(f"historicals[{idx}] must be a JSON object")
        if _is_interim_period(record.get("period")):
            continue
        enriched = _normalize_historical_record(record, payload, spec_lookup, share_lookup)
        if not enriched["year"]:
            enriched["year"] = 2020 + idx
            enriched["period"] = f"FY{enriched['year']}A"
        normalized.append(enriched)
    if not normalized:
        raise ValueError("Task 2 model sources must include at least one complete annual historical period")
    return sorted(normalized, key=lambda item: item["year"])


def _interim_historical_records(payload: dict[str, Any]) -> list[dict[str, Any]]:
    facts = payload.get("financial_facts") if isinstance(payload.get("financial_facts"), dict) else payload
    records = payload.get("historicals") or facts.get("historicals") or payload.get("historical_financials") or []
    if not isinstance(records, list):
        return []
    return [
        record
        for record in records
        if isinstance(record, dict) and _is_interim_period(record.get("period"))
    ]


def _forecast_labels(payload: dict[str, Any], latest_year: int) -> list[str]:
    forecast_records = _forecast_revenue_records(payload)
    if forecast_records:
        return [
            _normalize_period_label(record.get("period"), is_actual=False)
            for record in forecast_records
        ]
    raw = payload.get("forecast_years") or payload.get("projection_years_list")
    if isinstance(raw, list) and raw:
        return [
            _normalize_period_label(item, is_actual=False)
            for item in raw
        ]
    context = payload.get("task2_context_packet")
    if isinstance(context, dict):
        for path in (
            "model_horizon.forecast_years",
            "period_plan.forecast_periods",
            "forecast_years",
        ):
            raw_periods = _nested_field(context, path)
            if isinstance(raw_periods, list) and raw_periods:
                return [
                    _normalize_period_label(item, is_actual=False)
                    for item in raw_periods
                ]
        horizon = str(_nested_field(context, "period_plan.forecast_horizon", default=""))
        years = re.findall(r"20\d{2}E", horizon)
        if years:
            return [_normalize_period_label(year, is_actual=False) for year in years]
    periods = int(_as_float(payload.get("projection_periods"), 2))
    periods = max(1, min(periods, 7))
    return [f"FY{latest_year + offset}E" for offset in range(1, periods + 1)]


def _source_for(record: dict[str, Any]) -> str:
    source = _field(record, "source", "sources", "source_text", default="[UNSOURCED]")
    if source == "[UNSOURCED]":
        source = _nested_field(
            record,
            "income_statement.source",
            "balance_sheet.source",
            "cash_flow.source",
            default="[UNSOURCED]",
        )
    if isinstance(source, list):
        return "; ".join(str(item) for item in source) or "[UNSOURCED]"
    return str(source or "[UNSOURCED]")


def _forecast_revenue_records(payload: dict[str, Any]) -> list[dict[str, Any]]:
    revenue_build = _revenue_build_spec(payload)
    reconciliation = revenue_build.get("total_revenue_reconciliation")
    if isinstance(reconciliation, dict):
        forecast = reconciliation.get("forecast")
        if isinstance(forecast, list):
            return [item for item in forecast if isinstance(item, dict)]
        if isinstance(forecast, dict):
            records: list[dict[str, Any]] = []
            for period, item in forecast.items():
                if isinstance(item, dict):
                    records.append({"period": period, **item})
                else:
                    records.append({"period": period, "total": item})
            return records
    records = revenue_build.get("forecast")
    if isinstance(records, list):
        return [item for item in records if isinstance(item, dict)]
    if isinstance(records, dict):
        result: list[dict[str, Any]] = []
        for period, item in records.items():
            if isinstance(item, dict):
                result.append({"period": period, **item})
            else:
                result.append({"period": period, "total": item})
        return result
    return []


def _raw_scalar(value: Any) -> bool:
    return isinstance(value, (int, float, str)) and not isinstance(value, bool)


def _metric_pick(record: dict[str, Any], *names: str, pct: bool = False) -> tuple[bool, float]:
    for name in names:
        if name in record and record[name] not in (None, ""):
            return True, _as_decimal(record[name], 0.0) if pct else _as_float(record[name], 0.0)
    return False, 0.0


def _revenue_period_metrics(record: Any) -> dict[str, Any]:
    if not isinstance(record, dict):
        value = _as_float(record, 0.0)
        return {"revenue": value} if value else {}

    metrics: dict[str, Any] = {}
    has_revenue, revenue = _metric_pick(
        record,
        "revenue",
        "revenue_total",
        "segment_revenue",
        "amount",
        "value",
        "total",
    )
    has_cost, cost = _metric_pick(record, "cost", "cogs", "cost_of_revenue", "segment_cost")
    has_gp, gross_profit = _metric_pick(record, "gross_profit", "gp")
    has_gm, gross_margin = _metric_pick(
        record,
        "gross_margin",
        "gross_margin_pct",
        "margin",
        "margin_pct",
        pct=True,
    )
    has_mix, mix = _metric_pick(record, "mix", "revenue_mix", "revenue_mix_pct", pct=True)

    if has_gp and not has_cost and has_revenue:
        cost = revenue - gross_profit
        has_cost = True
    if has_cost and not has_gp and has_revenue:
        gross_profit = revenue - cost
        has_gp = True
    if has_gp and not has_gm and revenue:
        gross_margin = gross_profit / revenue
        has_gm = True

    if has_revenue:
        metrics["revenue"] = revenue
    if has_cost:
        metrics["cost"] = cost
    if has_gp:
        metrics["gross_profit"] = gross_profit
    if has_gm:
        metrics["gross_margin"] = gross_margin
    if has_mix:
        metrics["mix"] = mix

    reserved = {
        "period",
        "year",
        "source",
        "sources",
        "note",
        "notes",
        "source_ref",
        "estimation_method",
        "revenue",
        "revenue_total",
        "segment_revenue",
        "amount",
        "value",
        "total",
        "cost",
        "cogs",
        "cost_of_revenue",
        "segment_cost",
        "gross_profit",
        "gp",
        "gross_margin",
        "gross_margin_pct",
        "margin",
        "margin_pct",
        "mix",
        "revenue_mix",
        "revenue_mix_pct",
    }
    drivers: dict[str, float] = {}
    for key, value in record.items():
        if key in reserved or not _raw_scalar(value):
            continue
        parsed = _as_float(value, 0.0)
        if parsed or str(value).strip() in {"0", "0.0"}:
            drivers[key] = parsed
    if drivers:
        metrics["drivers"] = drivers
    return metrics


def _store_revenue_period_metrics(
    component: dict[str, Any],
    period: Any,
    record: Any,
    *,
    is_actual: bool,
) -> None:
    label = _normalize_period_label(period, is_actual=is_actual)
    if not label:
        return
    metrics = _revenue_period_metrics(record)
    if not metrics:
        return
    bucket_name = "historical_by_period" if is_actual else "forecast_by_period"
    existing = component[bucket_name].setdefault(label, {})
    drivers = metrics.pop("drivers", {})
    existing.update(metrics)
    if drivers:
        existing.setdefault("drivers", {}).update(drivers)


def _store_period_collection(
    component: dict[str, Any],
    raw: Any,
    *,
    is_actual: bool,
) -> None:
    if isinstance(raw, dict):
        for period, record in raw.items():
            _store_revenue_period_metrics(component, period, record, is_actual=is_actual)
    elif isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            period = item.get("period") or item.get("year")
            if not period:
                continue
            _store_revenue_period_metrics(component, period, item, is_actual=is_actual)


def _store_driver_collection(component: dict[str, Any], raw: Any) -> None:
    if not isinstance(raw, dict):
        return
    for driver_name, driver_payload in raw.items():
        if not isinstance(driver_payload, dict):
            continue
        for period, record in driver_payload.items():
            if not re.search(r"(19|20)\d{2}", str(period or "")):
                continue
            is_actual = not str(period).upper().endswith("E")
            label = _normalize_period_label(period, is_actual=is_actual)
            if isinstance(record, dict):
                value = _field(record, "value", "amount", "total")
            else:
                value = record
            if value in (None, ""):
                continue
            bucket_name = "historical_by_period" if is_actual else "forecast_by_period"
            existing = component[bucket_name].setdefault(label, {})
            existing.setdefault("drivers", {})[str(driver_name)] = _as_float(value, 0.0)


def _normalize_revenue_components(payload: dict[str, Any]) -> list[dict[str, Any]]:
    revenue_build = _revenue_build_spec(payload)
    segments = revenue_build.get("segments")
    components: list[dict[str, Any]] = []
    if isinstance(segments, list):
        for idx, segment in enumerate(segments):
            if not isinstance(segment, dict):
                continue
            display_name = str(
                segment.get("display_name")
                or segment.get("label")
                or segment.get("segment_key")
                or segment.get("segment_id")
                or segment.get("name")
                or f"Revenue Component {idx + 1}"
            ).strip()
            component = {
                "id": str(
                    segment.get("segment_id")
                    or segment.get("segment_key")
                    or segment.get("id")
                    or f"component_{idx + 1}"
                ),
                "display_name": display_name,
                "driver_type": str(segment.get("driver_type") or segment.get("type") or ""),
                "historical_by_period": {},
                "forecast_by_period": {},
                "driver_names": set(),
            }
            _store_period_collection(component, segment.get("historical"), is_actual=True)
            _store_period_collection(component, segment.get("historical_by_period"), is_actual=True)
            _store_period_collection(component, segment.get("historical_revenue"), is_actual=True)
            _store_period_collection(component, segment.get("forecast"), is_actual=False)
            _store_period_collection(component, segment.get("forecast_by_period"), is_actual=False)
            _store_period_collection(component, segment.get("forecast_revenue"), is_actual=False)
            for driver_key in (
                "metrics",
                "drivers",
                "volume_drivers",
                "price_drivers",
                "cost_drivers",
                "margin_drivers",
            ):
                _store_driver_collection(component, segment.get(driver_key))
            for period_bucket in ("historical_by_period", "forecast_by_period"):
                for metrics in component[period_bucket].values():
                    if isinstance(metrics, dict) and isinstance(metrics.get("drivers"), dict):
                        component["driver_names"].update(str(name) for name in metrics["drivers"])
            component["driver_names"] = sorted(component["driver_names"])
            components.append(component)
    if components:
        return components
    return [
        {
            "id": "core_revenue",
            "display_name": "Core Revenue",
            "driver_type": "growth_rate",
            "historical_by_period": {},
            "forecast_by_period": {},
            "driver_names": [],
        }
    ]


def _revenue_component_period_metrics(
    component: dict[str, Any],
    period_label: str,
    *,
    is_actual: bool,
) -> dict[str, Any]:
    bucket_name = "historical_by_period" if is_actual else "forecast_by_period"
    bucket = component.get(bucket_name)
    if not isinstance(bucket, dict):
        return {}
    normalized = _normalize_period_label(period_label, is_actual=is_actual)
    return bucket.get(normalized) if isinstance(bucket.get(normalized), dict) else {}


def _latest_component_gross_margin(component: dict[str, Any], default: float) -> float:
    historical = component.get("historical_by_period")
    if not isinstance(historical, dict):
        return default
    latest_year = -1
    latest_margin = default
    for period, metrics in historical.items():
        if not isinstance(metrics, dict):
            continue
        year = _period_year(period, 0)
        if year >= latest_year and metrics.get("gross_margin") not in (None, ""):
            latest_year = year
            latest_margin = _as_decimal(metrics.get("gross_margin"), default)
    return latest_margin


def _revenue_segment_records(payload: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = {}
    for component in _normalize_revenue_components(payload):
        records: list[dict[str, Any]] = []
        forecast = component.get("forecast_by_period")
        if isinstance(forecast, dict):
            for period, metrics in forecast.items():
                if isinstance(metrics, dict) and "revenue" in metrics:
                    records.append({"period": period, "value": metrics["revenue"]})
        result[str(component["display_name"])] = records
    return result


def _revenue_segment_value(
    payload: dict[str, Any],
    segment_index: int,
    period_label: str,
) -> float:
    components = _normalize_revenue_components(payload)
    if segment_index >= len(components):
        return 0.0
    metrics = _revenue_component_period_metrics(
        components[segment_index],
        period_label,
        is_actual=False,
    )
    return _as_float(metrics.get("revenue"), 0.0)


def _historical_revenue_segments(
    payload: dict[str, Any],
    period_label: str,
) -> list[float]:
    values: list[float] = []
    for component in _normalize_revenue_components(payload):
        metrics = _revenue_component_period_metrics(component, period_label, is_actual=True)
        values.append(_as_float(metrics.get("revenue"), 0.0))
    return values


def _assumption_forecast_values(
    payload: dict[str, Any],
    *,
    name_pattern: str,
) -> dict[str, float]:
    spec = _balance_sheet_spec(payload)
    raw = spec.get("assumption_requirements")
    if not isinstance(raw, list):
        return {}
    pattern = re.compile(name_pattern, re.IGNORECASE)
    for item in raw:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or item.get("description") or "")
        if not pattern.search(name):
            continue
        values = item.get("forecast_values")
        if not isinstance(values, dict):
            return {}
        return {
            _normalize_period_label(period, is_actual=False): _as_decimal(value, 0.0)
            for period, value in values.items()
        }
    return {}


def _gross_margin_forecast_values(payload: dict[str, Any]) -> dict[str, float]:
    spec = _income_statement_spec(payload)
    cogs = _nested_field(spec, "forecast_logic.cogs.assumptions", default=[])
    values: dict[str, float] = {}
    if isinstance(cogs, list):
        for item in cogs:
            text = str(item)
            period = re.search(r"(20\d{2}E)", text)
            margin = re.search(r"gross margin\s+([0-9.]+)%", text, re.IGNORECASE)
            if period and margin:
                values[_normalize_period_label(period.group(1), is_actual=False)] = (
                    _as_float(margin.group(1), 0.0) / 100.0
                )
    return values


def _forecast_logic_amount_values(
    payload: dict[str, Any],
    key: str,
) -> dict[str, float]:
    spec = _income_statement_spec(payload)
    assumptions = _nested_field(spec, f"forecast_logic.{key}.assumptions", default=[])
    values: dict[str, float] = {}
    if not isinstance(assumptions, list):
        return values
    for item in assumptions:
        text = str(item)
        match = re.search(r"(20\d{2}E)\s*[:：]\s*([+-]?\d+(?:\.\d+)?)", text)
        if match:
            values[_normalize_period_label(match.group(1), is_actual=False)] = _as_float(
                match.group(2),
                0.0,
            )
    return values


def _period_assumption_set(
    payload: dict[str, Any],
    latest: dict[str, Any],
    forecast_labels: list[str],
) -> dict[str, dict[str, float]]:
    base = _assumption_set(payload, latest)
    revenue_records = {
        _normalize_period_label(record.get("period"), is_actual=False): record
        for record in _forecast_revenue_records(payload)
    }
    gross_margin = _gross_margin_forecast_values(payload)
    ar_days = _assumption_forecast_values(payload, name_pattern=r"\bAR\b|DSO")
    inventory_days = _assumption_forecast_values(payload, name_pattern=r"Inventory|DIO")
    ap_days = _assumption_forecast_values(payload, name_pattern=r"\bAP\b|DPO")
    capex_pct = _assumption_forecast_values(payload, name_pattern=r"CapEx")
    dividend_payout = _assumption_forecast_values(payload, name_pattern=r"Dividend")
    net_finance_expense = _forecast_logic_amount_values(payload, "finance_expenses")

    assumptions: dict[str, dict[str, float]] = {}
    prior_revenue = _historical_value(latest, "revenue")
    for label in forecast_labels:
        current = dict(base)
        record = revenue_records.get(label)
        if record is not None:
            revenue = _as_float(record.get("total"), 0.0)
            if revenue and prior_revenue:
                current["revenue_growth"] = revenue / prior_revenue - 1
            current["revenue_total"] = revenue
            prior_revenue = revenue or prior_revenue
        current["gross_margin"] = gross_margin.get(label, current["gross_margin"])
        current["ar_days"] = _as_float(ar_days.get(label), current["ar_days"])
        current["inventory_days"] = _as_float(
            inventory_days.get(label),
            _latest_ratio(
                _historical_value(latest, "inventory"),
                _historical_value(latest, "cogs"),
                0.0,
            )
            * 365
            if _historical_value(latest, "cogs")
            else 90.0,
        )
        current["ap_days"] = _as_float(ap_days.get(label), current["ap_days"])
        current["capex_pct_revenue"] = capex_pct.get(label, current["capex_pct_revenue"])
        current["dividend_payout_pct"] = dividend_payout.get(
            label,
            current["dividend_payout_pct"] or 0.35,
        )
        current["net_finance_expense"] = net_finance_expense.get(
            label,
            current["net_finance_expense"],
        )
        assumptions[label] = current
    return assumptions


def _latest_ratio(numerator: float, denominator: float, default: float) -> float:
    if denominator == 0:
        return default
    return numerator / denominator


def _normalize_model_unit(value: Any) -> str:
    text = str(value or "").strip()
    if "亿元" in text:
        return "亿元"
    text = re.sub(r"\s*\([^)]*\)\s*", "", text).strip()
    return text or "millions"


def _model_metadata(payload: dict[str, Any]) -> dict[str, str]:
    company = payload.get("company")
    if isinstance(company, dict):
        company_name = str(company.get("short_name") or company.get("legal_name") or "Company")
        ticker = str(company.get("ticker") or _field(payload, "ticker", "symbol", default="TICKER"))
        market = str(company.get("market") or company.get("exchange") or _field(payload, "market", "exchange", default=""))
        currency = str(company.get("currency") or _field(payload, "currency", "reporting_currency", default="USD"))
        unit = str(company.get("reporting_unit") or _field(payload, "unit", "reporting_unit", default="millions"))
        fiscal_year_end = str(company.get("fiscal_year_end") or _field(payload, "fiscal_year_end", default="Dec"))
    else:
        context = payload.get("task2_context_packet")
        context_meta = context.get("company_metadata") if isinstance(context, dict) else {}
        if not isinstance(context_meta, dict):
            context_meta = {}
        company_name = str(
            _field(payload, "company", "company_name", default="")
            or _field(context_meta, "company", default="Company")
        )
        ticker = str(
            _field(payload, "ticker", "symbol", default="")
            or _field(context_meta, "ticker", default="TICKER")
        )
        market = str(
            _field(payload, "market", "exchange", default="")
            or _field(context_meta, "market", default="")
        )
        currency = str(
            _field(payload, "currency", "reporting_currency", default="")
            or _field(context_meta, "currency", default="USD")
        )
        unit = str(
            _field(payload, "unit", "reporting_unit", default="")
            or _field(context_meta, "reporting_unit", default="millions")
        )
        fiscal_year_end = str(
            _field(payload, "fiscal_year_end", default="")
            or _field(context_meta, "fiscal_year_end", default="Dec")
        )
    return {
        "company": company_name,
        "ticker": ticker,
        "market": market,
        "currency": currency,
        "unit": _normalize_model_unit(unit),
        "fiscal_year_end": fiscal_year_end,
    }


def _historical_value(record: dict[str, Any], key: str) -> float:
    aliases: dict[str, tuple[str, ...]] = {
        "revenue": ("revenue", "revenue_total", "net_revenue"),
        "cogs": ("cogs", "cost_of_revenue"),
        "gross_profit": ("gross_profit",),
        "operating_expenses": ("operating_expenses", "opex", "sgna"),
        "da": ("da", "d_and_a", "depreciation_amortization"),
        "ebit": ("ebit", "operating_income"),
        "ebitda": ("ebitda",),
        "interest_expense": ("interest_expense", "interest"),
        "interest_income": ("interest_income",),
        "net_finance_expense": ("net_finance_expense", "finance_expenses"),
        "pretax_income": ("pretax_income", "ebt", "income_before_tax"),
        "tax_expense": ("tax_expense", "tax"),
        "net_income": ("net_income",),
        "cash": ("cash", "cash_and_equivalents"),
        "ar": ("accounts_receivable", "ar"),
        "inventory": ("inventory",),
        "ap": ("accounts_payable", "ap"),
        "capex": ("capex", "capital_expenditures"),
        "ppe": ("ppe", "net_ppe", "property_plant_equipment"),
        "other_assets": ("other_assets",),
        "total_current_assets": ("total_current_assets",),
        "total_assets": ("total_assets",),
        "current_liabilities": ("current_liabilities", "total_current_liabilities"),
        "total_liabilities": ("total_liabilities",),
        "total_equity": ("total_equity",),
        "debt": ("debt", "total_debt"),
        "short_term_debt_raw": ("short_term_debt_raw", "short_term_debt"),
        "long_term_debt_raw": ("long_term_debt_raw", "long_term_debt"),
        "debt_to_market_equity": ("debt_to_market_equity",),
        "retained_earnings": ("retained_earnings",),
        "shares": ("diluted_shares", "shares_outstanding", "shares"),
    }
    return _as_float(_field(record, *aliases.get(key, (key,))), 0.0)


def _assumption_set(payload: dict[str, Any], latest: dict[str, Any]) -> dict[str, float]:
    assumptions = payload.get("assumptions")
    if not isinstance(assumptions, dict):
        assumptions = {}
    revenue = _historical_value(latest, "revenue")
    gross_profit = _historical_value(latest, "gross_profit") or revenue * 0.5
    opex = _historical_value(latest, "operating_expenses")
    da = _historical_value(latest, "da")
    capex = abs(_historical_value(latest, "capex"))
    pretax = _historical_value(latest, "pretax_income")
    tax = _historical_value(latest, "tax_expense")
    return {
        "revenue_growth": _as_decimal(assumptions.get("revenue_growth"), 0.05),
        "gross_margin": _as_decimal(
            assumptions.get("gross_margin"),
            _latest_ratio(gross_profit, revenue, 0.5),
        ),
        "opex_pct_revenue": _as_decimal(
            assumptions.get("opex_pct_revenue"),
            _latest_ratio(opex, revenue, 0.2),
        ),
        "tax_rate": _as_decimal(
            assumptions.get("tax_rate"),
            _latest_ratio(tax, pretax, 0.25),
        ),
        "ar_days": _as_float(assumptions.get("ar_days"), 45.0),
        "inventory_pct_revenue": _as_decimal(
            assumptions.get("inventory_pct_revenue"),
            0.1,
        ),
        "ap_days": _as_float(assumptions.get("ap_days"), 30.0),
        "da_pct_revenue": _as_decimal(
            assumptions.get("da_pct_revenue"),
            _latest_ratio(da, revenue, 0.03),
        ),
        "capex_pct_revenue": _as_decimal(
            assumptions.get("capex_pct_revenue"),
            _latest_ratio(capex, revenue, 0.04),
        ),
        "interest_rate": _as_decimal(assumptions.get("interest_rate"), 0.05),
        "net_finance_expense": _as_float(
            assumptions.get("net_finance_expense"),
            _historical_value(latest, "net_finance_expense"),
        ),
        "debt_repayment_pct": _as_decimal(
            assumptions.get("debt_repayment_pct"),
            0.0,
        ),
        "diluted_shares_growth": _as_decimal(
            assumptions.get("diluted_shares_growth"),
            0.0,
        ),
        "dividend_payout_pct": _as_decimal(
            assumptions.get("dividend_payout_pct"),
            0.0,
        ),
    }


def _validation_result(
    status: str,
    critical: list[dict[str, str]],
    warnings: list[dict[str, str]],
) -> str:
    return _json_result(
        {
            "status": status,
            "critical_count": len(critical),
            "warning_count": len(warnings),
            "critical": critical,
            "warnings": warnings,
        }
    )


def _read_json_file(path: Path) -> Any | None:
    text = read_text_artifact(path, missing_ok=True)
    if text is None:
        return None
    return json.loads(text)


def _statement_model_dir(run_dir: Path) -> Path:
    return run_dir / "02_financial_model"


def _task1_dir(run_dir: Path) -> Path:
    return run_dir / "01_company_research"


def _json_result(data: dict[str, Any]) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2)


def _inline_json_failure(data_json: str, field_name: str) -> dict[str, Any] | None:
    size = len((data_json or "").encode("utf-8"))
    if size <= INLINE_JSON_MAX_BYTES:
        return None
    return {
        "status": "FAIL",
        "field": field_name,
        "message": (
            f"{field_name} is {size} bytes, above the {INLINE_JSON_MAX_BYTES} byte "
            "inline JSON limit. Use compact structured input or an existing artifact path."
        ),
    }


def _required_task1_paths(run_dir: Path) -> dict[str, Path]:
    task1_dir = _task1_dir(run_dir)
    return {name: task1_dir / name for name in TASK1_REQUIRED_ARTIFACTS}


def _task1_missing_paths(run_dir: Path) -> list[str]:
    return [
        _relative_to_workspace(path)
        for path in _required_task1_paths(run_dir).values()
        if not artifact_exists(path)
    ]


def _infer_run_dir_from_task1_path(path: Path) -> Path:
    if path.name in TASK1_REQUIRED_ARTIFACTS:
        path = path.parent
    return path.parent if path.name == "01_company_research" else path


def _latest_task1_run(
    *,
    ticker: str,
    market: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> Path | None:
    runs_dir = _coverage_dir(market, ticker, output_dir) / "runs"
    candidates = [
        path
        for path in list_artifact_dir(runs_dir)
        if not _task1_missing_paths(path)
    ]
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: item.name)[-1]


def _wrong_root_artifacts_for(run_dir: Path) -> list[dict[str, str]]:
    wrong_run_dir = _project_wrong_root_path(run_dir)
    if wrong_run_dir is None or not wrong_run_dir.exists():
        return []

    artifacts: list[dict[str, str]] = []
    tracked = [
        *(wrong_run_dir / "01_company_research" / name for name in TASK1_REQUIRED_ARTIFACTS),
        *(
            wrong_run_dir / "02_financial_model" / name
            for name in (
                *TASK2_STATEMENT_ARTIFACTS.values(),
                "revenue_build_spec.json",
                *TASK2_MODEL_ARTIFACTS,
            )
        ),
    ]
    for wrong_path in tracked:
        if not wrong_path.exists():
            continue
        canonical_path = _workspace_root() / wrong_path.relative_to(_project_root())
        severity = "critical" if not canonical_path.exists() else "warning"
        artifacts.append(
            {
                "severity": severity,
                "wrong_root_path": str(wrong_path),
                "canonical_path": _relative_to_workspace(canonical_path),
            }
        )
    return artifacts


def _artifact_record(path: Path) -> dict[str, Any]:
    exists = artifact_exists(path)
    record: dict[str, Any] = {
        "path": _relative_to_workspace(path),
        "exists": exists,
    }
    if path.suffix == ".json" and exists:
        try:
            parsed = _read_json_file(path)
        except Exception as exc:
            record.update({"json_valid": False, "error": str(exc)})
        else:
            record["json_valid"] = isinstance(parsed, (dict, list))
    return record


def _append_finding(
    findings: list[dict[str, str]],
    *,
    category: str,
    issue: str,
    path: Path | None = None,
) -> None:
    finding = {"category": category, "issue": issue}
    if path is not None:
        finding["path"] = _relative_to_workspace(path)
    findings.append(finding)


def _safe_audit_text(value: Any, limit: int = 700) -> str:
    text = str(value or "").replace("\r", " ").strip()
    blocked = (
        "SystemMessage",
        "HumanMessage",
        "AIMessage",
        "ToolMessage",
        "inputs.messages",
        '"messages"',
        '"generations"',
        "langchain.schema.messages",
    )
    if any(token in text for token in blocked):
        return "[omitted runtime trace]"
    text = re.sub(r"\s+", " ", text)
    return text[:limit]


def _normalize_findings(raw: Any, *, limit: int = 50) -> list[dict[str, str]]:
    if not isinstance(raw, list):
        return []
    findings: list[dict[str, str]] = []
    for item in raw[:limit]:
        if isinstance(item, dict):
            category = _safe_audit_text(item.get("category") or item.get("type") or "Finding")
            issue = _safe_audit_text(item.get("issue") or item.get("message") or item)
            path = _safe_audit_text(item.get("path") or "")
            finding = {"category": category, "issue": issue}
            if path:
                finding["path"] = path
        else:
            finding = {"category": "Finding", "issue": _safe_audit_text(item)}
        findings.append(finding)
    return findings


def _load_task2_model_sources(run_dir: Path) -> dict[str, Any]:
    model_dir = _statement_model_dir(run_dir)
    paths = {
        "financial_facts": model_dir / "financial_facts.json",
        "task2_context_packet": model_dir / "task2_context_packet.json",
        "statement_spec_pack": model_dir / "statement_spec_pack.json",
        "revenue_build_spec": model_dir / "revenue_build_spec.json",
    }
    financial_facts = _read_json_file(paths["financial_facts"])
    context_packet = _read_json_file(paths["task2_context_packet"])
    statement_pack = _read_json_file(paths["statement_spec_pack"])
    revenue_build_spec = _read_json_file(paths["revenue_build_spec"])
    if not isinstance(financial_facts, dict):
        raise ValueError(
            f"financial_facts.json must be a JSON object at {paths['financial_facts']}"
        )
    if not isinstance(context_packet, dict):
        context_packet = {}
    if not isinstance(statement_pack, dict):
        statement_pack = {}
    if not isinstance(revenue_build_spec, dict):
        revenue_build_spec = {}
    payload: dict[str, Any] = {
        **financial_facts,
        "financial_facts": financial_facts,
        "task2_context_packet": context_packet,
    }
    if revenue_build_spec:
        payload["revenue_build_spec"] = revenue_build_spec
    if statement_pack:
        payload["statement_spec_pack"] = statement_pack
    return {
        "run_dir": run_dir,
        "model_dir": model_dir,
        "paths": paths,
        "financial_facts": financial_facts,
        "task2_context_packet": context_packet,
        "statement_spec_pack": statement_pack,
        "revenue_build_spec": revenue_build_spec,
        "payload": payload,
    }


def _statement_metadata_from_sources(sources: dict[str, Any]) -> dict[str, str]:
    return _model_metadata(sources["payload"])


def _statement_source_coverage(sources: dict[str, Any], statement_type: str) -> dict[str, Any]:
    facts = sources["financial_facts"]
    context = sources["task2_context_packet"]
    coverage: dict[str, Any] = {}
    raw_context_coverage = context.get("source_coverage")
    if isinstance(raw_context_coverage, dict):
        raw = raw_context_coverage.get(statement_type) or raw_context_coverage
        if isinstance(raw, dict):
            coverage.update(raw)
    if not coverage:
        coverage = {
            "status": "derived_from_financial_facts",
            "source_count": len(facts.get("sources") or []),
        }
    return coverage


def _statement_record_value(
    statement_type: str,
    canonical_key: str,
    record: dict[str, Any],
    previous_record: dict[str, Any] | None,
) -> float:
    if statement_type == "income_statement":
        mapping = {
            "revenue_total": "revenue",
            "gross_profit": "gross_profit",
            "ebit": "ebit",
            "ebitda": "ebitda",
            "interest_expense": "interest_expense",
            "pretax_income": "pretax_income",
            "tax_expense": "tax_expense",
            "net_income": "net_income",
            "da_total": "da",
        }
        return _historical_value(record, mapping[canonical_key])
    if statement_type == "balance_sheet":
        mapping = {
            "cash_and_equivalents": "cash",
            "total_current_assets": "total_current_assets",
            "total_assets": "total_assets",
            "total_current_liabilities": "current_liabilities",
            "total_debt": "debt",
            "retained_earnings": "retained_earnings",
            "total_equity": "total_equity",
        }
        if canonical_key == "total_liabilities_and_equity":
            total_assets = _historical_value(record, "total_assets")
            if total_assets:
                return total_assets
            return _historical_value(record, "total_liabilities") + _historical_value(
                record,
                "total_equity",
            )
        return _historical_value(record, mapping[canonical_key])

    if canonical_key == "net_income_cf":
        return _historical_value(record, "net_income")
    if canonical_key == "da_addback":
        return _historical_value(record, "da")
    if canonical_key == "nwc_change":
        return _historical_value(record, "nwc_change")
    if canonical_key == "cfo_total":
        explicit = _historical_value(record, "cfo_total")
        if explicit:
            return explicit
        return _historical_value(record, "net_income") + _historical_value(
            record,
            "da",
        ) + _historical_value(record, "nwc_change")
    if canonical_key == "capex":
        return _historical_value(record, "capex")
    if canonical_key == "cfi_total":
        explicit = _historical_value(record, "cfi_total")
        return explicit if explicit else -abs(_historical_value(record, "capex"))
    if canonical_key == "debt_proceeds_repayments":
        return _historical_value(record, "debt_proceeds_repayments")
    if canonical_key == "dividends":
        return _historical_value(record, "dividends")
    if canonical_key == "cff_total":
        explicit = _historical_value(record, "cff_total")
        if explicit:
            return explicit
        return _historical_value(record, "debt_proceeds_repayments") - _historical_value(
            record,
            "dividends",
        )
    if canonical_key == "beginning_cash":
        return _historical_value(previous_record or {}, "cash")
    if canonical_key == "ending_cash":
        return _historical_value(record, "cash")
    return 0.0


def _minimal_revenue_build_spec(sources: dict[str, Any]) -> dict[str, Any]:
    existing = _revenue_build_spec(sources["payload"])
    if existing:
        return existing
    return {
        "statement_type": "revenue_build",
        "segments": [],
        "forecast": [],
        "total_revenue_reconciliation": {"forecast": []},
    }


def _build_statement_spec_from_sources(
    statement_type: str,
    sources: dict[str, Any],
) -> dict[str, Any]:
    metadata = _statement_metadata_from_sources(sources)
    records = _historical_records(sources["payload"])
    canonical_keys = list(STATEMENT_CANONICAL_KEYS[statement_type])
    historical_inputs: list[dict[str, Any]] = []
    for idx, record in enumerate(records):
        previous = records[idx - 1] if idx else None
        for key in canonical_keys:
            historical_inputs.append(
                {
                    "period": record["period"],
                    "canonical_key": key,
                    "value": _statement_record_value(
                        statement_type,
                        key,
                        record,
                        previous,
                    ),
                    "source": record.get("source") or "[UNSOURCED]",
                    "currency": metadata["currency"],
                    "unit": metadata["unit"],
                }
            )

    spec: dict[str, Any] = {
        **metadata,
        "statement_type": statement_type,
        "canonical_row_keys": canonical_keys,
        "line_items": [
            {
                "canonical_key": key,
                "display_name": key.replace("_", " ").title(),
            }
            for key in canonical_keys
        ],
        "historical_inputs": historical_inputs,
        "forecast_logic": {"method": "derived_from_task2_model_sources"},
        "assumption_requirements": [],
        "cross_statement_dependencies": list(STATEMENT_DEPENDENCY_HINTS[statement_type]),
        "source_coverage": _statement_source_coverage(sources, statement_type),
        "unsourced_items": list(
            sources["financial_facts"].get("unsourced")
            or sources["financial_facts"].get("unsourced_items")
            or []
        ),
        "validation_status": "tool_generated",
    }
    if statement_type == "income_statement":
        spec["revenue_build_spec"] = _minimal_revenue_build_spec(sources)
    return spec


def _statement_context_payload(statement_type: str, run_dir: Path) -> dict[str, Any]:
    if statement_type not in STATEMENT_JSON_ALLOWED_TYPES:
        raise ValueError(
            "statement_type must be one of: "
            + ", ".join(STATEMENT_JSON_ALLOWED_TYPES)
        )

    task1_dir = _task1_dir(run_dir)
    paths = {
        "business_driver_map": task1_dir / "business_driver_map.json",
        "source_log": task1_dir / "source_log.json",
        "company_research": task1_dir / "company_research.md",
    }
    sources = _load_task2_model_sources(run_dir)
    facts = sources["financial_facts"]
    context = sources["task2_context_packet"]
    records = _historical_records(sources["payload"])
    artifact_paths = {
        **paths,
        **sources["paths"],
    }
    missing = [
        _relative_to_workspace(path)
        for name, path in paths.items()
        if not artifact_exists(path)
    ]
    return {
        "statement_type": statement_type,
        "run_dir": _relative_to_workspace(run_dir),
        "required_fields": list(STATEMENT_JSON_REQUIRED_FIELDS),
        "canonical_row_keys": list(STATEMENT_CANONICAL_KEYS[statement_type]),
        "cross_statement_dependency_hints": list(
            STATEMENT_DEPENDENCY_HINTS[statement_type]
        ),
        "metadata": _statement_metadata_from_sources(sources),
        "periods": [record["period"] for record in records],
        "historical_facts": [
            {
                "period": record["period"],
                "year": record["year"],
                "source": record.get("source") or "[UNSOURCED]",
                **{
                    key: _statement_record_value(statement_type, key, record, None)
                    for key in STATEMENT_CANONICAL_KEYS[statement_type]
                },
            }
            for record in records
        ],
        "source_coverage": _statement_source_coverage(sources, statement_type),
        "unsourced": list(
            facts.get("unsourced")
            or facts.get("unsourced_items")
            or context.get("unsourced")
            or []
        ),
        "artifact_paths": {
            name: _relative_to_workspace(path)
            for name, path in artifact_paths.items()
        },
        "artifacts": {
            name: _artifact_record(path)
            for name, path in artifact_paths.items()
        },
        "missing_artifacts": missing,
    }


def _normalize_statement_type(statement_type: str) -> str:
    aliases = {
        "is": "income_statement",
        "income_statement": "income_statement",
        "bs": "balance_sheet",
        "balance_sheet": "balance_sheet",
        "cf": "cash_flow",
        "cash_flow": "cash_flow",
        "cash_flow_statement": "cash_flow",
    }
    normalized = aliases.get(str(statement_type).strip().lower())
    if normalized is None:
        raise ValueError(
            "statement_type must be one of: income_statement, balance_sheet, cash_flow"
        )
    return normalized


def _canonical_keys(payload: dict[str, Any]) -> set[str]:
    raw = payload.get("canonical_row_keys")
    if isinstance(raw, dict):
        return {str(key) for key in raw}
    if isinstance(raw, list):
        return {str(item) for item in raw}
    return set()


def _supplemental_keys(payload: dict[str, Any]) -> set[str]:
    keys: set[str] = set()
    raw = payload.get("supplemental_line_items") or []
    if isinstance(raw, dict):
        raw = list(raw.values())
    if not isinstance(raw, list):
        return keys
    for item in raw:
        if isinstance(item, dict):
            key = _field(
                item,
                "canonical_key",
                "line_item_key",
                "key",
                "name",
                default="",
            )
            if key:
                keys.add(str(key))
        elif item:
            keys.add(str(item))
    return keys


def _dependency_values(payload: dict[str, Any]) -> list[str]:
    raw = payload.get("cross_statement_dependencies")
    if isinstance(raw, dict):
        values: list[str] = []
        for key, value in raw.items():
            values.append(str(key))
            if isinstance(value, list):
                values.extend(str(item) for item in value)
            elif isinstance(value, dict):
                values.extend(str(item) for item in value.values())
            else:
                values.append(str(value))
        return values
    if isinstance(raw, list):
        return [str(item) for item in raw]
    return []


def _historical_input_records(payload: dict[str, Any]) -> list[dict[str, Any]]:
    raw = payload.get("historical_inputs")
    if not isinstance(raw, list):
        return []
    return [item for item in raw if isinstance(item, dict)]


def _historical_input_key(record: dict[str, Any]) -> str:
    return str(
        _field(
            record,
            "canonical_key",
            "canonical_row_key",
            "line_item_key",
            "key",
            "name",
            default="",
        )
        or ""
    )


def _model_field_for_canonical_key(canonical_key: str) -> str | None:
    mapping = {
        "revenue_total": "revenue",
        "gross_profit": "gross_profit",
        "ebit": "ebit",
        "ebitda": "ebitda",
        "interest_expense": "interest_expense",
        "pretax_income": "pretax_income",
        "tax_expense": "tax_expense",
        "net_income": "net_income",
        "da_total": "da",
        "cash_and_equivalents": "cash",
        "accounts_receivable": "ar",
        "inventory": "inventory",
        "accounts_payable": "ap",
        "capex": "capex",
        "net_ppe": "ppe",
        "total_debt": "debt",
        "retained_earnings": "retained_earnings",
        "diluted_shares": "shares",
    }
    return mapping.get(canonical_key)


def _validate_statement_historical_inputs(
    payload: dict[str, Any],
    expected_statement_type: str,
    critical: list[dict[str, str]],
    warnings: list[dict[str, str]],
) -> None:
    records = _historical_input_records(payload)
    if not records:
        critical.append(
            {
                "category": "Historical Inputs",
                "issue": "historical_inputs must include sourced records.",
            }
        )
        return

    canonical_keys = _canonical_keys(payload)
    supplemental_keys = _supplemental_keys(payload)
    for idx, record in enumerate(records):
        missing: list[str] = []
        if not record.get("period"):
            missing.append("period")
        key = _historical_input_key(record)
        if not key:
            missing.append("canonical_key")
        elif key not in canonical_keys:
            parent_key = str(
                _field(
                    record,
                    "parent_canonical_key",
                    "parent_key",
                    "aggregate_key",
                    default="",
                )
                or ""
            )
            if parent_key in canonical_keys or key in supplemental_keys:
                pass
            else:
                warnings.append(
                    {
                        "category": "Historical Inputs",
                        "issue": (
                            f"historical_inputs[{idx}] canonical_key '{key}' is not "
                            f"listed in {expected_statement_type} canonical_row_keys "
                            "and has no valid parent_canonical_key."
                        ),
                    }
                )
        if "value" not in record and "amount" not in record:
            missing.append("value")
        if not _field(record, "source", "source_text", default=""):
            missing.append("source")
        if missing:
            critical.append(
                {
                    "category": "Historical Inputs",
                    "issue": (
                        f"historical_inputs[{idx}] missing required field(s): "
                        + ", ".join(missing)
                    ),
                }
            )


def _derive_financial_facts(
    *,
    specs: dict[str, dict[str, Any]],
    ticker: str,
    market: str,
) -> dict[str, Any]:
    metadata_source = next(iter(specs.values()), {})
    periods: dict[str, dict[str, Any]] = {}
    sources: list[str] = []
    unsourced: list[str] = []

    for statement_type, payload in specs.items():
        for item in _historical_input_records(payload):
            period = str(item.get("period") or "")
            if not period:
                continue
            record = periods.setdefault(
                period,
                {
                    "period": period,
                    "year": _period_year(period, 2020 + len(periods)),
                    "source": "[UNSOURCED]",
                },
            )
            source = str(_field(item, "source", "source_text", default="[UNSOURCED]"))
            if source and source != "[UNSOURCED]":
                sources.append(source)
                if record.get("source") == "[UNSOURCED]":
                    record["source"] = source
            elif source == "[UNSOURCED]":
                unsourced.append(f"{statement_type}:{period}:{_historical_input_key(item)}")

            model_field = _model_field_for_canonical_key(_historical_input_key(item))
            if not model_field:
                parent_key = str(
                    _field(
                        item,
                        "parent_canonical_key",
                        "parent_key",
                        "aggregate_key",
                        default="",
                    )
                    or ""
                )
                model_field = _model_field_for_canonical_key(parent_key)
            if model_field:
                record[model_field] = _as_float(
                    _field(item, "value", "amount", default=0.0),
                    0.0,
                )

        for item in payload.get("unsourced_items") or []:
            unsourced.append(f"{statement_type}:{item}")

    historicals = sorted(periods.values(), key=lambda item: item["year"])
    return {
        "company": str(_field(metadata_source, "company", "company_name", default="Company")),
        "ticker": str(_field(metadata_source, "ticker", "symbol", default=ticker)),
        "market": str(_field(metadata_source, "market", "exchange", default=market)),
        "currency": str(_field(metadata_source, "currency", "reporting_currency", default="USD")),
        "unit": str(_field(metadata_source, "unit", "reporting_unit", default="millions")),
        "fiscal_year_end": str(_field(metadata_source, "fiscal_year_end", default="Dec")),
        "historicals": historicals,
        "segments": [],
        "guidance": [],
        "projection_summary": {},
        "assumptions": dict(metadata_source.get("assumptions") or {}),
        "sources": sorted(set(sources)),
        "unsourced": sorted(set(unsourced)),
    }


def _derive_task2_context_packet(
    *,
    specs: dict[str, dict[str, Any]],
    pack: dict[str, Any],
    financial_facts: dict[str, Any],
) -> dict[str, Any]:
    return {
        "company": financial_facts.get("company", "Company"),
        "ticker": financial_facts.get("ticker", ""),
        "market": financial_facts.get("market", ""),
        "currency": financial_facts.get("currency", "USD"),
        "unit": financial_facts.get("unit", "millions"),
        "fiscal_year_end": financial_facts.get("fiscal_year_end", "Dec"),
        "periods": [item.get("period") for item in financial_facts.get("historicals", [])],
        "canonical_row_keys": {
            statement_type: list(STATEMENT_CANONICAL_KEYS[statement_type])
            for statement_type in STATEMENT_JSON_ALLOWED_TYPES
        },
        "source_coverage": {
            statement_type: payload.get("source_coverage", {})
            for statement_type, payload in specs.items()
        },
        "unsourced": financial_facts.get("unsourced", []),
        "reconciliation_status": pack.get("status"),
        "critical_count": pack.get("critical_count", 0),
        "warning_count": pack.get("warning_count", 0),
    }


def _validate_statement_payload(
    payload: dict[str, Any],
    expected_statement_type: str,
) -> dict[str, Any]:
    critical: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []

    actual_type = payload.get("statement_type")
    if actual_type != expected_statement_type:
        critical.append(
            {
                "category": "Statement Type",
                "issue": (
                    f"statement_type must be '{expected_statement_type}', "
                    f"got '{actual_type}'."
                ),
            }
        )

    for field_name in STATEMENT_JSON_REQUIRED_FIELDS:
        if field_name not in payload:
            critical.append(
                {
                    "category": "Missing Required Field",
                    "issue": f"Missing required field: {field_name}",
                }
            )

    required_keys = set(STATEMENT_CANONICAL_KEYS[expected_statement_type])
    present_keys = _canonical_keys(payload)
    missing_keys = sorted(required_keys - present_keys)
    if missing_keys:
        critical.append(
            {
                "category": "Canonical Row Keys",
                "issue": "Missing canonical row keys: " + ", ".join(missing_keys),
            }
        )

    _validate_statement_historical_inputs(
        payload,
        expected_statement_type,
        critical,
        warnings,
    )

    dependencies = " ".join(_dependency_values(payload)).lower()
    if not dependencies:
        critical.append(
            {
                "category": "Cross Statement Dependencies",
                "issue": "cross_statement_dependencies must declare tie-out links.",
            }
        )
    else:
        for hint in STATEMENT_DEPENDENCY_HINTS[expected_statement_type]:
            if hint.lower() not in dependencies:
                warnings.append(
                    {
                        "category": "Dependency Coverage",
                        "issue": f"Expected dependency hint not found: {hint}",
                    }
                )

    source_coverage = payload.get("source_coverage")
    if not source_coverage:
        critical.append(
            {
                "category": "Source Coverage",
                "issue": "source_coverage is required and cannot be empty.",
            }
        )

    unsourced = payload.get("unsourced_items")
    if unsourced:
        warnings.append(
            {
                "category": "Unsourced Items",
                "issue": f"{len(unsourced)} unsourced item(s) require model_audit.md disclosure.",
            }
        )

    status = "PASS" if not critical else "FAIL"
    return {
        "status": status,
        "statement_type": expected_statement_type,
        "critical_count": len(critical),
        "warning_count": len(warnings),
        "critical": critical,
        "warnings": warnings,
    }


def _statement_payload_from_run(run_dir: str, statement_type: str) -> dict[str, Any]:
    out_dir = _resolve_workspace_path(run_dir)
    sources = _load_task2_model_sources(out_dir)
    return _build_statement_spec_from_sources(statement_type, sources)


def _validate_statement_json(run_dir: str, expected_statement_type: str) -> str:
    payload = _statement_payload_from_run(run_dir, expected_statement_type)
    return _json_result(_validate_statement_payload(payload, expected_statement_type))


def _write_statement_json(
    *,
    statement_type: str,
    ticker: str,
    market: str,
    run_dir: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    out_dir = _find_run_dir(
        market=market,
        ticker=ticker,
        output_dir=output_dir,
        run_dir=run_dir,
    )
    sources = _load_task2_model_sources(out_dir)
    payload = _build_statement_spec_from_sources(statement_type, sources)
    validation = _validate_statement_payload(payload, statement_type)
    if validation["critical_count"]:
        return _json_result(
            {
                "status": "FAIL",
                "statement_type": statement_type,
                "validation": validation,
                "written": [],
            }
        )

    model_dir = _statement_model_dir(out_dir)
    _ensure_dir(model_dir)
    written: list[str] = []

    path = model_dir / STATEMENT_JSON_OUTPUTS[statement_type]
    _write_text(path, _json_result(payload) + "\n")
    written.append(_relative_to_workspace(path))

    if statement_type == "income_statement" and isinstance(
        payload.get("revenue_build_spec"),
        dict,
    ):
        revenue_path = model_dir / "revenue_build_spec.json"
        _write_text(revenue_path, _json_result(payload["revenue_build_spec"]) + "\n")
        written.append(_relative_to_workspace(revenue_path))

    return _json_result(
        {
            "status": "OK",
            "statement_type": statement_type,
            "validation": validation,
            "written": written,
        }
    )


def _timestamp() -> str:
    return (
        os.getenv("SINGLE_STOCK_COVERAGE_OUTPUT_TIMESTAMP")
        or datetime.now().strftime("%Y%m%d-%H%M%S")
    )


def _find_run_dir(
    *,
    market: str,
    ticker: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
    run_dir: str = "",
) -> Path:
    if run_dir:
        return _resolve_workspace_path(run_dir)

    key = f"{_safe_market(market)}:{_safe_ticker(ticker)}:{output_dir}"
    existing = _ACTIVE_RUNS.get(key)
    if existing:
        _ensure_dir(existing)
        return existing

    return _create_run_dir(
        company="",
        ticker=ticker,
        market=market,
        task_type="unspecified",
        triggering_event="",
        output_dir=output_dir,
    )


def _create_run_dir(
    *,
    company: str,
    ticker: str,
    market: str,
    task_type: str,
    triggering_event: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> Path:
    coverage_dir = _coverage_dir(market, ticker, output_dir)
    runs_dir = coverage_dir / "runs"
    _ensure_dir(runs_dir)

    base_timestamp = _timestamp()
    candidate = runs_dir / base_timestamp
    if artifact_exists(candidate) and os.getenv("SINGLE_STOCK_COVERAGE_OUTPUT_TIMESTAMP") is None:
        suffix = 2
        while artifact_exists(runs_dir / f"{base_timestamp}-{suffix}"):
            suffix += 1
        candidate = runs_dir / f"{base_timestamp}-{suffix}"
    _ensure_dir(candidate)

    key = f"{_safe_market(market)}:{_safe_ticker(ticker)}:{output_dir}"
    _ACTIVE_RUNS[key] = candidate

    manifest_path = candidate / "run_manifest.json"
    if not artifact_exists(manifest_path):
        manifest = {
            "run_id": candidate.name,
            "company": company,
            "ticker": ticker,
            "market": market,
            "task_type": task_type,
            "triggering_event": triggering_event,
            "subagents_called": [],
            "input_artifacts": [],
            "output_artifacts": [],
            "final_conclusion": "",
            "unsourced": [],
            "follow_up_checklist": [],
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        _write_text(
            manifest_path,
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        )

    coverage_state = coverage_dir / "coverage_state.json"
    if not artifact_exists(coverage_state):
        state = {
            "company": company,
            "ticker": ticker,
            "market": market,
            "coverage_status": "created",
            "latest_run": _relative_to_workspace(candidate),
            "latest_model_path": "",
            "latest_valuation_state": "",
            "latest_price_target": "",
            "latest_rating": "",
            "latest_recommendation": "",
            "latest_thesis_pillars": [],
            "key_assumptions": [],
            "next_catalysts": [],
            "stale_data_flags": [],
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        _write_text(
            coverage_state,
            json.dumps(state, ensure_ascii=False, indent=2) + "\n",
        )

    return candidate


@tool
def create_coverage_run_dir(
    company: str,
    ticker: str,
    market: str,
    task_type: str = "initiation",
    triggering_event: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Create or return a timestamped run directory for one stock coverage task.

    Args:
        company: Company name.
        ticker: Exchange ticker or code.
        market: Listing market, such as A-share, HK, US, or ADR.
        task_type: initiation, update, valuation_refresh, model_audit, etc.
        triggering_event: Optional event that triggered this run.
        output_dir: Coverage root relative to workspace. Defaults to ./out/coverage.

    Returns:
        JSON with workspace-relative coverage_dir, run_dir, manifest_path, and
        coverage_state_path.
    """
    run_dir = _create_run_dir(
        company=company,
        ticker=ticker,
        market=market,
        task_type=task_type,
        triggering_event=triggering_event,
        output_dir=output_dir,
    )
    coverage_dir = run_dir.parents[1]
    result = {
        "coverage_dir": _relative_to_workspace(coverage_dir),
        "run_dir": _relative_to_workspace(run_dir),
        "manifest_path": _relative_to_workspace(run_dir / "run_manifest.json"),
        "coverage_state_path": _relative_to_workspace(coverage_dir / "coverage_state.json"),
    }
    return json.dumps(result, ensure_ascii=False)


@tool
def write_markdown_artifact(
    markdown: str,
    filename: str,
    subdir: str,
    ticker: str,
    market: str,
    run_dir: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Write a Markdown artifact into a coverage run subdirectory."""
    out_dir = _find_run_dir(
        market=market,
        ticker=ticker,
        output_dir=output_dir,
        run_dir=run_dir,
    )
    artifact_dir = out_dir / _slugify(subdir, "artifacts")
    _ensure_dir(artifact_dir)
    safe_name = _slugify(Path(filename).stem, "artifact") + ".md"
    path = artifact_dir / safe_name
    _write_text(path, markdown)
    return _relative_to_workspace(path)


@tool
def write_json_artifact(
    data_json: str,
    filename: str,
    subdir: str,
    ticker: str,
    market: str,
    run_dir: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Validate and write a JSON artifact into a coverage run subdirectory."""
    failure = _inline_json_failure(data_json, "data_json")
    if failure:
        return _json_result(failure)
    parsed: Any = _json_loads(data_json, "data_json")
    out_dir = _find_run_dir(
        market=market,
        ticker=ticker,
        output_dir=output_dir,
        run_dir=run_dir,
    )
    artifact_dir = out_dir / _slugify(subdir, "artifacts")
    _ensure_dir(artifact_dir)
    safe_name = _slugify(Path(filename).stem, "artifact") + ".json"
    path = artifact_dir / safe_name
    _write_text(
        path,
        json.dumps(parsed, ensure_ascii=False, indent=2) + "\n",
    )
    return _relative_to_workspace(path)


@tool
def resolve_task2_handoff(
    ticker: str,
    market: str,
    run_dir: str = "",
    task1_dir: str = "",
    company_research_path: str = "",
    business_driver_map_path: str = "",
    source_log_path: str = "",
    create_new_run: bool = False,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Resolve Task 2 to one canonical workspace-level run directory.

    Direct Task 2 runs should continue the run that already contains Task 1
    artifacts. This tool never reconstructs Task 1 from prose. It either finds
    real Task 1 files or returns a blocking failure with exact missing paths.
    """
    explicit_paths = [
        company_research_path,
        business_driver_map_path,
        source_log_path,
    ]
    supplied = [path for path in explicit_paths if path]

    selected_run_dir: Path | None = None
    source = ""
    if run_dir:
        selected_run_dir = _infer_run_dir_from_task1_path(_resolve_workspace_path(run_dir))
        source = "run_dir"
    elif task1_dir:
        selected_run_dir = _infer_run_dir_from_task1_path(
            _resolve_workspace_path(task1_dir)
        )
        source = "task1_dir"
    elif supplied:
        selected_run_dir = _infer_run_dir_from_task1_path(
            _resolve_workspace_path(supplied[0])
        )
        source = "task1_file_paths"
    elif create_new_run:
        selected_run_dir = _create_run_dir(
            company="",
            ticker=ticker,
            market=market,
            task_type="initiation",
            triggering_event="",
            output_dir=output_dir,
        )
        source = "created_new_run"
    else:
        selected_run_dir = _latest_task1_run(
            ticker=ticker,
            market=market,
            output_dir=output_dir,
        )
        source = "latest_task1_run"

    if selected_run_dir is None:
        coverage_runs = _coverage_dir(market, ticker, output_dir) / "runs"
        return _json_result(
            {
                "status": "FAIL",
                "ticker": ticker,
                "market": market,
                "run_dir": "",
                "missing_artifacts": [
                    f"No run with complete Task 1 artifacts under {_relative_to_workspace(coverage_runs)}"
                ],
                "created_new_run": False,
                "source": source,
            }
        )

    selected_run_dir = _canonicalize_workspace_path(selected_run_dir)
    missing = _task1_missing_paths(selected_run_dir)
    file_path_mismatches: list[str] = []
    for supplied_path, expected_name in zip(explicit_paths, TASK1_REQUIRED_ARTIFACTS):
        if not supplied_path:
            continue
        resolved = _resolve_workspace_path(supplied_path)
        expected = _task1_dir(selected_run_dir) / expected_name
        if resolved.resolve() != expected.resolve():
            file_path_mismatches.append(
                f"{expected_name}: expected {_relative_to_workspace(expected)}, "
                f"got {_relative_to_workspace(resolved)}"
            )

    key = f"{_safe_market(market)}:{_safe_ticker(ticker)}:{output_dir}"
    _ACTIVE_RUNS[key] = selected_run_dir

    wrong_root_artifacts = _wrong_root_artifacts_for(selected_run_dir)
    status = "OK" if not missing and not file_path_mismatches else "FAIL"
    return _json_result(
        {
            "status": status,
            "ticker": ticker,
            "market": market,
            "run_dir": _relative_to_workspace(selected_run_dir),
            "task1_dir": _relative_to_workspace(_task1_dir(selected_run_dir)),
            "model_dir": _relative_to_workspace(_statement_model_dir(selected_run_dir)),
            "missing_artifacts": missing,
            "file_path_mismatches": file_path_mismatches,
            "wrong_root_artifacts": wrong_root_artifacts,
            "created_new_run": source == "created_new_run",
            "source": source,
        }
    )


@tool
def verify_task2_artifacts(run_dir: str, stage: str = "all") -> str:
    """Verify Task 2 artifacts in the canonical run directory.

    Args:
        run_dir: Coverage run directory, absolute or workspace-relative.
        stage: task1, financial_facts, statements, reconciliation, workbook, or all.
    """
    out_dir = _resolve_workspace_path(run_dir)
    model_dir = _statement_model_dir(out_dir)
    normalized_stage = str(stage or "all").strip().lower()
    critical: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []
    artifacts: dict[str, Any] = {}

    valid_stages = {
        "task1",
        "financial_facts",
        "statements",
        "reconciliation",
        "workbook",
        "all",
    }
    if normalized_stage not in valid_stages:
        raise ValueError("stage must be one of: " + ", ".join(sorted(valid_stages)))

    def should_check(name: str) -> bool:
        return normalized_stage in {"all", name}

    wrong_root_artifacts = _wrong_root_artifacts_for(out_dir)
    for item in wrong_root_artifacts:
        target = critical if item["severity"] == "critical" else warnings
        target.append(
            {
                "category": "Wrong Artifact Root",
                "issue": (
                    "Artifact exists under single-stock-coverage/out instead of "
                    "workspace-level out."
                ),
                "path": item["wrong_root_path"],
            }
        )

    if should_check("task1"):
        artifacts["task1"] = {}
        for name, path in _required_task1_paths(out_dir).items():
            artifacts["task1"][name] = _artifact_record(path)
            if not artifact_exists(path):
                _append_finding(
                    critical,
                    category="Missing Task 1 Artifact",
                    issue=f"Required Task 1 artifact missing: {name}",
                    path=path,
                )

    if should_check("financial_facts"):
        artifacts["financial_facts"] = {}
        for name in ("financial_facts.json", "task2_context_packet.json"):
            path = model_dir / name
            artifacts["financial_facts"][name] = _artifact_record(path)
            payload = _read_json_file(path)
            if not isinstance(payload, dict):
                _append_finding(
                    critical,
                    category="Missing Financial Context",
                    issue=f"{name} must exist and be valid JSON before statement modeling.",
                    path=path,
                )
        facts = _read_json_file(model_dir / "financial_facts.json")
        if isinstance(facts, dict) and not isinstance(facts.get("historicals"), list):
            _append_finding(
                critical,
                category="Invalid Financial Facts",
                issue="financial_facts.json must include a historicals list.",
                path=model_dir / "financial_facts.json",
            )

    if should_check("statements"):
        artifacts["statements"] = {}
        for statement_type, filename in TASK2_STATEMENT_ARTIFACTS.items():
            path = model_dir / filename
            artifacts["statements"][statement_type] = _artifact_record(path)
            payload = _read_json_file(path)
            if not isinstance(payload, dict):
                _append_finding(
                    critical,
                    category="Missing Statement JSON",
                    issue=f"{filename} must be written by its statement subagent.",
                    path=path,
                )
                continue
            validation = _validate_statement_payload(payload, statement_type)
            for finding in validation["critical"]:
                critical.append({"statement_type": statement_type, **finding})
            for finding in validation["warnings"]:
                warnings.append({"statement_type": statement_type, **finding})

    if should_check("reconciliation"):
        path = model_dir / "statement_spec_pack.json"
        artifacts["reconciliation"] = {"statement_spec_pack.json": _artifact_record(path)}
        pack = _read_json_file(path)
        if not isinstance(pack, dict):
            _append_finding(
                critical,
                category="Missing Reconciliation",
                issue="statement_spec_pack.json must exist before workbook build.",
                path=path,
            )
        else:
            for finding in pack.get("critical") or []:
                critical.append({"statement_type": "statement_pack", **finding})
            for finding in pack.get("warnings") or []:
                warnings.append({"statement_type": "statement_pack", **finding})

    if should_check("workbook"):
        artifacts["workbook"] = {}
        for name in ("integrated_model.xlsx", "model_audit.md"):
            path = model_dir / name
            artifacts["workbook"][name] = _artifact_record(path)
            if not artifact_exists(path):
                _append_finding(
                    critical,
                    category="Missing Workbook Artifact",
                    issue=f"{name} is required for Task 3 handoff.",
                    path=path,
                )

    status = "PASS" if not critical else "FAIL"
    return _json_result(
        {
            "status": status,
            "stage": normalized_stage,
            "run_dir": _relative_to_workspace(out_dir),
            "critical_count": len(critical),
            "warning_count": len(warnings),
            "critical": critical,
            "warnings": warnings,
            "artifacts": artifacts,
            "wrong_root_artifacts": wrong_root_artifacts,
        }
    )


@tool
def write_task2_model_audit(run_dir: str, audit_json: str) -> str:
    """Write a compact Task 2 model_audit.md from structured findings only."""
    failure = _inline_json_failure(audit_json, "audit_json")
    if failure:
        return _json_result(failure)
    audit = _json_loads(audit_json, "audit_json")
    if not isinstance(audit, dict):
        raise ValueError("audit_json must be a JSON object")

    out_dir = _resolve_workspace_path(run_dir)
    model_dir = _statement_model_dir(out_dir)
    _ensure_dir(model_dir)
    path = model_dir / "model_audit.md"

    status = _safe_audit_text(audit.get("status") or "UNKNOWN", 80)
    route = _safe_audit_text(audit.get("route") or audit.get("model_route") or "")
    handoff_ready = bool(audit.get("task3_handoff_ready"))
    critical = _normalize_findings(audit.get("critical"))
    warnings = _normalize_findings(audit.get("warnings"))
    artifacts = audit.get("artifacts") if isinstance(audit.get("artifacts"), dict) else {}
    next_steps = [
        _safe_audit_text(item, 300)
        for item in (audit.get("next_steps") or [])
        if _safe_audit_text(item, 300)
    ][:12]

    lines = [
        "# Task 2 Model Audit",
        "",
        f"- Status: {status}",
        f"- Run: `{_relative_to_workspace(out_dir)}`",
        f"- Task 3 handoff ready: {'yes' if handoff_ready else 'no'}",
    ]
    if route:
        lines.append(f"- Route: {route}")
    lines.extend(
        [
            f"- Critical findings: {len(critical)}",
            f"- Warnings: {len(warnings)}",
            "",
            "## Critical Findings",
        ]
    )
    if critical:
        for idx, finding in enumerate(critical, start=1):
            path_text = f" (`{finding['path']}`)" if finding.get("path") else ""
            lines.append(
                f"{idx}. {finding['category']}: {finding['issue']}{path_text}"
            )
    else:
        lines.append("None.")

    lines.extend(["", "## Warnings"])
    if warnings:
        for idx, finding in enumerate(warnings[:50], start=1):
            path_text = f" (`{finding['path']}`)" if finding.get("path") else ""
            lines.append(
                f"{idx}. {finding['category']}: {finding['issue']}{path_text}"
            )
    else:
        lines.append("None.")

    lines.extend(["", "## Artifact Status"])
    if artifacts:
        for name, value in artifacts.items():
            lines.append(f"- `{_safe_audit_text(name, 120)}`: {_safe_audit_text(value, 240)}")
    else:
        lines.append("No artifact status supplied.")

    lines.extend(["", "## Next Steps"])
    if next_steps:
        for idx, item in enumerate(next_steps, start=1):
            lines.append(f"{idx}. {item}")
    else:
        lines.append("No follow-up steps supplied.")

    markdown = "\n".join(lines).strip() + "\n"
    _write_text(path, markdown)
    return _json_result(
        {
            "status": "OK",
            "model_audit_path": _relative_to_workspace(path),
            "critical_count": len(critical),
            "warning_count": len(warnings),
            "task3_handoff_ready": handoff_ready,
        }
    )


@tool
def read_statement_context(
    statement_type: str,
    run_dir: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Read minimal Task 2 context for one independent statement JSON modeler.

    Args:
        statement_type: income_statement, balance_sheet, or cash_flow.
        run_dir: Coverage run directory, absolute or workspace-relative.
        output_dir: Coverage root used only when resolving relative paths.

    Returns:
        JSON with Task 1 artifacts, financial_facts.json, optional
        task2_context_packet.json, required fields, canonical keys, and dependency
        hints for the requested statement type.
    """
    del output_dir
    normalized_type = _normalize_statement_type(statement_type)
    out_dir = _resolve_workspace_path(run_dir)
    return _json_result(_statement_context_payload(normalized_type, out_dir))


@tool
def validate_income_statement_json(
    run_dir: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Validate Revenue Build and Income Statement specs derived from Task 2 files."""
    del output_dir
    return _validate_statement_json(run_dir, "income_statement")


@tool
def write_income_statement_json(
    ticker: str,
    market: str,
    run_dir: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Build, validate, and write income_statement_spec.json and revenue_build_spec.json."""
    return _write_statement_json(
        statement_type="income_statement",
        ticker=ticker,
        market=market,
        run_dir=run_dir,
        output_dir=output_dir,
    )


@tool
def validate_balance_sheet_json(
    run_dir: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Validate Balance Sheet spec derived from Task 2 files."""
    del output_dir
    return _validate_statement_json(run_dir, "balance_sheet")


@tool
def write_balance_sheet_json(
    ticker: str,
    market: str,
    run_dir: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Build, validate, and write balance_sheet_spec.json."""
    return _write_statement_json(
        statement_type="balance_sheet",
        ticker=ticker,
        market=market,
        run_dir=run_dir,
        output_dir=output_dir,
    )


@tool
def validate_cash_flow_json(
    run_dir: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Validate Cash Flow Statement spec derived from Task 2 files."""
    del output_dir
    return _validate_statement_json(run_dir, "cash_flow")


@tool
def write_cash_flow_json(
    ticker: str,
    market: str,
    run_dir: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Build, validate, and write cash_flow_statement_spec.json."""
    return _write_statement_json(
        statement_type="cash_flow",
        ticker=ticker,
        market=market,
        run_dir=run_dir,
        output_dir=output_dir,
    )


@tool
def reconcile_statement_specs(
    ticker: str,
    market: str,
    run_dir: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Reconcile independent IS, BS, and CF JSON specs before workbook build.

    Reads the three Task 2 statement JSON artifacts, checks canonical row keys,
    source gaps, and cross-statement dependencies, then writes
    `02_financial_model/statement_spec_pack.json`.
    """
    out_dir = _find_run_dir(
        market=market,
        ticker=ticker,
        output_dir=output_dir,
        run_dir=run_dir,
    )
    model_dir = _statement_model_dir(out_dir)
    paths = {
        "income_statement": model_dir / "income_statement_spec.json",
        "balance_sheet": model_dir / "balance_sheet_spec.json",
        "cash_flow": model_dir / "cash_flow_statement_spec.json",
    }
    specs: dict[str, dict[str, Any]] = {}
    critical: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []

    for statement_type, path in paths.items():
        payload = _read_json_file(path)
        if not isinstance(payload, dict):
            critical.append(
                {
                    "statement_type": statement_type,
                    "category": "Missing Statement JSON",
                    "issue": f"Required statement JSON not found: {_relative_to_workspace(path)}",
                }
            )
            continue
        specs[statement_type] = payload
        validation = _validate_statement_payload(payload, statement_type)
        critical.extend(
            {
                "statement_type": statement_type,
                **finding,
            }
            for finding in validation["critical"]
        )
        warnings.extend(
            {
                "statement_type": statement_type,
                **finding,
            }
            for finding in validation["warnings"]
        )

    if len(specs) == 3:
        dependency_text = {
            name: " ".join(_dependency_values(payload)).lower()
            for name, payload in specs.items()
        }
        cross_checks = {
            "IS net income -> CF net income": (
                "income_statement.net_income" in dependency_text["cash_flow"]
            ),
            "CF ending cash -> BS cash": (
                "cash_flow.ending_cash" in dependency_text["balance_sheet"]
            ),
            "RE roll-forward dependency": (
                "income_statement.net_income" in dependency_text["balance_sheet"]
                and "dividends" in dependency_text["balance_sheet"]
            ),
            "DCF revenue dependency": (
                "revenue_total" in " ".join(_canonical_keys(specs["income_statement"]))
            ),
            "DCF debt/cash dependency": (
                "total_debt" in " ".join(_canonical_keys(specs["balance_sheet"]))
                and "cash_and_equivalents" in " ".join(
                    _canonical_keys(specs["balance_sheet"])
                )
            ),
            "DCF cash flow dependency": (
                "capex" in " ".join(_canonical_keys(specs["cash_flow"]))
                and "nwc_change" in " ".join(_canonical_keys(specs["cash_flow"]))
            ),
        }
        for check_name, passed in cross_checks.items():
            if not passed:
                critical.append(
                    {
                        "statement_type": "statement_pack",
                        "category": "Cross Statement Tie",
                        "issue": f"Failed reconciliation check: {check_name}",
                    }
                )

        for statement_type, payload in specs.items():
            if payload.get("unsourced_items"):
                warnings.append(
                    {
                        "statement_type": statement_type,
                        "category": "Source Gap",
                        "issue": (
                            f"{statement_type} has {len(payload['unsourced_items'])} "
                            "unsourced item(s) to disclose in model_audit.md."
                        ),
                    }
                )
            forecast_logic = json.dumps(
                payload.get("forecast_logic", {}),
                ensure_ascii=False,
            ).lower()
            if "hardcode" in forecast_logic and "no hardcode" not in forecast_logic:
                warnings.append(
                    {
                        "statement_type": statement_type,
                        "category": "Forecast Hardcode Risk",
                        "issue": "forecast_logic mentions hardcode; parent must review before builder.",
                    }
                )

    status = "PASS" if not critical else "FAIL"
    pack = {
        "status": status,
        "run_dir": _relative_to_workspace(out_dir),
        "statement_specs": specs,
        "critical_count": len(critical),
        "warning_count": len(warnings),
        "critical": critical,
        "warnings": warnings,
        "builder_blocked": bool(critical),
    }
    _ensure_dir(model_dir)
    pack_path = model_dir / "statement_spec_pack.json"
    _write_text(pack_path, _json_result(pack) + "\n")

    facts_path = model_dir / "financial_facts.json"
    context_path = model_dir / "task2_context_packet.json"
    financial_facts = _read_json_file(facts_path)
    if not isinstance(financial_facts, dict):
        financial_facts = _derive_financial_facts(
            specs=specs,
            ticker=ticker,
            market=market,
        )
    context_packet = _read_json_file(context_path)
    if not isinstance(context_packet, dict):
        context_packet = _derive_task2_context_packet(
            specs=specs,
            pack=pack,
            financial_facts=financial_facts,
        )
    else:
        context_packet = {
            **context_packet,
            "reconciliation_status": pack.get("status"),
            "critical_count": pack.get("critical_count", 0),
            "warning_count": pack.get("warning_count", 0),
        }
    _write_text(facts_path, _json_result(financial_facts) + "\n")
    _write_text(context_path, _json_result(context_packet) + "\n")

    return _json_result(
        {
            **pack,
            "statement_spec_pack_path": _relative_to_workspace(pack_path),
            "financial_facts_path": _relative_to_workspace(facts_path),
            "task2_context_packet_path": _relative_to_workspace(context_path),
        }
    )


def _scoped_build_integrated_three_statement_model_reference(
    run_dir: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Build deterministic Task 2 integrated_model.xlsx from reconciled inputs."""
    del output_dir
    try:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError as exc:
        return _json_result(
            {"status": "ERROR", "message": f"openpyxl is not installed: {exc}"}
        )

    out_dir = _resolve_workspace_path(run_dir)
    sources = _load_task2_model_sources(out_dir)
    merged = _merge_payload_model_input(sources["payload"])
    metadata = _model_metadata(merged)
    model_dir = _statement_model_dir(out_dir)
    pack = _read_json_file(model_dir / "statement_spec_pack.json")
    if isinstance(pack, dict) and pack.get("builder_blocked"):
        return _json_result(
            {
                "status": "FAIL",
                "run_dir": _relative_to_workspace(out_dir),
                "workbook_path": "",
                "critical_count": int(pack.get("critical_count") or 0),
                "warning_count": int(pack.get("warning_count") or 0),
                "critical": pack.get("critical") or [],
                "warnings": pack.get("warnings") or [],
                "message": "statement_spec_pack.json has critical findings; builder is blocked.",
            }
        )

    historicals = _historical_records(merged)
    latest = historicals[-1]
    forecast_labels = _forecast_labels(merged, latest["year"])
    assumptions = _assumption_set(merged, latest)

    _ensure_dir(model_dir)
    workbook_path = model_dir / "integrated_model.xlsx"

    periods: list[dict[str, Any]] = [
        {"label": record["period"], "record": record, "is_actual": True}
        for record in historicals
    ]
    periods.extend(
        {"label": label, "record": {}, "is_actual": False}
        for label in forecast_labels
    )
    first_col = 3
    last_col = first_col + len(periods) - 1
    actual_count = len(historicals)
    period_columns = {
        str(period["label"]): _column_letter(idx)
        for idx, period in enumerate(periods, start=first_col)
    }
    row_map = json.loads(json.dumps(DEFAULT_THREE_STATEMENT_ROW_MAP))

    styles = {
        "section": PatternFill("solid", fgColor="1F4E79"),
        "header": PatternFill("solid", fgColor="D9E1F2"),
        "input": PatternFill("solid", fgColor="EAF3F8"),
        "white": Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "bold": Font(name="Arial", bold=True, color="000000", size=11),
        "input_font": Font(name="Arial", color="0000FF", size=11),
        "formula": Font(name="Arial", color="000000", size=11),
        "link": Font(name="Arial", color="008000", size=11),
        "normal": Font(name="Arial", color="000000", size=11),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "border": Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        ),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = {name: wb.create_sheet(name) for name in THREE_STATEMENT_TABS}

    def set_title(ws, title: str) -> None:
        ws["A1"] = title
        ws["A1"].fill = styles["section"]
        ws["A1"].font = styles["white"]
        ws["A1"].alignment = styles["left"]
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(last_col, 6),
        )
        ws["A3"] = f"Currency: {metadata['currency']} | Unit: {metadata['unit']}"

    def set_period_headers(ws) -> None:
        ws["A5"] = "Line Item"
        ws["B5"] = "Source / Type"
        for cell_ref in ("A5", "B5"):
            ws[cell_ref].fill = styles["header"]
            ws[cell_ref].font = styles["bold"]
        for idx, period in enumerate(periods, start=first_col):
            type_cell = ws.cell(
                row=4,
                column=idx,
                value="Actual" if period["is_actual"] else "Forecast",
            )
            type_cell.font = styles["bold"]
            type_cell.alignment = styles["center"]
            header_cell = ws.cell(row=5, column=idx, value=period["label"])
            header_cell.fill = styles["header"]
            header_cell.font = styles["bold"]
            header_cell.alignment = styles["center"]

    def format_sheet(ws) -> None:
        ws.freeze_panes = "C6"
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        for col in range(first_col, last_col + 1):
            ws.column_dimensions[_column_letter(col)].width = 14
        for row in ws.iter_rows():
            for cell in row:
                cell.border = styles["border"]
                cell.alignment = styles["center"]
                if cell.value is not None and cell.number_format == "General":
                    cell.number_format = "#,##0.0;(#,##0.0);-"

    def section(ws, row: int, title: str) -> None:
        ws.cell(row=row, column=1, value=title)
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["section"]
            cell.font = styles["white"]

    def label(ws, row: int, text: str, source_type: str = "") -> None:
        ws.cell(row=row, column=1, value=text).font = styles["normal"]
        ws.cell(row=row, column=2, value=source_type).font = styles["normal"]

    def input_cell(cell, value: Any, source: str = "") -> None:
        cell.value = value
        cell.font = styles["input_font"]
        cell.fill = styles["input"]
        cell.alignment = styles["center"]
        if source:
            cell.comment = Comment(f"Source: {source}", "single-stock-coverage")

    def formula_cell(cell, formula: str, link: bool = False) -> None:
        cell.value = formula
        cell.font = styles["link"] if link else styles["formula"]
        cell.alignment = styles["center"]

    def pct_row(ws, row: int) -> None:
        for col in range(first_col, last_col + 1):
            ws.cell(row=row, column=col).number_format = "0.0%"

    for ws in sheets.values():
        set_title(ws, ws.title)
        set_period_headers(ws)

    cover = sheets["Cover"]
    for row, key, value in (
        (6, "Company", metadata["company"]),
        (7, "Ticker", metadata["ticker"]),
        (8, "Market", metadata["market"]),
        (9, "Fiscal Year End", metadata["fiscal_year_end"]),
        (10, "Model Date", datetime.now().strftime("%Y-%m-%d")),
        (12, "Artifact", "02_financial_model/integrated_model.xlsx"),
    ):
        cover.cell(row=row, column=1, value=key)
        cover.cell(row=row, column=2, value=value)

    sources = sheets["Sources"]
    for col, header in enumerate(("Source ID", "Period", "Source", "Notes"), start=1):
        sources.cell(row=6, column=col, value=header).fill = styles["header"]
        sources.cell(row=6, column=col).font = styles["bold"]
    unsourced_items = list(merged.get("unsourced") or [])
    for row, record in enumerate(historicals, start=7):
        source = _source_for(record)
        sources.cell(row=row, column=1, value=f"SRC-{row - 6:03d}")
        sources.cell(row=row, column=2, value=record["period"])
        sources.cell(row=row, column=3, value=source)
        sources.cell(row=row, column=4, value="Historical financial record")
        if source == "[UNSOURCED]":
            unsourced_items.append(f"{record['period']}: missing source")

    assumptions_ws = sheets["Assumptions"]
    assumptions_ws["A6"] = "Scenario"
    input_cell(assumptions_ws["B6"], "Base", "Task 2 assumption set")
    assumption_rows = {
        "Revenue Drivers": (9, {"Revenue Growth": (10, assumptions["revenue_growth"], "0.0%")}),
        "Margin Drivers": (
            13,
            {
                "Gross Margin": (14, assumptions["gross_margin"], "0.0%"),
                "Operating Expenses % Revenue": (
                    15,
                    assumptions["opex_pct_revenue"],
                    "0.0%",
                ),
                "Tax Rate": (16, assumptions["tax_rate"], "0.0%"),
            },
        ),
        "Working Capital Drivers": (
            19,
            {
                "AR Days": (20, assumptions["ar_days"], "0.0"),
                "Inventory % Revenue": (
                    21,
                    assumptions["inventory_pct_revenue"],
                    "0.0%",
                ),
                "AP Days": (22, assumptions["ap_days"], "0.0"),
            },
        ),
        "CapEx/D&A Drivers": (
            25,
            {
                "D&A % Revenue": (26, assumptions["da_pct_revenue"], "0.0%"),
                "CapEx % Revenue": (27, assumptions["capex_pct_revenue"], "0.0%"),
            },
        ),
        "Debt/Interest Drivers": (
            30,
            {
                "Interest Rate": (31, assumptions["interest_rate"], "0.0%"),
                "Debt Repayment % Beginning Debt": (
                    32,
                    assumptions["debt_repayment_pct"],
                    "0.0%",
                ),
            },
        ),
        "Share Count Drivers": (
            35,
            {
                "Diluted Shares Growth": (
                    36,
                    assumptions["diluted_shares_growth"],
                    "0.0%",
                ),
                "Dividend Payout % Net Income": (
                    37,
                    assumptions["dividend_payout_pct"],
                    "0.0%",
                ),
            },
        ),
    }
    for title, (header_row, rows) in assumption_rows.items():
        section(assumptions_ws, header_row, title)
        for name, (row, value, number_format) in rows.items():
            label(assumptions_ws, row, name, "Input")
            for col in range(first_col + actual_count, last_col + 1):
                cell = assumptions_ws.cell(row=row, column=col)
                input_cell(cell, value, "Task 2 child specs")
                cell.number_format = number_format

    revenue_build = sheets["Revenue Build"]
    section(revenue_build, 7, "Revenue Build")
    label(revenue_build, 10, "Core Revenue", "Input / Formula")
    label(revenue_build, 12, "Total Revenue", "Formula")
    label(revenue_build, 13, "YoY Growth", "Formula")
    pct_row(revenue_build, 13)
    for idx, period in enumerate(periods, start=first_col):
        col = _column_letter(idx)
        prev_col = _column_letter(idx - 1) if idx > first_col else ""
        if period["is_actual"]:
            input_cell(
                revenue_build.cell(row=10, column=idx),
                _historical_value(period["record"], "revenue"),
                _source_for(period["record"]),
            )
        else:
            formula_cell(
                revenue_build.cell(row=10, column=idx),
                f"={prev_col}12*(1+{_sheet_ref('Assumptions', f'{col}10')})",
            )
        formula_cell(revenue_build.cell(row=12, column=idx), f"={col}10")
        formula_cell(
            revenue_build.cell(row=13, column=idx),
            "=0" if idx == first_col else f"=IF({prev_col}12=0,0,{col}12/{prev_col}12-1)",
        )

    wc = sheets["Working Capital"]
    section(wc, 7, "Working Capital")
    for row, text in (
        (8, "Accounts Receivable"),
        (9, "Inventory"),
        (10, "Accounts Payable"),
        (11, "Net Working Capital"),
        (12, "Change in NWC"),
    ):
        label(wc, row, text, "Input / Formula")
    for idx, period in enumerate(periods, start=first_col):
        col = _column_letter(idx)
        if period["is_actual"]:
            record = period["record"]
            revenue = _historical_value(record, "revenue")
            values = (
                (8, _historical_value(record, "ar") or revenue * assumptions["ar_days"] / 365),
                (9, _historical_value(record, "inventory") or revenue * assumptions["inventory_pct_revenue"]),
                (10, _historical_value(record, "ap") or revenue * assumptions["ap_days"] / 365),
            )
            for row, value in values:
                input_cell(wc.cell(row=row, column=idx), value, _source_for(record))
        else:
            formula_cell(
                wc.cell(row=8, column=idx),
                f"={_sheet_ref('Income Statement', f'{col}8')}*{_sheet_ref('Assumptions', f'{col}20')}/365",
                True,
            )
            formula_cell(
                wc.cell(row=9, column=idx),
                f"={_sheet_ref('Income Statement', f'{col}8')}*{_sheet_ref('Assumptions', f'{col}21')}",
                True,
            )
            formula_cell(
                wc.cell(row=10, column=idx),
                f"={_sheet_ref('Income Statement', f'{col}8')}*{_sheet_ref('Assumptions', f'{col}22')}/365",
                True,
            )
        formula_cell(wc.cell(row=11, column=idx), f"={col}8+{col}9-{col}10")
        prev_col = _column_letter(idx - 1) if idx > first_col else ""
        formula_cell(
            wc.cell(row=12, column=idx),
            "=0" if idx == first_col else f"={col}11-{prev_col}11",
        )

    ppe = sheets["PP&E & D&A"]
    section(ppe, 7, "PP&E & D&A")
    for row, text in (
        (8, "Beginning PP&E"),
        (9, "CapEx"),
        (10, "D&A"),
        (11, "Ending PP&E"),
    ):
        label(ppe, row, text, "Input / Formula")
    for idx, period in enumerate(periods, start=first_col):
        col = _column_letter(idx)
        if idx == first_col:
            input_cell(
                ppe.cell(row=8, column=idx),
                _historical_value(period["record"], "ppe"),
                _source_for(period["record"]),
            )
        else:
            prev_col = _column_letter(idx - 1)
            formula_cell(ppe.cell(row=8, column=idx), f"={prev_col}11")
        if period["is_actual"]:
            record = period["record"]
            input_cell(
                ppe.cell(row=9, column=idx),
                abs(_historical_value(record, "capex")),
                _source_for(record),
            )
            input_cell(
                ppe.cell(row=10, column=idx),
                _historical_value(record, "da"),
                _source_for(record),
            )
        else:
            formula_cell(
                ppe.cell(row=9, column=idx),
                f"={_sheet_ref('Income Statement', f'{col}8')}*{_sheet_ref('Assumptions', f'{col}27')}",
                True,
            )
            formula_cell(
                ppe.cell(row=10, column=idx),
                f"={_sheet_ref('Income Statement', f'{col}8')}*{_sheet_ref('Assumptions', f'{col}26')}",
                True,
            )
        formula_cell(ppe.cell(row=11, column=idx), f"={col}8+{col}9-{col}10")

    debt = sheets["Debt & Interest"]
    section(debt, 7, "Debt & Interest")
    for row, text in (
        (8, "Beginning Debt"),
        (9, "Debt Issuance"),
        (10, "Debt Repayment"),
        (11, "Ending Debt"),
        (12, "Interest Expense"),
    ):
        label(debt, row, text, "Input / Formula")
    for idx, period in enumerate(periods, start=first_col):
        col = _column_letter(idx)
        if idx == first_col:
            input_cell(
                debt.cell(row=8, column=idx),
                _historical_value(period["record"], "debt"),
                _source_for(period["record"]),
            )
        else:
            prev_col = _column_letter(idx - 1)
            formula_cell(debt.cell(row=8, column=idx), f"={prev_col}11")
        if period["is_actual"]:
            input_cell(debt.cell(row=9, column=idx), 0, _source_for(period["record"]))
            input_cell(debt.cell(row=10, column=idx), 0, _source_for(period["record"]))
            input_cell(
                debt.cell(row=12, column=idx),
                _historical_value(period["record"], "interest_expense"),
                _source_for(period["record"]),
            )
        else:
            formula_cell(debt.cell(row=9, column=idx), "=0")
            formula_cell(
                debt.cell(row=10, column=idx),
                f"={col}8*{_sheet_ref('Assumptions', f'{col}32')}",
            )
            formula_cell(
                debt.cell(row=12, column=idx),
                f"=(({col}8+{col}11)/2)*{_sheet_ref('Assumptions', f'{col}31')}",
            )
        formula_cell(debt.cell(row=11, column=idx), f"={col}8+{col}9-{col}10")

    share = sheets["Share Count"]
    section(share, 7, "Share Count")
    for row, text in (
        (8, "Beginning Diluted Shares"),
        (9, "Share Issuance"),
        (10, "Buybacks"),
        (11, "Ending Diluted Shares"),
        (12, "Dividends"),
    ):
        label(share, row, text, "Input / Formula")
    for idx, period in enumerate(periods, start=first_col):
        col = _column_letter(idx)
        if idx == first_col:
            input_cell(
                share.cell(row=8, column=idx),
                _historical_value(period["record"], "shares"),
                _source_for(period["record"]),
            )
        else:
            prev_col = _column_letter(idx - 1)
            formula_cell(share.cell(row=8, column=idx), f"={prev_col}11")
        if period["is_actual"]:
            input_cell(share.cell(row=9, column=idx), 0, _source_for(period["record"]))
            input_cell(share.cell(row=10, column=idx), 0, _source_for(period["record"]))
            input_cell(
                share.cell(row=12, column=idx),
                _as_float(_field(period["record"], "dividends"), 0),
                _source_for(period["record"]),
            )
            formula_cell(share.cell(row=11, column=idx), f"={col}8+{col}9-{col}10")
        else:
            formula_cell(share.cell(row=9, column=idx), "=0")
            formula_cell(share.cell(row=10, column=idx), "=0")
            formula_cell(
                share.cell(row=11, column=idx),
                f"={col}8*(1+{_sheet_ref('Assumptions', f'{col}36')})",
            )
            formula_cell(
                share.cell(row=12, column=idx),
                f"=MAX(0,{_sheet_ref('Income Statement', f'{col}20')}*{_sheet_ref('Assumptions', f'{col}37')})",
                True,
            )

    income = sheets["Income Statement"]
    section(income, 7, "Income Statement")
    for row, text in (
        (8, "Revenue"),
        (9, "COGS"),
        (10, "Gross Profit"),
        (12, "Operating Expenses"),
        (13, "D&A"),
        (14, "EBIT"),
        (15, "EBITDA"),
        (17, "Interest Expense"),
        (18, "Pretax Income"),
        (19, "Tax Expense"),
        (20, "Net Income"),
        (22, "Diluted Shares"),
        (23, "EPS Diluted"),
    ):
        label(income, row, text, "Input / Formula")
    for idx, period in enumerate(periods, start=first_col):
        col = _column_letter(idx)
        if period["is_actual"]:
            record = period["record"]
            revenue = _historical_value(record, "revenue")
            gross_profit = (
                _historical_value(record, "gross_profit")
                or revenue * assumptions["gross_margin"]
            )
            da = _historical_value(record, "da")
            ebit = (
                _historical_value(record, "ebit")
                or gross_profit - _historical_value(record, "operating_expenses") - da
            )
            interest = _historical_value(record, "interest_expense")
            pretax = _historical_value(record, "pretax_income") or ebit - interest
            tax = _historical_value(record, "tax_expense")
            net_income = _historical_value(record, "net_income") or pretax - tax
            values = {
                8: revenue,
                9: max(revenue - gross_profit, 0),
                10: gross_profit,
                12: max(gross_profit - ebit - da, 0),
                13: da,
                14: ebit,
                15: _historical_value(record, "ebitda") or ebit + da,
                17: interest,
                18: pretax,
                19: tax,
                20: net_income,
            }
            for row, value in values.items():
                input_cell(income.cell(row=row, column=idx), value, _source_for(record))
        else:
            formula_cell(
                income.cell(row=8, column=idx),
                f"={_sheet_ref('Revenue Build', f'{col}12')}",
                True,
            )
            formula_cell(
                income.cell(row=9, column=idx),
                f"={col}8*(1-{_sheet_ref('Assumptions', f'{col}14')})",
            )
            formula_cell(income.cell(row=10, column=idx), f"={col}8-{col}9")
            formula_cell(
                income.cell(row=12, column=idx),
                f"={col}8*{_sheet_ref('Assumptions', f'{col}15')}",
            )
            formula_cell(
                income.cell(row=13, column=idx),
                f"={_sheet_ref('PP&E & D&A', f'{col}10')}",
                True,
            )
            formula_cell(income.cell(row=14, column=idx), f"={col}10-{col}12-{col}13")
            formula_cell(income.cell(row=15, column=idx), f"={col}14+{col}13")
            formula_cell(
                income.cell(row=17, column=idx),
                f"={_sheet_ref('Debt & Interest', f'{col}12')}",
                True,
            )
            formula_cell(income.cell(row=18, column=idx), f"={col}14-{col}17")
            formula_cell(
                income.cell(row=19, column=idx),
                f"=MAX(0,{col}18*{_sheet_ref('Assumptions', f'{col}16')})",
            )
            formula_cell(income.cell(row=20, column=idx), f"={col}18-{col}19")
        formula_cell(
            income.cell(row=22, column=idx),
            f"={_sheet_ref('Share Count', f'{col}11')}",
            True,
        )
        formula_cell(income.cell(row=23, column=idx), f"=IF({col}22=0,0,{col}20/{col}22)")

    bs = sheets["Balance Sheet"]
    section(bs, 7, "Assets")
    section(bs, 17, "Liabilities")
    section(bs, 23, "Equity")
    for row, text in (
        (8, "Cash and Equivalents"),
        (9, "Accounts Receivable"),
        (10, "Inventory"),
        (11, "Total Current Assets"),
        (13, "PP&E Net"),
        (14, "Other Assets"),
        (15, "Total Assets"),
        (18, "Accounts Payable"),
        (19, "Total Current Liabilities"),
        (20, "Total Debt"),
        (21, "Total Liabilities"),
        (24, "Common Stock + APIC"),
        (25, "Retained Earnings"),
        (26, "Total Equity"),
        (27, "Total Liabilities and Equity"),
    ):
        label(bs, row, text, "Input / Formula")
    for idx, period in enumerate(periods, start=first_col):
        col = _column_letter(idx)
        if period["is_actual"]:
            input_cell(
                bs.cell(row=8, column=idx),
                _historical_value(period["record"], "cash"),
                _source_for(period["record"]),
            )
        else:
            formula_cell(
                bs.cell(row=8, column=idx),
                f"={_sheet_ref('Cash Flow Statement', f'{col}25')}",
                True,
            )
        formula_cell(bs.cell(row=9, column=idx), f"={_sheet_ref('Working Capital', f'{col}8')}", True)
        formula_cell(bs.cell(row=10, column=idx), f"={_sheet_ref('Working Capital', f'{col}9')}", True)
        formula_cell(bs.cell(row=11, column=idx), f"=SUM({col}8:{col}10)")
        formula_cell(bs.cell(row=13, column=idx), f"={_sheet_ref('PP&E & D&A', f'{col}11')}", True)
        formula_cell(bs.cell(row=14, column=idx), "=0")
        formula_cell(bs.cell(row=15, column=idx), f"={col}11+{col}13+{col}14")
        formula_cell(bs.cell(row=18, column=idx), f"={_sheet_ref('Working Capital', f'{col}10')}", True)
        formula_cell(bs.cell(row=19, column=idx), f"={col}18")
        formula_cell(bs.cell(row=20, column=idx), f"={_sheet_ref('Debt & Interest', f'{col}11')}", True)
        formula_cell(bs.cell(row=21, column=idx), f"={col}19+{col}20")
        if idx == first_col:
            input_cell(
                bs.cell(row=25, column=idx),
                _historical_value(period["record"], "retained_earnings"),
                _source_for(period["record"]),
            )
        elif period["is_actual"]:
            input_cell(
                bs.cell(row=25, column=idx),
                _historical_value(period["record"], "retained_earnings"),
                _source_for(period["record"]),
            )
        else:
            prev_col = _column_letter(idx - 1)
            formula_cell(
                bs.cell(row=25, column=idx),
                f"={prev_col}25+{_sheet_ref('Income Statement', f'{col}20')}-{_sheet_ref('Share Count', f'{col}12')}",
                True,
            )
        formula_cell(bs.cell(row=24, column=idx), f"={col}15-{col}21-{col}25")
        formula_cell(bs.cell(row=26, column=idx), f"={col}24+{col}25")
        formula_cell(bs.cell(row=27, column=idx), f"={col}21+{col}26")

    cf = sheets["Cash Flow Statement"]
    section(cf, 7, "Operating Activities")
    section(cf, 13, "Investing Activities")
    section(cf, 17, "Financing Activities")
    for row, text in (
        (8, "Net Income"),
        (9, "D&A Addback"),
        (10, "Change in NWC"),
        (11, "CFO Total"),
        (14, "CapEx"),
        (15, "CFI Total"),
        (18, "Debt Issuance"),
        (19, "Debt Repayment"),
        (20, "Dividends"),
        (21, "CFF Total"),
        (23, "Net Change in Cash"),
        (24, "Beginning Cash"),
        (25, "Ending Cash"),
    ):
        label(cf, row, text, "Formula")
    for idx, _period in enumerate(periods, start=first_col):
        col = _column_letter(idx)
        formula_cell(cf.cell(row=8, column=idx), f"={_sheet_ref('Income Statement', f'{col}20')}", True)
        formula_cell(cf.cell(row=9, column=idx), f"={_sheet_ref('PP&E & D&A', f'{col}10')}", True)
        formula_cell(cf.cell(row=10, column=idx), f"=-{_sheet_ref('Working Capital', f'{col}12')}", True)
        formula_cell(cf.cell(row=11, column=idx), f"=SUM({col}8:{col}10)")
        formula_cell(cf.cell(row=14, column=idx), f"=-{_sheet_ref('PP&E & D&A', f'{col}9')}", True)
        formula_cell(cf.cell(row=15, column=idx), f"={col}14")
        formula_cell(cf.cell(row=18, column=idx), f"={_sheet_ref('Debt & Interest', f'{col}9')}", True)
        formula_cell(cf.cell(row=19, column=idx), f"=-{_sheet_ref('Debt & Interest', f'{col}10')}", True)
        formula_cell(cf.cell(row=20, column=idx), f"=-{_sheet_ref('Share Count', f'{col}12')}", True)
        formula_cell(cf.cell(row=21, column=idx), f"=SUM({col}18:{col}20)")
        formula_cell(cf.cell(row=23, column=idx), f"={col}11+{col}15+{col}21")
        if idx == first_col:
            formula_cell(
                cf.cell(row=24, column=idx),
                f"={_sheet_ref('Balance Sheet', f'{col}8')}-{col}23",
                True,
            )
        else:
            prev_col = _column_letter(idx - 1)
            formula_cell(cf.cell(row=24, column=idx), f"={prev_col}25")
        formula_cell(cf.cell(row=25, column=idx), f"={col}24+{col}23")

    dcf = sheets["DCF Inputs"]
    section(dcf, 7, "DCF Inputs")
    for row, text in (
        (8, "Revenue"),
        (9, "EBIT"),
        (10, "Tax Rate"),
        (11, "D&A"),
        (12, "CapEx"),
        (13, "Change in NWC"),
        (14, "Total Debt"),
        (15, "Cash"),
        (16, "Diluted Shares"),
        (17, "Scenario Label"),
    ):
        label(dcf, row, text, "Formula")
    for idx in range(first_col, last_col + 1):
        col = _column_letter(idx)
        refs = {
            8: _sheet_ref("Income Statement", f"{col}8"),
            9: _sheet_ref("Income Statement", f"{col}14"),
            10: _sheet_ref("Assumptions", f"{col}16"),
            11: _sheet_ref("Income Statement", f"{col}13"),
            12: _sheet_ref("Cash Flow Statement", f"{col}14"),
            13: _sheet_ref("Cash Flow Statement", f"{col}10"),
            14: _sheet_ref("Balance Sheet", f"{col}20"),
            15: _sheet_ref("Balance Sheet", f"{col}8"),
            16: _sheet_ref("Share Count", f"{col}11"),
            17: _sheet_ref("Assumptions", "$B$6"),
        }
        for row, ref in refs.items():
            formula_cell(dcf.cell(row=row, column=idx), f"={ref}", True)

    checks = sheets["Checks"]
    checks["A4"] = "Failing Check Count"
    formula_cell(checks["B4"], f"=COUNTIF(C8:{_column_letter(last_col)}15,\"<>0\")")
    for row, text in (
        (8, "BS Balance"),
        (9, "Cash Tie-Out"),
        (10, "NI Link"),
        (11, "RE Roll-Forward"),
        (12, "CapEx/PP&E Tie"),
        (13, "Debt Tie"),
        (14, "Revenue Tie"),
        (15, "D&A Tie"),
    ):
        label(checks, row, text, "Formula")
    for idx in range(first_col, last_col + 1):
        col = _column_letter(idx)
        prev_col = _column_letter(idx - 1) if idx > first_col else ""
        formulas = {
            8: f"={_sheet_ref('Balance Sheet', f'{col}15')}-{_sheet_ref('Balance Sheet', f'{col}27')}",
            9: f"={_sheet_ref('Cash Flow Statement', f'{col}25')}-{_sheet_ref('Balance Sheet', f'{col}8')}",
            10: f"={_sheet_ref('Income Statement', f'{col}20')}-{_sheet_ref('Cash Flow Statement', f'{col}8')}",
            11: "=0" if idx == first_col else f"={_sheet_ref('Balance Sheet', f'{prev_col}25')}+{_sheet_ref('Income Statement', f'{col}20')}-{_sheet_ref('Share Count', f'{col}12')}-{_sheet_ref('Balance Sheet', f'{col}25')}",
            12: f"={_sheet_ref('PP&E & D&A', f'{col}11')}-{_sheet_ref('PP&E & D&A', f'{col}8')}-{_sheet_ref('PP&E & D&A', f'{col}9')}+{_sheet_ref('PP&E & D&A', f'{col}10')}",
            13: f"={_sheet_ref('Debt & Interest', f'{col}11')}-{_sheet_ref('Debt & Interest', f'{col}8')}-{_sheet_ref('Debt & Interest', f'{col}9')}+{_sheet_ref('Debt & Interest', f'{col}10')}",
            14: f"={_sheet_ref('Income Statement', f'{col}8')}-{_sheet_ref('Revenue Build', f'{col}12')}",
            15: f"={_sheet_ref('Income Statement', f'{col}13')}-{_sheet_ref('PP&E & D&A', f'{col}10')}",
        }
        for row, formula in formulas.items():
            formula_cell(checks.cell(row=row, column=idx), formula)

    for name, ref in {
        "ScenarioSelector": "'Assumptions'!$B$6",
        "RevDriverBlock": f"'Assumptions'!$C$10:${_column_letter(last_col)}$10",
        "MarginDriverBlock": f"'Assumptions'!$C$14:${_column_letter(last_col)}$16",
        "NWCDriverBlock": f"'Assumptions'!$C$20:${_column_letter(last_col)}$22",
        "CapExDriverBlock": f"'Assumptions'!$C$26:${_column_letter(last_col)}$27",
        "DebtDriverBlock": f"'Assumptions'!$C$31:${_column_letter(last_col)}$33",
    }.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    for ws in sheets.values():
        format_sheet(ws)

    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    _save_workbook(wb, workbook_path)

    return _json_result(
        {
            "status": "OK",
            "workbook_path": _relative_to_workspace(workbook_path),
            "row_map": row_map,
            "period_columns": period_columns,
            "warnings": [],
            "unsourced_items": sorted(set(str(item) for item in unsourced_items)),
        }
    )


def _scoped_validate_integrated_three_statement_model_reference(
    excel_path: str,
    row_map_json: str = "",
) -> str:
    """Validate deterministic Task 2 integrated_model.xlsx structure and formulas."""
    try:
        import openpyxl
    except ImportError as exc:
        return _json_result(
            {"status": "ERROR", "message": f"openpyxl is not installed: {exc}"}
        )

    path = _resolve_workspace_path(excel_path)
    if not artifact_exists(path):
        return _validation_result(
            "FAIL",
            [
                {
                    "sheet": "",
                    "cell": "",
                    "category": "Missing File",
                    "issue": f"Workbook not found: {path}",
                }
            ],
            [],
        )

    row_map = json.loads(json.dumps(DEFAULT_THREE_STATEMENT_ROW_MAP))
    if row_map_json:
        failure = _inline_json_failure(row_map_json, "row_map_json")
        if failure:
            return _validation_result(
                "FAIL",
                [
                    {
                        "sheet": "",
                        "cell": "",
                        "category": "Inline JSON Too Large",
                        "issue": failure["message"],
                    }
                ],
                [],
            )
        parsed = _json_loads(row_map_json, "row_map_json")
        if (
            isinstance(parsed, dict)
            and isinstance(parsed.get("row_map"), dict)
        ):
            row_map = parsed["row_map"]
        elif isinstance(parsed, dict):
            row_map = parsed

    wb = _load_workbook(path, data_only=False)
    critical: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []

    for sheet_name in THREE_STATEMENT_TABS:
        if sheet_name not in wb.sheetnames:
            critical.append(
                {
                    "sheet": sheet_name,
                    "cell": "",
                    "category": "Missing Required Tab",
                    "issue": f"Required tab '{sheet_name}' is absent.",
                }
            )
    if critical:
        return _validation_result("FAIL", critical, warnings)

    defined_names = set(wb.defined_names)
    for name in REQUIRED_MODEL_NAMES:
        if name not in defined_names:
            critical.append(
                {
                    "sheet": "Assumptions",
                    "cell": name,
                    "category": "Missing Named Range",
                    "issue": f"Named range '{name}' is required for Task 2 handoff.",
                }
            )

    checks = wb["Checks"]
    forecast_cols = [
        col
        for col in range(3, checks.max_column + 1)
        if str(checks.cell(row=4, column=col).value or "").lower() == "forecast"
    ]
    if not forecast_cols:
        warnings.append(
            {
                "sheet": "Checks",
                "cell": "4:4",
                "category": "Period Metadata",
                "issue": "No forecast columns were marked in row 4; validating all modeled period columns.",
            }
        )
        forecast_cols = list(range(3, checks.max_column + 1))

    formula_requirements: list[tuple[str, list[int]]] = [
        (
            "Revenue Build",
            [
                row_map["revenue_build"]["core_revenue"],
                row_map["revenue_build"]["revenue_total"],
                row_map["revenue_build"]["revenue_growth"],
            ],
        ),
        ("Income Statement", list(row_map["income_statement"].values())),
        ("Balance Sheet", list(row_map["balance_sheet"].values())),
        ("Cash Flow Statement", list(row_map["cash_flow"].values())),
        ("Working Capital", list(row_map["working_capital"].values())),
        ("PP&E & D&A", list(row_map["ppe_da"].values())),
        ("Debt & Interest", list(row_map["debt_interest"].values())),
        ("Share Count", list(row_map["share_count"].values())),
        ("DCF Inputs", list(row_map["dcf_inputs"].values())),
        (
            "Checks",
            [
                row
                for key, row in row_map["checks"].items()
                if key != "failing_check_count"
            ],
        ),
    ]
    for sheet_name, rows in formula_requirements:
        ws = wb[sheet_name]
        for row in sorted(set(rows)):
            for col in forecast_cols:
                value = ws.cell(row=row, column=col).value
                if not _formula(value):
                    critical.append(
                        {
                            "sheet": sheet_name,
                            "cell": f"{_column_letter(col)}{row}",
                            "category": "Projection Hardcode",
                            "issue": "Forecast/model cell must be an Excel formula.",
                        }
                    )

    bs = wb["Balance Sheet"]
    cf = wb["Cash Flow Statement"]
    for col in forecast_cols:
        col_letter = _column_letter(col)
        bs_cash_row = row_map["balance_sheet"]["cash_and_equivalents"]
        cf_cash_row = row_map["cash_flow"]["ending_cash"]
        bs_cash = bs.cell(row=bs_cash_row, column=col).value
        cf_cash = cf.cell(row=cf_cash_row, column=col).value
        if not (
            _formula(bs_cash)
            and "Cash Flow Statement" in bs_cash
            and f"{col_letter}{cf_cash_row}" in bs_cash
        ):
            critical.append(
                {
                    "sheet": "Balance Sheet",
                    "cell": f"{col_letter}{bs_cash_row}",
                    "category": "Cash Tie-Out",
                    "issue": "Forecast BS cash must link to Cash Flow Statement ending cash.",
                }
            )
        if not _formula(cf_cash):
            critical.append(
                {
                    "sheet": "Cash Flow Statement",
                    "cell": f"{col_letter}{cf_cash_row}",
                    "category": "Cash Tie-Out",
                    "issue": "Forecast CF ending cash must be formula-driven.",
                }
            )

    dcf = wb["DCF Inputs"]
    for row in row_map["dcf_inputs"].values():
        for col in forecast_cols:
            value = dcf.cell(row=row, column=col).value
            if not _formula(value):
                critical.append(
                    {
                        "sheet": "DCF Inputs",
                        "cell": f"{_column_letter(col)}{row}",
                        "category": "DCF Input Hardcode",
                        "issue": "DCF Inputs must pull from linked model cells.",
                    }
                )

    return _validation_result("PASS" if not critical else "FAIL", critical, warnings)


@tool
def update_run_manifest(
    patch_json: str,
    ticker: str,
    market: str,
    run_dir: str = "",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Merge a JSON patch object into the current run manifest."""
    failure = _inline_json_failure(patch_json, "patch_json")
    if failure:
        return _json_result(failure)
    patch: Any = _json_loads(patch_json, "patch_json")
    if not isinstance(patch, dict):
        raise ValueError("patch_json must be a JSON object")

    out_dir = _find_run_dir(
        market=market,
        ticker=ticker,
        output_dir=output_dir,
        run_dir=run_dir,
    )
    manifest_path = out_dir / "run_manifest.json"
    manifest = json.loads(read_text_artifact(manifest_path))
    for key, value in patch.items():
        if isinstance(value, list) and isinstance(manifest.get(key), list):
            manifest[key].extend(value)
        else:
            manifest[key] = value
    manifest["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _write_text(
        manifest_path,
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
    )
    return _relative_to_workspace(manifest_path)


@tool
def build_integrated_three_statement_model(
    run_dir: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Build deterministic Task 2 integrated_model.xlsx from statement specs."""
    del output_dir
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError as exc:
        return _json_result({"status": "ERROR", "message": f"openpyxl is not installed: {exc}"})

    out_dir = _resolve_workspace_path(run_dir)
    sources = _load_task2_model_sources(out_dir)
    merged = _merge_payload_model_input(sources["payload"])
    model_dir = sources["model_dir"]
    _ensure_dir(model_dir)
    workbook_path = model_dir / "integrated_model.xlsx"

    file_facts = _read_json_file(model_dir / "financial_facts.json")
    file_context = _read_json_file(model_dir / "task2_context_packet.json")
    file_pack = _read_json_file(model_dir / "statement_spec_pack.json")
    file_revenue = _read_json_file(model_dir / "revenue_build_spec.json")
    if isinstance(file_facts, dict) and not isinstance(merged.get("financial_facts"), dict):
        merged = {**file_facts, **merged, "financial_facts": file_facts}
    if isinstance(file_context, dict) and not isinstance(merged.get("task2_context_packet"), dict):
        merged["task2_context_packet"] = file_context
    if isinstance(file_pack, dict) and not _statement_spec_pack(merged):
        merged["statement_spec_pack"] = file_pack
    if isinstance(file_revenue, dict) and file_revenue and not isinstance(merged.get("revenue_build_spec"), dict):
        merged["revenue_build_spec"] = file_revenue
    metadata = _model_metadata(merged)
    historicals = _historical_records(merged)
    interim_records = _interim_historical_records(merged)
    latest = historicals[-1]
    forecast_labels = _forecast_labels(merged, latest["year"])
    period_assumptions = _period_assumption_set(merged, latest, forecast_labels)
    assumptions = _assumption_set(merged, latest)
    revenue_components = _normalize_revenue_components(merged)

    periods: list[dict[str, Any]] = [
        {"label": record["period"], "record": record, "is_actual": True}
        for record in historicals
    ]
    periods.extend(
        {"label": label, "record": {}, "is_actual": False}
        for label in forecast_labels
    )
    first_col = 3
    last_col = first_col + len(periods) - 1
    actual_count = len(historicals)
    period_columns = {
        period["label"]: _column_letter(idx)
        for idx, period in enumerate(periods, start=first_col)
    }
    row_map = json.loads(json.dumps(DEFAULT_THREE_STATEMENT_ROW_MAP))
    unsourced_items = list(merged.get("unsourced") or [])

    revenue_component_rows: list[dict[str, Any]] = []
    revenue_row_map: dict[str, int] = {}
    rb_row = 8
    for component_idx, component in enumerate(revenue_components):
        driver_names = [
            str(name)
            for name in component.get("driver_names", [])
            if str(name) not in {"revenue", "cost", "gross_profit", "gross_margin", "mix"}
        ]
        rows = {
            "revenue": rb_row,
            "cost": rb_row + 1,
            "gross_profit": rb_row + 2,
            "gross_margin": rb_row + 3,
            "mix": rb_row + 4,
            "drivers": {},
        }
        rb_row += 5
        for driver_name in driver_names:
            rows["drivers"][driver_name] = rb_row
            rb_row += 1
        revenue_component_rows.append({"component": component, "rows": rows})
        key_prefix = f"component_{component_idx}"
        revenue_row_map[f"{key_prefix}_revenue"] = rows["revenue"]
        revenue_row_map[f"{key_prefix}_cost"] = rows["cost"]
        revenue_row_map[f"{key_prefix}_gross_profit"] = rows["gross_profit"]
        revenue_row_map[f"{key_prefix}_gross_margin"] = rows["gross_margin"]
        revenue_row_map[f"{key_prefix}_mix"] = rows["mix"]
        for driver_name, driver_row in rows["drivers"].items():
            safe_driver = re.sub(r"[^A-Za-z0-9_]+", "_", driver_name).strip("_").lower()
            revenue_row_map[f"{key_prefix}_driver_{safe_driver or 'metric'}"] = driver_row
    reconciliation_adjustment_row = rb_row
    revenue_total_row = rb_row + 1
    revenue_growth_row = rb_row + 2
    total_cost_row = rb_row + 3
    total_gross_profit_row = rb_row + 4
    total_gross_margin_row = rb_row + 5
    revenue_row_map.update(
        {
            "core_revenue": revenue_component_rows[0]["rows"]["revenue"],
            "reconciliation_adjustment": reconciliation_adjustment_row,
            "revenue_total": revenue_total_row,
            "revenue_growth": revenue_growth_row,
            "total_cost": total_cost_row,
            "total_gross_profit": total_gross_profit_row,
            "total_gross_margin": total_gross_margin_row,
        }
    )
    row_map["revenue_build"] = revenue_row_map
    row_map["revenue_components"] = [
        {
            "id": item["component"]["id"],
            "display_name": item["component"]["display_name"],
            "driver_type": item["component"]["driver_type"],
            "rows": item["rows"],
        }
        for item in revenue_component_rows
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = {name: wb.create_sheet(name) for name in THREE_STATEMENT_TABS}
    navy_fill = PatternFill("solid", fgColor="1F4E79")
    blue_fill = PatternFill("solid", fgColor="D9E1F2")
    input_font = Font(name="Arial", color="0000FF")
    formula_font = Font(name="Arial", color="000000")
    link_font = Font(name="Arial", color="008000")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", bold=True, color="000000")
    center = Alignment(horizontal="center", vertical="center")

    def init_sheet(ws, title: str) -> None:
        ws["A1"] = title
        ws["A1"].fill = navy_fill
        ws["A1"].font = header_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(last_col, 6))
        ws["A3"] = f"Currency: {metadata['currency']} | Unit: {metadata['unit']}"
        ws["A5"] = "Line Item"
        ws["B5"] = "Source / Type"
        for cell in (ws["A5"], ws["B5"]):
            cell.fill = blue_fill
            cell.font = bold_font
        for col_idx, period in enumerate(periods, start=first_col):
            col = _column_letter(col_idx)
            ws.cell(row=4, column=col_idx, value="Actual" if period["is_actual"] else "Forecast")
            ws.cell(row=5, column=col_idx, value=period["label"])
            ws[f"{col}4"].font = bold_font
            ws[f"{col}5"].font = bold_font
            ws[f"{col}5"].fill = blue_fill
        ws.freeze_panes = "C6"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        for col_idx in range(first_col, last_col + 1):
            ws.column_dimensions[_column_letter(col_idx)].width = 14

    def section(ws, row: int, title: str) -> None:
        ws.cell(row=row, column=1, value=title)
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = navy_fill
            cell.font = header_font

    def label(ws, row: int, text: str, source_type: str = "Formula") -> None:
        ws.cell(row=row, column=1, value=text)
        ws.cell(row=row, column=2, value=source_type)

    def input_cell(ws, row: int, col_idx: int, value: Any) -> None:
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = input_font
        cell.alignment = center

    def formula_cell(ws, row: int, col_idx: int, formula: str, link: bool = False) -> None:
        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = link_font if link else formula_font
        cell.alignment = center

    for ws in sheets.values():
        init_sheet(ws, ws.title)

    cover = sheets["Cover"]
    cover["A1"] = f"{metadata['company']} ({metadata['ticker']}) Integrated 3-Statement Model"
    for row, label_text, value in (
        (6, "Company", metadata["company"]),
        (7, "Ticker", metadata["ticker"]),
        (8, "Market", metadata["market"]),
        (9, "Fiscal Year End", metadata["fiscal_year_end"]),
        (10, "Workbook", "02_financial_model/integrated_model.xlsx"),
        (11, "Currency", metadata["currency"]),
        (12, "Unit", metadata["unit"]),
        (13, "Historical Periods", f"{periods[0]['label']} - {historicals[-1]['period']}"),
        (14, "Forecast Periods", f"{forecast_labels[0]} - {forecast_labels[-1]}"),
    ):
        label(cover, row, label_text, "Metadata")
        cover.cell(row=row, column=2, value=value)

    section(cover, 16, "Model Status")
    label(cover, 17, "Validation Status", "Formula")
    formula_cell(cover, 17, 2, '=IF(Checks!B4=0,"PASS","FAIL")', True)
    label(cover, 18, "Failing Check Count", "Formula")
    formula_cell(cover, 18, 2, "=Checks!B4", True)
    label(cover, 19, "Model Type", "Metadata")
    cover.cell(row=19, column=2, value="Integrated IS / BS / CF + DCF Inputs")

    section(cover, 21, "Model Snapshot")
    cover.cell(row=22, column=1, value="Line Item")
    cover.cell(row=22, column=2, value="Linked Source")
    for col_idx, period in enumerate(periods, start=first_col):
        cover.cell(row=22, column=col_idx, value=period["label"])
        cover.cell(row=22, column=col_idx).fill = blue_fill
        cover.cell(row=22, column=col_idx).font = bold_font
    for cell in (cover["A22"], cover["B22"]):
        cell.fill = blue_fill
        cell.font = bold_font
    snapshot_rows = {
        23: ("Revenue", "Income Statement!8", "Income Statement", 8, "#,##0.0"),
        24: ("Gross Margin", "IS Gross Profit / Revenue", "", 0, "0.0%"),
        25: ("EBIT", "Income Statement!14", "Income Statement", 14, "#,##0.0"),
        26: ("EBIT Margin", "IS EBIT / Revenue", "", 0, "0.0%"),
        27: ("Net Income", "Income Statement!22", "Income Statement", 22, "#,##0.0"),
        28: ("Cash", "Balance Sheet!8", "Balance Sheet", 8, "#,##0.0"),
        29: ("Total Assets", "Balance Sheet!15", "Balance Sheet", 15, "#,##0.0"),
        30: ("Total Debt", "Balance Sheet!20", "Balance Sheet", 20, "#,##0.0"),
        31: ("Diluted Shares", "Share Count!11", "Share Count", 11, "#,##0.0"),
    }
    for row, (line_item, source_text, sheet_name, source_row, number_format) in snapshot_rows.items():
        label(cover, row, line_item, source_text)
        for col_idx in range(first_col, last_col + 1):
            col = _column_letter(col_idx)
            if row == 24:
                formula = f"=IF({col}23=0,0,{_sheet_ref('Income Statement', f'{col}10')}/{col}23)"
            elif row == 26:
                formula = f"=IF({col}23=0,0,{col}25/{col}23)"
            else:
                formula = f"={_sheet_ref(sheet_name, f'{col}{source_row}')}"
            formula_cell(cover, row, col_idx, formula, True)
            cover.cell(row=row, column=col_idx).number_format = number_format

    section(cover, 34, "Workbook Navigation")
    cover["A35"] = "Sheet"
    cover["B35"] = "Purpose"
    cover["C35"] = "Primary Content"
    for cell in (cover["A35"], cover["B35"], cover["C35"]):
        cell.fill = blue_fill
        cell.font = bold_font
    navigation_rows = [
        ("Assumptions", "Forecast drivers", "Growth, margins, working capital, capex, debt, share count"),
        ("Revenue Build", "Revenue bridge", "Spec-driven revenue components, cost, margin, mix and YoY growth"),
        ("Income Statement", "Profitability model", "Revenue through EPS"),
        ("Balance Sheet", "Balance sheet model", "Assets, liabilities, equity, retained earnings"),
        ("Cash Flow Statement", "Cash flow model", "CFO, CFI, CFF, beginning and ending cash"),
        ("Working Capital", "NWC schedule", "AR, inventory, AP, change in NWC"),
        ("PP&E & D&A", "Fixed asset schedule", "CapEx, D&A, ending PP&E"),
        ("Debt & Interest", "Debt schedule", "Debt movement and interest expense"),
        ("Share Count", "Share schedule", "Diluted shares and dividends"),
        ("DCF Inputs", "Valuation handoff", "Linked model outputs for Task 3"),
        ("Checks", "Model audit", "BS balance, cash tie-out, roll-forward and linkage checks"),
        ("Sources", "Source audit", "Historical source status and coverage summary"),
    ]
    for row, (sheet_name, purpose, primary_content) in enumerate(navigation_rows, start=36):
        cover.cell(row=row, column=1, value=sheet_name)
        cover.cell(row=row, column=2, value=purpose)
        cover.cell(row=row, column=3, value=primary_content)

    section(cover, 50, "Formatting Legend")
    legend_rows = [
        (51, "Hardcoded inputs / editable assumptions", "Blue font"),
        (52, "Formulas / calculations", "Black font"),
        (53, "Same-workbook links", "Green font"),
    ]
    for row, label_text, meaning in legend_rows:
        label(cover, row, label_text, meaning)
    cover["A51"].font = input_font
    cover["A52"].font = formula_font
    cover["A53"].font = link_font

    sources = sheets["Sources"]
    unsourced_count = 0
    for record in historicals:
        if _source_for(record) == "[UNSOURCED]":
            unsourced_count += 1
    section(sources, 7, "Source Coverage Summary")
    source_summary_rows = [
        (8, "Historical periods loaded", len(historicals)),
        (9, "Forecast periods loaded", len(forecast_labels)),
        (10, "Unsourced historical periods", unsourced_count),
        (11, "Source status", "Needs source tags" if unsourced_count else "Source tags present"),
    ]
    for row, label_text, value in source_summary_rows:
        label(sources, row, label_text, "Summary")
        sources.cell(row=row, column=3, value=value)
    section(sources, 13, "Historical Source Detail")
    for idx, record in enumerate(historicals, start=14):
        label(sources, idx, record["period"], "Source")
        source = _source_for(record)
        sources.cell(row=idx, column=3, value=source)
        if source == "[UNSOURCED]":
            unsourced_items.append(f"{record['period']}: missing source")
    if interim_records:
        interim_start = 15 + len(historicals)
        section(sources, interim_start, "Interim Source Detail")
        for idx, record in enumerate(interim_records, start=interim_start + 1):
            period_label = str(record.get("period") or record.get("year") or f"Interim {idx - interim_start}")
            label(sources, idx, period_label, "Interim")
            sources.cell(row=idx, column=3, value=_source_for(record))
            sources.cell(row=idx, column=4, value="Excluded from annual model columns")

    assumptions_ws = sheets["Assumptions"]
    label(assumptions_ws, 6, "Scenario", "Input")
    assumptions_ws["B6"] = "Base"
    assumption_rows = {
        10: ("Revenue Growth", "revenue_growth", "0.0%"),
        14: ("Gross Margin", "gross_margin", "0.0%"),
        15: ("Operating Expenses % Revenue", "opex_pct_revenue", "0.0%"),
        16: ("Tax Rate", "tax_rate", "0.0%"),
        20: ("AR Days", "ar_days", "0.0"),
        21: ("Inventory Days", "inventory_days", "0.0"),
        22: ("AP Days", "ap_days", "0.0"),
        26: ("D&A % Revenue", "da_pct_revenue", "0.0%"),
        27: ("CapEx % Revenue", "capex_pct_revenue", "0.0%"),
        31: ("Interest Rate", "interest_rate", "0.0%"),
        32: ("Debt Repayment % Beginning Debt", "debt_repayment_pct", "0.0%"),
        33: ("Net Finance Expense / (Income)", "net_finance_expense", "#,##0.0"),
        36: ("Diluted Shares Growth", "diluted_shares_growth", "0.0%"),
        37: ("Dividend Payout % Net Income", "dividend_payout_pct", "0.0%"),
    }
    for header_row, title in (
        (9, "Revenue Drivers"),
        (13, "Margin Drivers"),
        (19, "Working Capital Drivers"),
        (25, "CapEx/D&A Drivers"),
        (30, "Debt/Interest Drivers"),
        (35, "Share Count Drivers"),
        (39, "Revenue Component Inputs"),
    ):
        section(assumptions_ws, header_row, title)
    for row, (display, key, number_format) in assumption_rows.items():
        label(assumptions_ws, row, display, "Input")
        for col_idx in range(first_col + actual_count, last_col + 1):
            label_text = periods[col_idx - first_col]["label"]
            value = period_assumptions.get(label_text, assumptions).get(key, assumptions.get(key, 0.0))
            input_cell(assumptions_ws, row, col_idx, value)
            assumptions_ws.cell(row=row, column=col_idx).number_format = number_format

    revenue_assumption_rows: list[dict[str, Any]] = []
    assumption_row = 40
    for item in revenue_component_rows:
        component = item["component"]
        display_name = str(component["display_name"])
        latest_margin = _latest_component_gross_margin(component, assumptions["gross_margin"])
        rows = {
            "revenue": assumption_row,
            "gross_margin": assumption_row + 1,
        }
        label(assumptions_ws, rows["revenue"], f"{display_name} Revenue", "Input")
        label(assumptions_ws, rows["gross_margin"], f"{display_name} Gross Margin", "Input")
        for col_idx in range(first_col + actual_count, last_col + 1):
            period_label = periods[col_idx - first_col]["label"]
            forecast_metrics = _revenue_component_period_metrics(
                component,
                period_label,
                is_actual=False,
            )
            input_cell(
                assumptions_ws,
                rows["revenue"],
                col_idx,
                _as_float(forecast_metrics.get("revenue"), 0.0),
            )
            assumptions_ws.cell(row=rows["revenue"], column=col_idx).number_format = "#,##0.0"
            margin_value = _as_decimal(
                forecast_metrics.get("gross_margin"),
                latest_margin,
            )
            input_cell(assumptions_ws, rows["gross_margin"], col_idx, margin_value)
            assumptions_ws.cell(row=rows["gross_margin"], column=col_idx).number_format = "0.0%"
        assumption_row += 2
        driver_rows: dict[str, int] = {}
        for driver_name in item["rows"]["drivers"]:
            driver_row = assumption_row
            driver_rows[driver_name] = driver_row
            label(assumptions_ws, driver_row, f"{display_name} {driver_name}", "Input")
            for col_idx in range(first_col + actual_count, last_col + 1):
                period_label = periods[col_idx - first_col]["label"]
                forecast_metrics = _revenue_component_period_metrics(
                    component,
                    period_label,
                    is_actual=False,
                )
                drivers = forecast_metrics.get("drivers") if isinstance(forecast_metrics, dict) else {}
                value = _as_float(drivers.get(driver_name), 0.0) if isinstance(drivers, dict) else 0.0
                input_cell(assumptions_ws, driver_row, col_idx, value)
            assumption_row += 1
        revenue_assumption_rows.append(
            {
                "component": component,
                "rows": rows,
                "driver_rows": driver_rows,
            }
        )

    revenue_build = sheets["Revenue Build"]
    section(revenue_build, 7, "Revenue Build")
    for item in revenue_component_rows:
        component = item["component"]
        rows = item["rows"]
        display_name = str(component["display_name"])
        label(revenue_build, rows["revenue"], f"{display_name} Revenue")
        label(revenue_build, rows["cost"], f"{display_name} COGS / Cost")
        label(revenue_build, rows["gross_profit"], f"{display_name} Gross Profit")
        label(revenue_build, rows["gross_margin"], f"{display_name} Gross Margin")
        label(revenue_build, rows["mix"], f"{display_name} Revenue Mix")
        for driver_name, driver_row in rows["drivers"].items():
            label(revenue_build, driver_row, f"{display_name} {driver_name}")
    label(revenue_build, reconciliation_adjustment_row, "Revenue Reconciliation Adjustment")
    label(revenue_build, revenue_total_row, "Total Revenue")
    label(revenue_build, revenue_growth_row, "YoY Growth")
    label(revenue_build, total_cost_row, "Total COGS / Cost")
    label(revenue_build, total_gross_profit_row, "Total Gross Profit")
    label(revenue_build, total_gross_margin_row, "Total Gross Margin")

    wc = sheets["Working Capital"]
    section(wc, 7, "Working Capital")
    for row, text in ((8, "Accounts Receivable"), (9, "Inventory"), (10, "Accounts Payable"), (11, "Net Working Capital"), (12, "Change in NWC")):
        label(wc, row, text)

    ppe = sheets["PP&E & D&A"]
    section(ppe, 7, "PP&E & D&A")
    for row, text in ((8, "Beginning PP&E"), (9, "CapEx"), (10, "D&A"), (11, "Ending PP&E")):
        label(ppe, row, text)

    debt = sheets["Debt & Interest"]
    section(debt, 7, "Debt & Interest")
    for row, text in (
        (8, "Beginning Debt"),
        (9, "Debt Issuance"),
        (10, "Debt Repayment"),
        (11, "Ending Debt"),
        (12, "Short-Term Debt Raw"),
        (13, "Interest Expense"),
        (14, "Interest Income"),
        (15, "Net Finance Expense / (Income)"),
        (16, "Debt Data Quality Flag"),
    ):
        label(debt, row, text)

    share = sheets["Share Count"]
    section(share, 7, "Share Count")
    for row, text in ((8, "Beginning Diluted Shares"), (9, "Share Issuance"), (10, "Buybacks"), (11, "Ending Diluted Shares"), (12, "Dividends")):
        label(share, row, text)

    income = sheets["Income Statement"]
    section(income, 7, "Income Statement")
    for row, text in (
        (8, "Revenue"),
        (9, "COGS"),
        (10, "Gross Profit"),
        (12, "Operating Expenses"),
        (13, "D&A"),
        (14, "EBIT"),
        (15, "EBITDA"),
        (17, "Interest Expense"),
        (18, "Interest Income"),
        (19, "Net Finance Expense / (Income)"),
        (20, "Pretax Income"),
        (21, "Tax Expense"),
        (22, "Net Income"),
        (24, "Diluted Shares"),
        (25, "EPS Diluted"),
    ):
        label(income, row, text)

    bs = sheets["Balance Sheet"]
    section(bs, 7, "Assets")
    section(bs, 17, "Liabilities")
    section(bs, 23, "Equity")
    for row, text in ((8, "Cash and Equivalents"), (9, "Accounts Receivable"), (10, "Inventory"), (11, "Total Current Assets"), (13, "PP&E Net"), (14, "Other Assets"), (15, "Total Assets"), (18, "Accounts Payable"), (19, "Total Current Liabilities"), (20, "Total Debt"), (21, "Total Liabilities"), (24, "Common Stock + APIC"), (25, "Retained Earnings"), (26, "Total Equity"), (27, "Total Liabilities and Equity")):
        label(bs, row, text)

    cf = sheets["Cash Flow Statement"]
    section(cf, 7, "Operating Activities")
    section(cf, 13, "Investing Activities")
    section(cf, 17, "Financing Activities")
    for row, text in ((8, "Net Income"), (9, "D&A Addback"), (10, "Change in NWC"), (11, "CFO Total"), (14, "CapEx"), (15, "CFI Total"), (18, "Debt Issuance"), (19, "Debt Repayment"), (20, "Dividends"), (21, "CFF Total"), (23, "Net Change in Cash"), (24, "Beginning Cash"), (25, "Ending Cash")):
        label(cf, row, text)

    dcf = sheets["DCF Inputs"]
    section(dcf, 7, "DCF Inputs")
    for row, text in ((8, "Revenue"), (9, "EBIT"), (10, "Tax Rate"), (11, "D&A"), (12, "CapEx"), (13, "Change in NWC"), (14, "Total Debt"), (15, "Cash"), (16, "Diluted Shares"), (17, "Scenario Label")):
        label(dcf, row, text)

    checks = sheets["Checks"]
    label(checks, 4, "Failing Check Count", "Formula")
    for row, text in ((8, "BS Balance"), (9, "Cash Tie-Out"), (10, "NI Link"), (11, "RE Roll-Forward"), (12, "CapEx/PP&E Tie"), (13, "Debt Tie"), (14, "Revenue Tie"), (15, "D&A Tie")):
        label(checks, row, text)

    def sum_rows_formula(col: str, rows: list[int]) -> str:
        if not rows:
            return "=0"
        return f"=SUM({','.join(f'{col}{row}' for row in rows)})"

    def component_metric_value(
        component: dict[str, Any],
        period_label: str,
        metric: str,
        *,
        is_actual: bool,
        default: float = 0.0,
    ) -> float:
        metrics = _revenue_component_period_metrics(component, period_label, is_actual=is_actual)
        return _as_float(metrics.get(metric), default) if isinstance(metrics, dict) else default

    def component_has_forecast_metric(
        component: dict[str, Any],
        period_label: str,
        metric: str,
    ) -> bool:
        metrics = _revenue_component_period_metrics(component, period_label, is_actual=False)
        return isinstance(metrics, dict) and metric in metrics and metrics[metric] not in (None, "")

    component_revenue_rows = [item["rows"]["revenue"] for item in revenue_component_rows]
    component_cost_rows = [item["rows"]["cost"] for item in revenue_component_rows]
    component_gross_profit_rows = [item["rows"]["gross_profit"] for item in revenue_component_rows]

    for col_idx, period in enumerate(periods, start=first_col):
        col = _column_letter(col_idx)
        prev_col = _column_letter(col_idx - 1) if col_idx > first_col else ""
        period_idx = col_idx - first_col
        is_actual = bool(period["is_actual"])
        record = period["record"]

        if is_actual:
            revenue = _historical_value(record, "revenue")
            gross_profit = _historical_value(record, "gross_profit") or revenue * assumptions["gross_margin"]
            ebit = _historical_value(record, "ebit") or gross_profit - _historical_value(record, "operating_expenses") - _historical_value(record, "da")
            net_finance_expense = _historical_value(record, "net_finance_expense")
            pretax = _historical_value(record, "pretax_income") or ebit - net_finance_expense
            net_income = _historical_value(record, "net_income") or pretax - _historical_value(record, "tax_expense")
            ending_debt = _historical_value(record, "debt")
            beginning_debt = ending_debt
            if period_idx > 0 and bool(periods[period_idx - 1]["is_actual"]):
                beginning_debt = _historical_value(periods[period_idx - 1]["record"], "debt")
            debt_issuance = max(ending_debt - beginning_debt, 0.0)
            debt_repayment = max(beginning_debt - ending_debt, 0.0)
            component_revenue_total = 0.0
            for item in revenue_component_rows:
                component = item["component"]
                rows = item["rows"]
                comp_revenue = component_metric_value(
                    component,
                    period["label"],
                    "revenue",
                    is_actual=True,
                    default=revenue if component["id"] == "core_revenue" else 0.0,
                )
                comp_gp = component_metric_value(component, period["label"], "gross_profit", is_actual=True)
                comp_cost = component_metric_value(component, period["label"], "cost", is_actual=True)
                comp_margin = _as_decimal(
                    _revenue_component_period_metrics(component, period["label"], is_actual=True).get("gross_margin")
                    if _revenue_component_period_metrics(component, period["label"], is_actual=True)
                    else None,
                    assumptions["gross_margin"],
                )
                if not comp_gp and comp_revenue and comp_cost:
                    comp_gp = comp_revenue - comp_cost
                if not comp_cost and comp_revenue and comp_gp:
                    comp_cost = comp_revenue - comp_gp
                if not comp_cost and comp_revenue:
                    comp_cost = comp_revenue * (1 - comp_margin)
                if not comp_gp and comp_revenue:
                    comp_gp = comp_revenue - comp_cost
                component_revenue_total += comp_revenue
                input_cell(revenue_build, rows["revenue"], col_idx, comp_revenue)
                input_cell(revenue_build, rows["cost"], col_idx, comp_cost)
                input_cell(revenue_build, rows["gross_profit"], col_idx, comp_gp)
                input_cell(
                    revenue_build,
                    rows["gross_margin"],
                    col_idx,
                    comp_gp / comp_revenue if comp_revenue else 0.0,
                )
                formula_cell(
                    revenue_build,
                    rows["mix"],
                    col_idx,
                    f"=IF({col}{revenue_total_row}=0,0,{col}{rows['revenue']}/{col}{revenue_total_row})",
                )
                for driver_name, driver_row in rows["drivers"].items():
                    metrics = _revenue_component_period_metrics(component, period["label"], is_actual=True)
                    drivers = metrics.get("drivers") if isinstance(metrics, dict) else {}
                    input_cell(
                        revenue_build,
                        driver_row,
                        col_idx,
                        _as_float(drivers.get(driver_name), 0.0) if isinstance(drivers, dict) else 0.0,
                    )
            formula_cell(
                revenue_build,
                reconciliation_adjustment_row,
                col_idx,
                f"={_sheet_ref('Income Statement', f'{col}8')}-{sum_rows_formula(col, component_revenue_rows)[1:]}",
                True,
            )
            for ws, row, value in (
                (income, 8, revenue),
                (income, 10, gross_profit),
                (income, 13, _historical_value(record, "da")),
                (income, 14, ebit),
                (income, 15, _historical_value(record, "ebitda") or ebit + _historical_value(record, "da")),
                (income, 17, _historical_value(record, "interest_expense")),
                (income, 18, _historical_value(record, "interest_income")),
                (income, 19, net_finance_expense),
                (income, 20, pretax),
                (income, 21, _historical_value(record, "tax_expense")),
                (income, 22, net_income),
                (wc, 8, _historical_value(record, "ar")),
                (wc, 9, _historical_value(record, "inventory")),
                (wc, 10, _historical_value(record, "ap")),
                (bs, 8, _historical_value(record, "cash")),
                (bs, 11, _historical_value(record, "total_current_assets")),
                (bs, 14, _historical_value(record, "other_assets")),
                (bs, 15, _historical_value(record, "total_assets")),
                (bs, 19, _historical_value(record, "current_liabilities")),
                (bs, 20, _historical_value(record, "debt")),
                (bs, 21, _historical_value(record, "total_liabilities")),
                (bs, 25, _historical_value(record, "retained_earnings")),
                (bs, 26, _historical_value(record, "total_equity")),
                (ppe, 8, _historical_value(record, "ppe")),
                (ppe, 9, abs(_historical_value(record, "capex"))),
                (ppe, 10, _historical_value(record, "da")),
                (ppe, 11, _historical_value(record, "ppe")),
                (debt, 8, beginning_debt),
                (debt, 9, debt_issuance),
                (debt, 10, debt_repayment),
                (debt, 11, ending_debt),
                (debt, 12, _historical_value(record, "short_term_debt_raw")),
                (debt, 13, _historical_value(record, "interest_expense")),
                (debt, 14, _historical_value(record, "interest_income")),
                (debt, 15, net_finance_expense),
                (share, 8, _historical_value(record, "shares")),
            ):
                input_cell(ws, row, col_idx, value)
        else:
            for item, assumption_item in zip(revenue_component_rows, revenue_assumption_rows):
                component = item["component"]
                rows = item["rows"]
                assumption_rows_for_component = assumption_item["rows"]
                if component_has_forecast_metric(component, period["label"], "revenue"):
                    revenue_assumption_ref = _sheet_ref(
                        "Assumptions",
                        f"{col}{assumption_rows_for_component['revenue']}",
                    )
                    formula_cell(
                        revenue_build,
                        rows["revenue"],
                        col_idx,
                        f"={revenue_assumption_ref}",
                        True,
                    )
                else:
                    formula_cell(
                        revenue_build,
                        rows["revenue"],
                        col_idx,
                        f"={prev_col}{rows['revenue']}*(1+{_sheet_ref('Assumptions', f'{col}10')})",
                    )
                if component_has_forecast_metric(component, period["label"], "cost"):
                    forecast_cost = component_metric_value(component, period["label"], "cost", is_actual=False)
                    input_cell(assumptions_ws, assumption_rows_for_component["gross_margin"], col_idx, 1 - forecast_cost / max(component_metric_value(component, period["label"], "revenue", is_actual=False), 1e-9))
                gross_margin_assumption_ref = _sheet_ref(
                    "Assumptions",
                    f"{col}{assumption_rows_for_component['gross_margin']}",
                )
                formula_cell(
                    revenue_build,
                    rows["cost"],
                    col_idx,
                    f"={col}{rows['revenue']}*(1-{gross_margin_assumption_ref})",
                    True,
                )
                formula_cell(revenue_build, rows["gross_profit"], col_idx, f"={col}{rows['revenue']}-{col}{rows['cost']}")
                formula_cell(revenue_build, rows["gross_margin"], col_idx, f"=IF({col}{rows['revenue']}=0,0,{col}{rows['gross_profit']}/{col}{rows['revenue']})")
                formula_cell(revenue_build, rows["mix"], col_idx, f"=IF({col}{revenue_total_row}=0,0,{col}{rows['revenue']}/{col}{revenue_total_row})")
                for driver_name, driver_row in rows["drivers"].items():
                    driver_assumption_row = assumption_item["driver_rows"].get(driver_name)
                    if driver_assumption_row:
                        formula_cell(
                            revenue_build,
                            driver_row,
                            col_idx,
                            f"={_sheet_ref('Assumptions', f'{col}{driver_assumption_row}')}",
                            True,
                        )
                    else:
                        formula_cell(revenue_build, driver_row, col_idx, "=0")
            formula_cell(revenue_build, reconciliation_adjustment_row, col_idx, "=0")
            formula_cell(income, 8, col_idx, f"={_sheet_ref('Revenue Build', f'{col}{revenue_total_row}')}", True)
            formula_cell(income, 9, col_idx, f"={_sheet_ref('Revenue Build', f'{col}{total_cost_row}')}", True)
            formula_cell(income, 10, col_idx, f"={_sheet_ref('Revenue Build', f'{col}{total_gross_profit_row}')}", True)
            formula_cell(income, 13, col_idx, f"={_sheet_ref('PP&E & D&A', f'{col}10')}", True)
            formula_cell(income, 14, col_idx, f"={col}10-{col}12-{col}13")
            formula_cell(income, 15, col_idx, f"={col}14+{col}13")
            formula_cell(income, 17, col_idx, f"={_sheet_ref('Debt & Interest', f'{col}13')}", True)
            formula_cell(income, 18, col_idx, f"={_sheet_ref('Debt & Interest', f'{col}14')}", True)
            formula_cell(income, 19, col_idx, f"={_sheet_ref('Debt & Interest', f'{col}15')}", True)
            formula_cell(income, 20, col_idx, f"={col}14-{col}19")
            formula_cell(income, 21, col_idx, f"=MAX(0,{col}20*{_sheet_ref('Assumptions', f'{col}16')})")
            formula_cell(income, 22, col_idx, f"={col}20-{col}21")
            formula_cell(bs, 8, col_idx, f"={_sheet_ref('Cash Flow Statement', f'{col}25')}", True)
            formula_cell(bs, 25, col_idx, f"={prev_col}25+{_sheet_ref('Income Statement', f'{col}22')}-{_sheet_ref('Share Count', f'{col}12')}", True)
            formula_cell(ppe, 8, col_idx, f"={prev_col}11")
            formula_cell(ppe, 9, col_idx, f"={_sheet_ref('Income Statement', f'{col}8')}*{_sheet_ref('Assumptions', f'{col}27')}", True)
            formula_cell(ppe, 10, col_idx, f"={_sheet_ref('Income Statement', f'{col}8')}*{_sheet_ref('Assumptions', f'{col}26')}", True)
            formula_cell(debt, 8, col_idx, f"={prev_col}11")
            formula_cell(debt, 13, col_idx, f"=(({col}8+{col}11)/2)*{_sheet_ref('Assumptions', f'{col}31')}")
            formula_cell(debt, 14, col_idx, f"=MAX(0,{col}13-{col}15)")
            formula_cell(debt, 15, col_idx, f"={_sheet_ref('Assumptions', f'{col}33')}", True)
            formula_cell(share, 8, col_idx, f"={prev_col}11")

        formula_cell(
            revenue_build,
            revenue_total_row,
            col_idx,
            sum_rows_formula(col, component_revenue_rows + [reconciliation_adjustment_row]),
        )
        formula_cell(
            revenue_build,
            revenue_growth_row,
            col_idx,
            "=0" if col_idx == first_col else f"=IF({prev_col}{revenue_total_row}=0,0,{col}{revenue_total_row}/{prev_col}{revenue_total_row}-1)",
        )
        formula_cell(revenue_build, total_cost_row, col_idx, sum_rows_formula(col, component_cost_rows))
        formula_cell(revenue_build, total_gross_profit_row, col_idx, sum_rows_formula(col, component_gross_profit_rows))
        formula_cell(
            revenue_build,
            total_gross_margin_row,
            col_idx,
            f"=IF({col}{revenue_total_row}=0,0,{col}{total_gross_profit_row}/{col}{revenue_total_row})",
        )
        if is_actual:
            formula_cell(income, 9, col_idx, f"={col}8-{col}10")
        formula_cell(income, 12, col_idx, f"={col}8*{_sheet_ref('Assumptions', f'{col}15')}" if not is_actual else f"=MAX(0,{col}10-{col}14-{col}13)")
        formula_cell(income, 24, col_idx, f"={_sheet_ref('Share Count', f'{col}11')}", True)
        formula_cell(income, 25, col_idx, f"=IF({col}24=0,0,{col}22/{col}24)")
        if not is_actual:
            formula_cell(wc, 8, col_idx, f"={_sheet_ref('Income Statement', f'{col}8')}*{_sheet_ref('Assumptions', f'{col}20')}/365", True)
            formula_cell(wc, 9, col_idx, f"={_sheet_ref('Income Statement', f'{col}9')}*{_sheet_ref('Assumptions', f'{col}21')}/365", True)
            formula_cell(wc, 10, col_idx, f"={_sheet_ref('Income Statement', f'{col}9')}*{_sheet_ref('Assumptions', f'{col}22')}/365", True)
        formula_cell(wc, 11, col_idx, f"={col}8+{col}9-{col}10")
        formula_cell(wc, 12, col_idx, "=0" if col_idx == first_col else f"={col}11-{prev_col}11")
        if not is_actual:
            formula_cell(ppe, 11, col_idx, f"={col}8+{col}9-{col}10")
        if not is_actual:
            formula_cell(debt, 9, col_idx, "=0")
            formula_cell(debt, 10, col_idx, f"={col}8*{_sheet_ref('Assumptions', f'{col}32')}")
            formula_cell(debt, 11, col_idx, f"={col}8+{col}9-{col}10")
            formula_cell(debt, 12, col_idx, f"={col}11")
        formula_cell(debt, 16, col_idx, f'=IF({col}12>MAX({col}11*5,0.01),"CHECK SHORT-TERM RAW > TOTAL DEBT","OK")')
        formula_cell(share, 9, col_idx, "=0")
        formula_cell(share, 10, col_idx, "=0")
        formula_cell(share, 11, col_idx, f"={col}8*(1+{_sheet_ref('Assumptions', f'{col}36')})" if not is_actual else f"={col}8+{col}9-{col}10")
        formula_cell(share, 12, col_idx, f"=MAX(0,{_sheet_ref('Income Statement', f'{col}22')}*{_sheet_ref('Assumptions', f'{col}37')})", True)
        formula_cell(bs, 9, col_idx, f"={_sheet_ref('Working Capital', f'{col}8')}", True)
        formula_cell(bs, 10, col_idx, f"={_sheet_ref('Working Capital', f'{col}9')}", True)
        if not is_actual:
            formula_cell(bs, 11, col_idx, f"=SUM({col}8:{col}10)")
        formula_cell(bs, 13, col_idx, f"={_sheet_ref('PP&E & D&A', f'{col}11')}", True)
        if not is_actual:
            formula_cell(bs, 14, col_idx, f"={prev_col}14" if prev_col else "=0")
            formula_cell(bs, 15, col_idx, f"={col}11+{col}13+{col}14")
        formula_cell(bs, 18, col_idx, f"={_sheet_ref('Working Capital', f'{col}10')}", True)
        if not is_actual:
            formula_cell(bs, 19, col_idx, f"={col}18")
            formula_cell(bs, 20, col_idx, f"={_sheet_ref('Debt & Interest', f'{col}11')}", True)
            formula_cell(bs, 21, col_idx, f"={col}19+{col}20")
        formula_cell(bs, 24, col_idx, f"={col}15-{col}21-{col}25")
        if not is_actual:
            formula_cell(bs, 26, col_idx, f"={col}24+{col}25")
        formula_cell(bs, 27, col_idx, f"={col}21+{col}26")
        formula_cell(cf, 8, col_idx, f"={_sheet_ref('Income Statement', f'{col}22')}", True)
        formula_cell(cf, 9, col_idx, f"={_sheet_ref('PP&E & D&A', f'{col}10')}", True)
        formula_cell(cf, 10, col_idx, f"=-{_sheet_ref('Working Capital', f'{col}12')}", True)
        formula_cell(cf, 11, col_idx, f"=SUM({col}8:{col}10)")
        formula_cell(cf, 14, col_idx, f"=-{_sheet_ref('PP&E & D&A', f'{col}9')}", True)
        formula_cell(cf, 15, col_idx, f"={col}14")
        formula_cell(cf, 18, col_idx, f"={_sheet_ref('Debt & Interest', f'{col}9')}", True)
        formula_cell(cf, 19, col_idx, f"=-{_sheet_ref('Debt & Interest', f'{col}10')}", True)
        formula_cell(cf, 20, col_idx, f"=-{_sheet_ref('Share Count', f'{col}12')}", True)
        formula_cell(cf, 21, col_idx, f"=SUM({col}18:{col}20)")
        formula_cell(cf, 23, col_idx, f"={col}11+{col}15+{col}21")
        formula_cell(cf, 24, col_idx, f"={_sheet_ref('Balance Sheet', f'{col}8')}-{col}23" if col_idx == first_col else f"={prev_col}25", True)
        formula_cell(
            cf,
            25,
            col_idx,
            f"={_sheet_ref('Balance Sheet', f'{col}8')}" if is_actual else f"={col}24+{col}23",
            True if is_actual else False,
        )
        for row, ref in {
            8: _sheet_ref("Income Statement", f"{col}8"),
            9: _sheet_ref("Income Statement", f"{col}14"),
            11: _sheet_ref("Income Statement", f"{col}13"),
            12: _sheet_ref("Cash Flow Statement", f"{col}14"),
            13: _sheet_ref("Cash Flow Statement", f"{col}10"),
            14: _sheet_ref("Balance Sheet", f"{col}20"),
            15: _sheet_ref("Balance Sheet", f"{col}8"),
            16: _sheet_ref("Share Count", f"{col}11"),
            17: _sheet_ref("Assumptions", "$B$6"),
        }.items():
            formula_cell(dcf, row, col_idx, f"={ref}", True)
        if is_actual:
            formula_cell(
                dcf,
                10,
                col_idx,
                f"=IF({_sheet_ref('Income Statement', f'{col}20')}=0,0,{_sheet_ref('Income Statement', f'{col}21')}/{_sheet_ref('Income Statement', f'{col}20')})",
                True,
            )
        else:
            formula_cell(dcf, 10, col_idx, f"={_sheet_ref('Assumptions', f'{col}16')}", True)

    checks_headers = ("Check", "Period", "Actual", "Expected", "Difference", "Tolerance", "Status", "Notes")
    for idx, header in enumerate(checks_headers, start=1):
        cell = checks.cell(row=7, column=idx, value=header)
        cell.fill = blue_fill
        cell.font = bold_font

    check_row = 8

    def add_check(
        name: str,
        period_label: str,
        actual_formula: str,
        expected_formula: str,
        *,
        tolerance: float = 0.05,
        notes: str = "",
    ) -> None:
        nonlocal check_row
        checks.cell(row=check_row, column=1, value=name)
        checks.cell(row=check_row, column=2, value=period_label)
        formula_cell(checks, check_row, 3, actual_formula, True)
        formula_cell(checks, check_row, 4, expected_formula, True)
        formula_cell(checks, check_row, 5, f"=MAX(C{check_row}-D{check_row},D{check_row}-C{check_row})")
        input_cell(checks, check_row, 6, tolerance)
        formula_cell(checks, check_row, 7, f'=IF(E{check_row}<=F{check_row},"PASS","FAIL")')
        checks.cell(row=check_row, column=8, value=notes)
        check_row += 1

    for col_idx, period in enumerate(periods, start=first_col):
        col = _column_letter(col_idx)
        prev_col = _column_letter(col_idx - 1) if col_idx > first_col else ""
        label_text = str(period["label"])
        is_actual_period = bool(period["is_actual"])
        add_check(
            "BS Balance",
            label_text,
            f"={_sheet_ref('Balance Sheet', f'{col}15')}",
            f"={_sheet_ref('Balance Sheet', f'{col}27')}",
        )
        add_check(
            "Cash Tie-Out",
            label_text,
            f"={_sheet_ref('Cash Flow Statement', f'{col}25')}",
            f"={_sheet_ref('Balance Sheet', f'{col}8')}",
        )
        add_check(
            "NI Link",
            label_text,
            f"={_sheet_ref('Cash Flow Statement', f'{col}8')}",
            f"={_sheet_ref('Income Statement', f'{col}22')}",
        )
        re_expected = (
            f"={_sheet_ref('Balance Sheet', f'{col}25')}"
            if col_idx == first_col or is_actual_period
            else f"={_sheet_ref('Balance Sheet', f'{prev_col}25')}+{_sheet_ref('Income Statement', f'{col}22')}-{_sheet_ref('Share Count', f'{col}12')}"
        )
        add_check(
            "RE Roll-Forward",
            label_text,
            f"={_sheet_ref('Balance Sheet', f'{col}25')}",
            re_expected,
            notes="Historical retained earnings sourced, not rebuilt" if is_actual_period else "",
        )
        ppe_expected = (
            f"={_sheet_ref('PP&E & D&A', f'{col}11')}"
            if is_actual_period
            else f"={_sheet_ref('PP&E & D&A', f'{col}8')}+{_sheet_ref('PP&E & D&A', f'{col}9')}-{_sheet_ref('PP&E & D&A', f'{col}10')}"
        )
        add_check(
            "CapEx/PP&E Tie",
            label_text,
            f"={_sheet_ref('PP&E & D&A', f'{col}11')}",
            ppe_expected,
            notes="Historical PP&E sourced, not rebuilt" if is_actual_period else "",
        )
        add_check(
            "Debt Tie",
            label_text,
            f"={_sheet_ref('Debt & Interest', f'{col}11')}",
            f"={_sheet_ref('Debt & Interest', f'{col}8')}+{_sheet_ref('Debt & Interest', f'{col}9')}-{_sheet_ref('Debt & Interest', f'{col}10')}",
        )
        add_check(
            "Revenue Tie",
            label_text,
            f"={_sheet_ref('Income Statement', f'{col}8')}",
            f"={_sheet_ref('Revenue Build', f'{col}{revenue_total_row}')}",
        )
        add_check(
            "D&A Tie",
            label_text,
            f"={_sheet_ref('Income Statement', f'{col}13')}",
            f"={_sheet_ref('PP&E & D&A', f'{col}10')}",
        )

    formula_cell(checks, 4, 2, f'=COUNTIF(G8:G{check_row - 1},"FAIL")')
    for name, ref in {
        "ScenarioSelector": "'Assumptions'!$B$6",
        "RevDriverBlock": f"'Assumptions'!$C$10:${_column_letter(last_col)}$10",
        "MarginDriverBlock": f"'Assumptions'!$C$14:${_column_letter(last_col)}$16",
        "NWCDriverBlock": f"'Assumptions'!$C$20:${_column_letter(last_col)}$22",
        "CapExDriverBlock": f"'Assumptions'!$C$26:${_column_letter(last_col)}$27",
        "DebtDriverBlock": f"'Assumptions'!$C$31:${_column_letter(last_col)}$33",
    }.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    for ws in sheets.values():
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = center
                    if cell.number_format == "General":
                        cell.number_format = "#,##0.0;(#,##0.0);-"

    formula_cache = _evaluate_workbook_formula_caches(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    cached_formula_count = _save_workbook_with_formula_cache(
        wb,
        workbook_path,
        formula_cache,
    )
    return _json_result(
        {
            "status": "OK",
            "workbook_path": _relative_to_workspace(workbook_path),
            "row_map": row_map,
            "period_columns": period_columns,
            "cached_formula_count": cached_formula_count,
            "warnings": [],
            "unsourced_items": sorted(set(str(item) for item in unsourced_items)),
        }
    )


@tool
def update_integrated_three_statement_model(
    prior_workbook_path: str,
    run_dir: str,
    update_scope_json: str = "{}",
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Refresh an existing Task 2 workbook into the current run directory.

    The update executor uses this only after statement specs are reconciled. It
    does not fetch data; it copies the prior workbook, preserves formulas, marks
    the workbook for recalculation, and returns the current-run workbook path for
    validation/audit.
    """
    del output_dir
    source_path = _resolve_workspace_path(prior_workbook_path)
    if not artifact_exists(source_path):
        return _json_result(
            {
                "status": "FAIL",
                "workbook_path": "",
                "critical_count": 1,
                "critical": [
                    {
                        "category": "Missing Prior Workbook",
                        "issue": f"Prior workbook not found: {source_path}",
                    }
                ],
                "warnings": [],
            }
        )

    try:
        import openpyxl
    except ImportError as exc:
        return _json_result({"status": "ERROR", "message": f"openpyxl is not installed: {exc}"})

    out_dir = _resolve_workspace_path(run_dir)
    model_dir = _statement_model_dir(out_dir)
    _ensure_dir(model_dir)
    workbook_path = model_dir / "integrated_model.xlsx"
    if source_path.resolve() != workbook_path.resolve():
        _copy_file_artifact(source_path, workbook_path)

    failure = _inline_json_failure(update_scope_json or "{}", "update_scope_json")
    if failure:
        return _json_result(failure)
    update_scope = _json_loads(update_scope_json or "{}", "update_scope_json")
    if not isinstance(update_scope, dict):
        raise ValueError("update_scope_json must be a JSON object")
    sources = _load_task2_model_sources(out_dir)
    model_input = sources["payload"]
    statement_pack = sources["statement_spec_pack"]
    if not isinstance(statement_pack, dict):
        raise ValueError("statement_spec_pack.json must be a JSON object")
    warnings = list(statement_pack.get("warnings") or [])
    if statement_pack.get("builder_blocked"):
        return _json_result(
            {
                "status": "FAIL",
                "workbook_path": "",
                "critical_count": int(statement_pack.get("critical_count") or 0),
                "critical": statement_pack.get("critical") or [],
                "warnings": statement_pack.get("warnings") or [],
                "message": "statement_spec_pack.json has critical findings; update is blocked.",
            }
        )

    wb = _load_workbook(workbook_path)
    updated_cells: list[str] = []
    if model_input:
        try:
            merged = _merge_payload_model_input(model_input)
            historicals = _historical_records(merged)
        except ValueError as exc:
            warnings.append(
                {
                    "category": "Model Input",
                    "issue": f"Could not apply financial_facts.json to workbook: {exc}",
                }
            )
        else:
            period_ws = wb["Income Statement"] if "Income Statement" in wb.sheetnames else wb.active
            period_cols = {
                str(period_ws.cell(row=5, column=col_idx).value): col_idx
                for col_idx in range(3, period_ws.max_column + 1)
                if period_ws.cell(row=5, column=col_idx).value
            }
            update_map = {
                "Revenue Build": {
                    "core_revenue": "revenue",
                    "revenue_total": "revenue",
                },
                "Income Statement": {
                    "revenue_total": "revenue",
                    "gross_profit": "gross_profit",
                    "da_total": "da",
                    "ebit": "ebit",
                    "ebitda": "ebitda",
                    "interest_expense": "interest_expense",
                    "pretax_income": "pretax_income",
                    "tax_expense": "tax_expense",
                    "net_income": "net_income",
                    "diluted_shares": "shares",
                },
                "Balance Sheet": {
                    "cash_and_equivalents": "cash",
                    "accounts_receivable": "ar",
                    "inventory": "inventory",
                    "accounts_payable": "ap",
                    "net_ppe": "ppe",
                    "total_debt": "debt",
                    "retained_earnings": "retained_earnings",
                },
                "Cash Flow Statement": {
                    "net_income_cf": "net_income",
                    "da_addback": "da",
                    "capex": "capex",
                    "ending_cash": "cash",
                },
                "Working Capital": {
                    "accounts_receivable": "ar",
                    "inventory": "inventory",
                    "accounts_payable": "ap",
                },
                "PP&E & D&A": {
                    "capex": "capex",
                    "da_total": "da",
                    "ending_ppe": "ppe",
                },
                "Debt & Interest": {
                    "ending_debt": "debt",
                },
                "Share Count": {
                    "ending_diluted_shares": "shares",
                },
            }
            sheet_key_map = {
                "Revenue Build": "revenue_build",
                "Income Statement": "income_statement",
                "Balance Sheet": "balance_sheet",
                "Cash Flow Statement": "cash_flow",
                "Working Capital": "working_capital",
                "PP&E & D&A": "ppe_da",
                "Debt & Interest": "debt_interest",
                "Share Count": "share_count",
            }
            for record in historicals:
                col_idx = period_cols.get(str(record["period"]))
                if not col_idx:
                    warnings.append(
                        {
                            "category": "Period Mapping",
                            "issue": f"Workbook has no existing period column for {record['period']}.",
                        }
                    )
                    continue
                for sheet_name, row_to_field in update_map.items():
                    if sheet_name not in wb.sheetnames:
                        continue
                    ws = wb[sheet_name]
                    sheet_rows = DEFAULT_THREE_STATEMENT_ROW_MAP[
                        sheet_key_map[sheet_name]
                    ]
                    for row_key, field_name in row_to_field.items():
                        if field_name not in record or row_key not in sheet_rows:
                            continue
                        ws.cell(row=sheet_rows[row_key], column=col_idx, value=record[field_name])
                        updated_cells.append(
                            f"{sheet_name}!{_column_letter(col_idx)}{sheet_rows[row_key]}"
                        )

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    if "Cover" in wb.sheetnames:
        cover = wb["Cover"]
        cover["A14"] = "Update Scope"
        cover["B14"] = json.dumps(update_scope, ensure_ascii=False)
        if model_input:
            cover["A15"] = "Updated Model Input"
            cover["B15"] = "financial_facts/task2_context_packet"
    formula_cache = _evaluate_workbook_formula_caches(wb)
    cached_formula_count = _save_workbook_with_formula_cache(
        wb,
        workbook_path,
        formula_cache,
    )

    return _json_result(
        {
            "status": "OK",
            "workbook_path": _relative_to_workspace(workbook_path),
            "prior_workbook_path": _relative_to_workspace(source_path),
            "update_scope": update_scope,
            "updated_cells": updated_cells,
            "critical_count": 0,
            "warning_count": len(warnings),
            "critical": [],
            "warnings": warnings,
            "cached_formula_count": cached_formula_count,
        }
    )


@tool
def validate_integrated_three_statement_model(
    excel_path: str,
    row_map_json: str = "",
) -> str:
    """Validate Task 2 integrated_model.xlsx for formula and tie-out discipline."""
    try:
        import openpyxl
    except ImportError as exc:
        return _json_result({"status": "ERROR", "message": f"openpyxl is not installed: {exc}"})

    path = _resolve_workspace_path(excel_path)
    if not artifact_exists(path):
        return _validation_result(
            "FAIL",
            [{"sheet": "", "cell": "", "category": "Missing File", "issue": f"Workbook not found: {path}"}],
            [],
        )

    row_map = DEFAULT_THREE_STATEMENT_ROW_MAP
    if row_map_json:
        failure = _inline_json_failure(row_map_json, "row_map_json")
        if failure:
            return _validation_result(
                "FAIL",
                [
                    {
                        "sheet": "",
                        "cell": "",
                        "category": "Inline JSON Too Large",
                        "issue": failure["message"],
                    }
                ],
                [],
            )
        parsed = _json_loads(row_map_json, "row_map_json")
        if isinstance(parsed, dict) and isinstance(parsed.get("row_map"), dict):
            row_map = parsed["row_map"]
        elif isinstance(parsed, dict):
            row_map = parsed

    workbook_bytes = read_bytes_artifact(path) if backend_is_daytona() else None
    if workbook_bytes is not None:
        wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=False)
        wb_values = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    else:
        wb = openpyxl.load_workbook(path, data_only=False)
        wb_values = openpyxl.load_workbook(path, data_only=True)
    critical: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []
    source_payload: dict[str, Any] = {}
    facts = _read_json_file(path.parent / "financial_facts.json")
    context = _read_json_file(path.parent / "task2_context_packet.json")
    pack = _read_json_file(path.parent / "statement_spec_pack.json")
    revenue_spec = _read_json_file(path.parent / "revenue_build_spec.json")
    if isinstance(facts, dict):
        source_payload.update(facts)
        source_payload["financial_facts"] = facts
    if isinstance(context, dict):
        source_payload["task2_context_packet"] = context
    if isinstance(revenue_spec, dict) and revenue_spec:
        source_payload["revenue_build_spec"] = revenue_spec
    if isinstance(pack, dict):
        source_payload["statement_spec_pack"] = pack
    source_historicals: list[dict[str, Any]] = []
    source_forecast_labels: list[str] = []
    source_metadata: dict[str, str] = {}
    if source_payload:
        try:
            source_historicals = _historical_records(source_payload)
            source_forecast_labels = _forecast_labels(
                source_payload,
                source_historicals[-1]["year"],
            )
            source_metadata = _model_metadata(source_payload)
        except Exception as exc:
            warnings.append(
                {
                    "sheet": "",
                    "cell": "",
                    "category": "Source Validation",
                    "issue": f"Could not load source payload for deep workbook validation: {exc}",
                }
            )
    for sheet_name in THREE_STATEMENT_TABS:
        if sheet_name not in wb.sheetnames:
            critical.append(
                {
                    "sheet": sheet_name,
                    "cell": "",
                    "category": "Missing Required Tab",
                    "issue": f"Required tab '{sheet_name}' is absent.",
                }
            )
    if critical:
        return _validation_result("FAIL", critical, warnings)

    revenue_build = wb["Revenue Build"]
    if not row_map_json:
        inferred_revenue_rows: dict[str, int] = {}
        component_metadata: list[dict[str, Any]] = []
        current_component: dict[str, Any] | None = None
        for row_idx in range(1, revenue_build.max_row + 1):
            text = str(revenue_build.cell(row=row_idx, column=1).value or "")
            if not text:
                continue
            if text == "Total Revenue":
                inferred_revenue_rows["revenue_total"] = row_idx
            elif text == "YoY Growth":
                inferred_revenue_rows["revenue_growth"] = row_idx
            elif text == "Total COGS / Cost":
                inferred_revenue_rows["total_cost"] = row_idx
            elif text == "Total Gross Profit":
                inferred_revenue_rows["total_gross_profit"] = row_idx
            elif text == "Total Gross Margin":
                inferred_revenue_rows["total_gross_margin"] = row_idx
            elif text == "Revenue Reconciliation Adjustment":
                inferred_revenue_rows["reconciliation_adjustment"] = row_idx
            elif text.endswith(" Revenue") and text != "Total Revenue":
                display = text[: -len(" Revenue")]
                current_component = {
                    "display_name": display,
                    "rows": {"revenue": row_idx, "drivers": {}},
                }
                component_metadata.append(current_component)
                key_prefix = f"component_{len(component_metadata) - 1}"
                inferred_revenue_rows.setdefault("core_revenue", row_idx)
                inferred_revenue_rows[f"{key_prefix}_revenue"] = row_idx
            elif current_component and text.startswith(f"{current_component['display_name']} "):
                suffix = text[len(str(current_component["display_name"])) + 1 :]
                key_prefix = f"component_{len(component_metadata) - 1}"
                if suffix == "COGS / Cost":
                    current_component["rows"]["cost"] = row_idx
                    inferred_revenue_rows[f"{key_prefix}_cost"] = row_idx
                elif suffix == "Gross Profit":
                    current_component["rows"]["gross_profit"] = row_idx
                    inferred_revenue_rows[f"{key_prefix}_gross_profit"] = row_idx
                elif suffix == "Gross Margin":
                    current_component["rows"]["gross_margin"] = row_idx
                    inferred_revenue_rows[f"{key_prefix}_gross_margin"] = row_idx
                elif suffix == "Revenue Mix":
                    current_component["rows"]["mix"] = row_idx
                    inferred_revenue_rows[f"{key_prefix}_mix"] = row_idx
        if inferred_revenue_rows.get("revenue_total"):
            row_map["revenue_build"] = inferred_revenue_rows
            row_map["revenue_components"] = component_metadata

    defined_names = set(wb.defined_names)
    for name in REQUIRED_MODEL_NAMES:
        if name not in defined_names:
            critical.append(
                {
                    "sheet": "Assumptions",
                    "cell": name,
                    "category": "Missing Named Range",
                    "issue": f"Named range '{name}' is required.",
                }
            )

    cover = wb["Cover"]
    cover_company = str(cover["B6"].value or "")
    cover_ticker = str(cover["B7"].value or "")
    cover_market = str(cover["B8"].value or "")
    unit_text = str(cover["A3"].value or "")
    if not cover_company or "{" in cover_company or cover_ticker in {"", "TICKER"}:
        critical.append(
            {
                "sheet": "Cover",
                "cell": "B6:B8",
                "category": "Default Metadata",
                "issue": "Workbook contains placeholder or serialized metadata instead of company/ticker/market values.",
            }
        )
    if source_metadata:
        expected = {
            "B6": source_metadata.get("company"),
            "B7": source_metadata.get("ticker"),
            "B8": source_metadata.get("market"),
        }
        actual = {"B6": cover_company, "B7": cover_ticker, "B8": cover_market}
        for cell_ref, expected_value in expected.items():
            if expected_value and expected_value not in actual[cell_ref]:
                critical.append(
                    {
                        "sheet": "Cover",
                        "cell": cell_ref,
                        "category": "Metadata Mismatch",
                        "issue": f"Expected '{expected_value}' but found '{actual[cell_ref]}'.",
                    }
                )
        expected_currency = source_metadata.get("currency")
        expected_unit = source_metadata.get("unit")
        if (expected_currency and expected_currency not in unit_text) or (
            expected_unit and expected_unit not in unit_text
        ):
            critical.append(
                {
                    "sheet": "Cover",
                    "cell": "A3",
                    "category": "Unit Mismatch",
                    "issue": f"Expected currency/unit {expected_currency}/{expected_unit}, found '{unit_text}'.",
                }
            )

    checks = wb["Checks"]
    income = wb["Income Statement"]
    forecast_cols = [
        col_idx
        for col_idx in range(3, income.max_column + 1)
        if str(income.cell(row=4, column=col_idx).value or "").lower() == "forecast"
    ]
    if not forecast_cols:
        warnings.append(
            {
                "sheet": "Income Statement",
                "cell": "4:4",
                "category": "Period Metadata",
                "issue": "No forecast columns marked; validating all modeled columns.",
            }
        )
        forecast_cols = list(range(3, income.max_column + 1))

    period_labels = {
        str(income.cell(row=5, column=col_idx).value or ""): col_idx
        for col_idx in range(3, income.max_column + 1)
    }
    expected_check_headers = [
        "Check",
        "Period",
        "Actual",
        "Expected",
        "Difference",
        "Tolerance",
        "Status",
        "Notes",
    ]
    actual_check_headers = [
        str(checks.cell(row=7, column=col_idx).value or "")
        for col_idx in range(1, len(expected_check_headers) + 1)
    ]
    if actual_check_headers != expected_check_headers:
        critical.append(
            {
                "sheet": "Checks",
                "cell": "A7:H7",
                "category": "Checks Structure",
                "issue": "Checks sheet must use Check/Period/Actual/Expected/Difference/Tolerance/Status/Notes audit columns.",
            }
        )
    if not _formula(checks["B4"].value) or "COUNTIF" not in str(checks["B4"].value).upper():
        critical.append(
            {
                "sheet": "Checks",
                "cell": "B4",
                "category": "Checks Structure",
                "issue": "Failing Check Count must aggregate Status values from the Checks audit table.",
            }
        )
    for label in source_forecast_labels:
        if label not in period_labels:
            critical.append(
                {
                    "sheet": "Checks",
                    "cell": "5:5",
                    "category": "Missing Forecast Period",
                    "issue": f"Expected forecast period {label} from statement specs.",
                }
            )
    for record in source_historicals:
        label = record["period"]
        if label not in period_labels:
            critical.append(
                {
                    "sheet": "Checks",
                    "cell": "5:5",
                    "category": "Missing Historical Period",
                    "issue": f"Expected historical period {label} from financial facts.",
                }
            )

    assumptions_ws = wb["Assumptions"]
    for row_idx in range(1, assumptions_ws.max_row + 1):
        label_value = str(assumptions_ws.cell(row=row_idx, column=1).value or "")
        if re.fullmatch(r"Assumption\s+\d+", label_value):
            critical.append(
                {
                    "sheet": "Assumptions",
                    "cell": f"A{row_idx}",
                    "category": "Placeholder Assumption",
                    "issue": "Assumption rows must have business-driver labels, not generated placeholders.",
                }
            )

    historical_revenue_values: list[float] = []
    for record in source_historicals:
        col_idx = period_labels.get(record["period"])
        if not col_idx:
            continue
        value = income.cell(row=row_map["income_statement"]["revenue_total"], column=col_idx).value
        if _formula(value):
            critical.append(
                {
                    "sheet": "Income Statement",
                    "cell": f"{_column_letter(col_idx)}{row_map['income_statement']['revenue_total']}",
                    "category": "Historical Formula",
                    "issue": "Historical revenue must be a hardcoded sourced actual, not a formula.",
                }
            )
            continue
        actual = _as_float(value, 0.0)
        expected_revenue = _historical_value(record, "revenue")
        historical_revenue_values.append(actual)
        if expected_revenue and abs(actual - expected_revenue) > max(0.01, abs(expected_revenue) * 0.001):
            critical.append(
                {
                    "sheet": "Income Statement",
                    "cell": f"{_column_letter(col_idx)}{row_map['income_statement']['revenue_total']}",
                    "category": "Historical Value Mismatch",
                    "issue": f"Revenue {actual} does not match source {expected_revenue} for {record['period']}.",
                }
            )
    if source_historicals and historical_revenue_values and all(abs(value) < 1e-9 for value in historical_revenue_values):
        critical.append(
            {
                "sheet": "Income Statement",
                "cell": "C8",
                "category": "Historical Values Missing",
                "issue": "Historical revenue values are all zero.",
            }
        )

    bs_for_debt = wb["Balance Sheet"]
    debt_ws = wb["Debt & Interest"]
    values_bs_for_debt = wb_values["Balance Sheet"]
    values_debt_ws = wb_values["Debt & Interest"]
    source_spec_lookup = _historical_input_lookup(source_payload) if source_payload else {}
    for record in source_historicals:
        col_idx = period_labels.get(record["period"])
        if not col_idx:
            continue
        col = _column_letter(col_idx)
        expected_debt = _historical_value(record, "debt")
        spec_debt = source_spec_lookup.get(_period_key(record["period"]), {}).get("total_debt")
        if expected_debt and spec_debt is not None and abs(spec_debt - expected_debt) > max(0.01, abs(expected_debt) * 0.001):
            warnings.append(
                {
                    "sheet": "Balance Sheet",
                    "cell": f"{col}{row_map['balance_sheet']['total_debt']}",
                    "category": "Statement Spec Debt Conflict",
                    "issue": f"financial_facts debt {expected_debt} conflicts with statement spec total_debt {spec_debt} for {record['period']}; workbook uses financial_facts as canonical historical source.",
                }
            )
        workbook_debt = _as_float(
            values_bs_for_debt.cell(row=row_map["balance_sheet"]["total_debt"], column=col_idx).value,
            0.0,
        )
        debt_schedule_debt = _as_float(
            values_debt_ws.cell(row=row_map["debt_interest"]["ending_debt"], column=col_idx).value,
            0.0,
        )
        if expected_debt and abs(workbook_debt - expected_debt) > max(0.01, abs(expected_debt) * 0.001):
            critical.append(
                {
                    "sheet": "Balance Sheet",
                    "cell": f"{col}{row_map['balance_sheet']['total_debt']}",
                    "category": "Debt Source Mismatch",
                    "issue": f"Workbook total debt {workbook_debt} does not match source interest-bearing debt {expected_debt} for {record['period']}.",
                }
            )
        if expected_debt and abs(debt_schedule_debt - expected_debt) > max(0.01, abs(expected_debt) * 0.001):
            critical.append(
                {
                    "sheet": "Debt & Interest",
                    "cell": f"{col}{row_map['debt_interest']['ending_debt']}",
                    "category": "Debt Schedule Mismatch",
                    "issue": f"Debt schedule ending debt {debt_schedule_debt} does not match source interest-bearing debt {expected_debt} for {record['period']}.",
                }
            )
        short_term_raw = _historical_value(record, "short_term_debt_raw")
        if short_term_raw > max(expected_debt * 5, 0.01):
            warnings.append(
                {
                    "sheet": "Debt & Interest",
                    "cell": f"{col}{row_map['debt_interest']['short_term_debt_raw']}",
                    "category": "Debt Data Quality",
                    "issue": (
                        f"Short-term debt raw {short_term_raw} materially exceeds modeled "
                        f"interest-bearing debt {expected_debt} for {record['period']}; "
                        "preserve raw field and verify MCP metric/unit before using it as Total Debt."
                    ),
                }
            )
        expected_net_finance = _historical_value(record, "net_finance_expense")
        actual_net_finance = _as_float(
            income.cell(row=row_map["income_statement"]["net_finance_expense"], column=col_idx).value,
            0.0,
        )
        if expected_net_finance and abs(actual_net_finance - expected_net_finance) > max(0.01, abs(expected_net_finance) * 0.001):
            critical.append(
                {
                    "sheet": "Income Statement",
                    "cell": f"{col}{row_map['income_statement']['net_finance_expense']}",
                    "category": "Net Finance Expense Mismatch",
                    "issue": f"Workbook net finance expense {actual_net_finance} does not match source {expected_net_finance} for {record['period']}.",
                }
            )

    revenue_build = wb["Revenue Build"]
    workbook_components = row_map.get("revenue_components")
    if not isinstance(workbook_components, list):
        workbook_components = []
    if source_payload:
        allowed_component_labels = {
            str(component["display_name"])
            for component in _normalize_revenue_components(source_payload)
        }
        workbook_component_labels = {
            str(component.get("display_name"))
            for component in workbook_components
            if isinstance(component, dict) and component.get("display_name")
        }
        leaked_labels = sorted(workbook_component_labels - allowed_component_labels)
        if leaked_labels:
            critical.append(
                {
                    "sheet": "Revenue Build",
                    "cell": "A:A",
                    "category": "Revenue Component Label Leakage",
                    "issue": (
                        "Revenue Build contains component labels not declared in "
                        f"revenue_build_spec.json: {', '.join(leaked_labels)}."
                    ),
                }
            )
    forecast_records = _forecast_revenue_records(source_payload) if source_payload else []
    for record in forecast_records:
        label = _normalize_period_label(record.get("period"), is_actual=False)
        col_idx = period_labels.get(label)
        if not col_idx:
            continue
        segment_sum = 0.0
        for component in workbook_components:
            if not isinstance(component, dict):
                continue
            rows = component.get("rows")
            if not isinstance(rows, dict):
                continue
            revenue_row = rows.get("revenue")
            if not isinstance(revenue_row, int):
                continue
            segment_sum += _as_float(
                wb_values["Revenue Build"].cell(
                    row=revenue_row,
                    column=col_idx,
                ).value,
                0.0,
            )
        expected_total = _as_float(record.get("total"), 0.0)
        total_formula = revenue_build.cell(
            row=row_map["revenue_build"]["revenue_total"],
            column=col_idx,
        ).value
        if expected_total and abs(segment_sum - expected_total) > max(0.01, expected_total * 0.001):
            critical.append(
                {
                    "sheet": "Revenue Build",
                    "cell": f"{_column_letter(col_idx)}:{_column_letter(col_idx)}",
                    "category": "Forecast Revenue Mismatch",
                    "issue": f"Segment revenue inputs sum to {segment_sum}, expected {expected_total} for {label}.",
                }
            )
        if not (_formula(total_formula) and "SUM" in total_formula.upper()):
            critical.append(
                {
                    "sheet": "Revenue Build",
                    "cell": f"{_column_letter(col_idx)}{row_map['revenue_build']['revenue_total']}",
                    "category": "Revenue Build Link",
                    "issue": "Forecast total revenue must sum formula-driven segment rows.",
                }
            )

    formula_requirements: list[tuple[str, list[int]]] = [
        ("Revenue Build", list(row_map["revenue_build"].values())),
        ("Income Statement", list(row_map["income_statement"].values())),
        ("Balance Sheet", list(row_map["balance_sheet"].values())),
        ("Cash Flow Statement", list(row_map["cash_flow"].values())),
        ("Working Capital", list(row_map["working_capital"].values())),
        ("PP&E & D&A", list(row_map["ppe_da"].values())),
        ("Debt & Interest", list(row_map["debt_interest"].values())),
        ("Share Count", list(row_map["share_count"].values())),
        ("DCF Inputs", list(row_map["dcf_inputs"].values())),
    ]
    for sheet_name, rows in formula_requirements:
        ws = wb[sheet_name]
        for row in sorted(set(rows)):
            for col_idx in forecast_cols:
                value = ws.cell(row=row, column=col_idx).value
                if not _formula(value):
                    critical.append(
                        {
                            "sheet": sheet_name,
                            "cell": f"{_column_letter(col_idx)}{row}",
                            "category": "Projection Hardcode",
                            "issue": "Forecast/model cell must be an Excel formula.",
                        }
                    )

    cache_missing: list[str] = []
    model_cols = list(range(3, max(period_labels.values() or [checks.max_column]) + 1))
    for sheet_name, rows in formula_requirements:
        ws = wb[sheet_name]
        values_ws = wb_values[sheet_name]
        for row in sorted(set(rows)):
            for col_idx in model_cols:
                if _formula(ws.cell(row=row, column=col_idx).value) and values_ws.cell(
                    row=row,
                    column=col_idx,
                ).value is None:
                    cache_missing.append(f"{sheet_name}!{_column_letter(col_idx)}{row}")
    if _formula(checks["B4"].value) and wb_values["Checks"]["B4"].value is None:
        cache_missing.append("Checks!B4")
    if cache_missing:
        critical.append(
            {
                "sheet": "",
                "cell": "",
                "category": "Formula Cache Missing",
                "issue": (
                    f"{len(cache_missing)} formula cells have no cached result; "
                    "the workbook can appear blank in data_only readers or previews. "
                    f"Examples: {', '.join(cache_missing[:8])}."
                ),
            }
        )

    try:
        from openpyxl.formula.tokenizer import Tokenizer
    except Exception:
        Tokenizer = None  # type: ignore[assignment]
    if Tokenizer is not None:
        blank_references: list[str] = []
        scanned_sheets = (
            "DCF Inputs",
            "Checks",
            "Debt & Interest",
            "Income Statement",
            "Balance Sheet",
            "Cash Flow Statement",
        )
        for sheet_name in scanned_sheets:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    formula = cell.value
                    if not _formula(formula):
                        continue
                    try:
                        tokens = Tokenizer(formula).items
                    except Exception:
                        continue
                    for token in tokens:
                        if token.subtype != "RANGE":
                            continue
                        ref_text = str(token.value)
                        if ":" in ref_text:
                            continue
                        if "!" in ref_text:
                            ref_sheet, ref_cell = ref_text.split("!", 1)
                            ref_sheet = ref_sheet.strip("'")
                        else:
                            ref_sheet, ref_cell = sheet_name, ref_text
                        ref_cell = ref_cell.replace("$", "")
                        if ref_sheet not in wb.sheetnames:
                            continue
                        try:
                            source_value = wb[ref_sheet][ref_cell].value
                        except Exception:
                            continue
                        if source_value is None:
                            blank_references.append(
                                f"{sheet_name}!{cell.coordinate}->{ref_sheet}!{ref_cell}"
                            )
        if blank_references:
            critical.append(
                {
                    "sheet": "",
                    "cell": "",
                    "category": "Blank Formula Reference",
                    "issue": (
                        f"{len(blank_references)} formulas reference blank source cells. "
                        f"Examples: {', '.join(blank_references[:8])}."
                    ),
                }
            )

    bs = wb["Balance Sheet"]
    cf = wb["Cash Flow Statement"]
    for col_idx in forecast_cols:
        col = _column_letter(col_idx)
        bs_cash = bs.cell(row=row_map["balance_sheet"]["cash_and_equivalents"], column=col_idx).value
        cf_cash = cf.cell(row=row_map["cash_flow"]["ending_cash"], column=col_idx).value
        if not (
            _formula(bs_cash)
            and "Cash Flow Statement" in bs_cash
            and f"{col}{row_map['cash_flow']['ending_cash']}" in bs_cash
        ):
            critical.append(
                {
                    "sheet": "Balance Sheet",
                    "cell": f"{col}{row_map['balance_sheet']['cash_and_equivalents']}",
                    "category": "Cash Tie-Out",
                    "issue": "Forecast BS cash must link to CF ending cash.",
                }
            )
        if not _formula(cf_cash):
            critical.append(
                {
                    "sheet": "Cash Flow Statement",
                    "cell": f"{col}{row_map['cash_flow']['ending_cash']}",
                    "category": "Cash Tie-Out",
                    "issue": "Forecast CF ending cash must be formula-driven.",
                }
            )

    dcf = wb["DCF Inputs"]
    for row in row_map["dcf_inputs"].values():
        for col_idx in forecast_cols:
            if not _formula(dcf.cell(row=row, column=col_idx).value):
                critical.append(
                    {
                        "sheet": "DCF Inputs",
                        "cell": f"{_column_letter(col_idx)}{row}",
                        "category": "DCF Input Hardcode",
                        "issue": "DCF Inputs must pull from linked model cells.",
                    }
                )

    return _validation_result("PASS" if not critical else "FAIL", critical, warnings)


@tool
def write_coverage_state(
    state_json: str,
    ticker: str,
    market: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Overwrite coverage_state.json for the stock after validating JSON."""
    failure = _inline_json_failure(state_json, "state_json")
    if failure:
        return _json_result(failure)
    state: Any = _json_loads(state_json, "state_json")
    if not isinstance(state, dict):
        raise ValueError("state_json must be a JSON object")

    coverage_dir = _coverage_dir(market, ticker, output_dir)
    _ensure_dir(coverage_dir)
    state.setdefault("ticker", ticker)
    state.setdefault("market", market)
    state["updated_at"] = datetime.now().isoformat(timespec="seconds")
    path = coverage_dir / "coverage_state.json"
    _write_text(
        path,
        json.dumps(state, ensure_ascii=False, indent=2) + "\n",
    )
    return _relative_to_workspace(path)
